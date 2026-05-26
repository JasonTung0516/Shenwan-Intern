from __future__ import annotations

from pathlib import Path
from typing import Iterable, Sequence
import math
import re
import warnings

import numpy as np
import pandas as pd

try:
    from .config import *
except ImportError:  # allow running a file directly from this folder
    from config import *

try:
    from .core_base import *
    from .core_neutralization import *
    from .portfolio_five_group import *
    from .portfolio_industry_neutral import *
except ImportError:
    from core_base import *
    from core_neutralization import *
    from portfolio_five_group import *
    from portfolio_industry_neutral import *

# This chunk only compares the two completed five-group tests.
# It does NOT rerun factor calculation, IC, or portfolio backtest.

from pathlib import Path
import numpy as np
import pandas as pd


# =============================================================================
# CONFIG
# =============================================================================

COMPARE_OUTDIR = OUTDIR / "q5_minus_q1_comparison"
COMPARE_OUTDIR.mkdir(parents=True, exist_ok=True)

Q5_COMPARE_EXCEL = COMPARE_OUTDIR / "q5_minus_q1_performance_comparison.xlsx"

# Existing output files, used only if result dictionaries are not in memory.
ORIGINAL_FIVE_GROUP_EXCEL = (
    OUTDIR / "five_group_test" / "five_group_portfolio_test.xlsx"
)

INDUSTRY_NEUTRAL_FIVE_GROUP_EXCEL = (
    OUTDIR / "five_group_test_industry_neutral" / "five_group_portfolio_test_industry_neutral.xlsx"
)


# =============================================================================
# Helper functions
# =============================================================================

def calc_q5q1_max_drawdown(nav: pd.Series) -> float:
    nav = pd.Series(nav).dropna()

    if nav.empty:
        return np.nan

    running_max = nav.cummax()
    drawdown = nav / running_max - 1

    return float(drawdown.min())


def calc_q5q1_perf_from_daily_returns(
    daily_returns: pd.DataFrame,
    method_name: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Recalculate Q5_minus_Q1 performance from Daily_Group_Returns.

    Q5_minus_Q1 daily return:
        ret_Q5_minus_Q1[t] = ret_Q5[t] - ret_Q1[t]

    Metrics:
        total_return
        annual_return
        annual_vol
        sharpe_no_rf
        max_drawdown
        avg_daily_return
        daily_win_rate
    """
    daily = daily_returns.copy()
    daily["trade_day"] = pd.to_datetime(daily["trade_day"])

    summary_records = []
    nav_records = []

    for h, sub in daily.groupby("horizon"):
        wide = (
            sub.pivot_table(
                index="trade_day",
                columns="group",
                values="ret",
                aggfunc="mean",
            )
            .sort_index()
        )

        if "Q5" not in wide.columns or "Q1" not in wide.columns:
            continue

        q5q1_ret = wide["Q5"] - wide["Q1"]
        q5q1_ret = q5q1_ret.replace([np.inf, -np.inf], np.nan).dropna()

        if q5q1_ret.empty:
            continue

        nav = (1.0 + q5q1_ret).cumprod()
        n_days = len(q5q1_ret)

        total_return = float(nav.iloc[-1] - 1.0)
        annual_return = float(nav.iloc[-1] ** (252.0 / n_days) - 1.0)
        annual_vol = float(q5q1_ret.std(ddof=1) * np.sqrt(252.0))
        sharpe_no_rf = annual_return / annual_vol if annual_vol > 0 else np.nan
        max_drawdown = calc_q5q1_max_drawdown(nav)

        summary_records.append(
            {
                "method": method_name,
                "horizon": int(h),
                "portfolio": "Q5_minus_Q1",
                "n_days": n_days,
                "total_return": total_return,
                "annual_return": annual_return,
                "annual_vol": annual_vol,
                "sharpe_no_rf": sharpe_no_rf,
                "max_drawdown": max_drawdown,
                "avg_daily_return": float(q5q1_ret.mean()),
                "daily_win_rate": float((q5q1_ret > 0).mean()),
            }
        )

        nav_records.append(
            pd.DataFrame(
                {
                    "method": method_name,
                    "horizon": int(h),
                    "trade_day": q5q1_ret.index,
                    "q5_minus_q1_ret": q5q1_ret.values,
                    "q5_minus_q1_nav": nav.values,
                }
            )
        )

    summary_df = pd.DataFrame(summary_records)

    if nav_records:
        nav_df = pd.concat(nav_records, ignore_index=True)
    else:
        nav_df = pd.DataFrame()

    return summary_df, nav_df


def load_daily_returns_from_memory_or_excel(
    result_name: str,
    excel_path: Path,
) -> pd.DataFrame:
    """
    Priority:
    1. Use existing result dictionary in memory.
    2. Otherwise read Daily_Group_Returns from existing Excel.
    """
    if (
        result_name in globals()
        and isinstance(globals()[result_name], dict)
        and "daily_returns" in globals()[result_name]
    ):
        return globals()[result_name]["daily_returns"].copy()

    if not excel_path.exists():
        raise FileNotFoundError(f"Cannot find Excel file: {excel_path}")

    return pd.read_excel(excel_path, sheet_name="Daily_Group_Returns")


# =============================================================================
# 1. Load daily group returns
# =============================================================================

original_daily_returns = load_daily_returns_from_memory_or_excel(
    result_name="five_group_result",
    excel_path=ORIGINAL_FIVE_GROUP_EXCEL,
)

industry_neutral_daily_returns = load_daily_returns_from_memory_or_excel(
    result_name="industry_neutral_five_group_result",
    excel_path=INDUSTRY_NEUTRAL_FIVE_GROUP_EXCEL,
)


# =============================================================================
# 2. Recalculate Q5_minus_Q1 performance
# =============================================================================

original_q5_summary, original_q5_nav = calc_q5q1_perf_from_daily_returns(
    daily_returns=original_daily_returns,
    method_name="Original equal-weight",
)

industry_neutral_q5_summary, industry_neutral_q5_nav = calc_q5q1_perf_from_daily_returns(
    daily_returns=industry_neutral_daily_returns,
    method_name="Industry-neutral weight",
)

raw_q5_summary = pd.concat(
    [original_q5_summary, industry_neutral_q5_summary],
    ignore_index=True,
).sort_values(["horizon", "method"]).reset_index(drop=True)

q5_nav = pd.concat(
    [original_q5_nav, industry_neutral_q5_nav],
    ignore_index=True,
)


# =============================================================================
# 3. Build wide comparison table
# =============================================================================

comparison = original_q5_summary.merge(
    industry_neutral_q5_summary,
    on="horizon",
    suffixes=("_orig", "_ind"),
)

comparison_wide = pd.DataFrame(
    {
        "horizon": comparison["horizon"],

        "orig_n_days": comparison["n_days_orig"],
        "ind_neutral_n_days": comparison["n_days_ind"],

        "orig_total_return": comparison["total_return_orig"],
        "ind_total_return": comparison["total_return_ind"],
        "diff_total_return": comparison["total_return_ind"] - comparison["total_return_orig"],

        "orig_annual_return": comparison["annual_return_orig"],
        "ind_annual_return": comparison["annual_return_ind"],
        "diff_annual_return": comparison["annual_return_ind"] - comparison["annual_return_orig"],

        "orig_annual_vol": comparison["annual_vol_orig"],
        "ind_annual_vol": comparison["annual_vol_ind"],
        "diff_annual_vol": comparison["annual_vol_ind"] - comparison["annual_vol_orig"],

        "orig_sharpe_no_rf": comparison["sharpe_no_rf_orig"],
        "ind_sharpe_no_rf": comparison["sharpe_no_rf_ind"],
        "diff_sharpe_no_rf": comparison["sharpe_no_rf_ind"] - comparison["sharpe_no_rf_orig"],

        "orig_max_drawdown": comparison["max_drawdown_orig"],
        "ind_max_drawdown": comparison["max_drawdown_ind"],
        "diff_max_drawdown": comparison["max_drawdown_ind"] - comparison["max_drawdown_orig"],

        "orig_avg_daily_return": comparison["avg_daily_return_orig"],
        "ind_avg_daily_return": comparison["avg_daily_return_ind"],
        "diff_avg_daily_return": comparison["avg_daily_return_ind"] - comparison["avg_daily_return_orig"],

        "orig_daily_win_rate": comparison["daily_win_rate_orig"],
        "ind_daily_win_rate": comparison["daily_win_rate_ind"],
        "diff_daily_win_rate": comparison["daily_win_rate_ind"] - comparison["daily_win_rate_orig"],
    }
)

comparison_wide = comparison_wide.sort_values("horizon").reset_index(drop=True)


# =============================================================================
# 4. Methodology sheet
# =============================================================================

methodology = pd.DataFrame(
    [
        {
            "Item": "Original equal-weight",
            "Description": "The original five-group test. Each quintile is equal-weighted across all stocks.",
        },
        {
            "Item": "Industry-neutral weight",
            "Description": "The industry-neutral five-group test. Within each industry stocks are equal-weighted; across industries the same target industry weights are applied to each quintile.",
        },
        {
            "Item": "Q5_minus_Q1",
            "Description": "Long Q5 high-signal group and short Q1 low-signal group.",
        },
        {
            "Item": "annual_return",
            "Description": "Annualized return from the Q5_minus_Q1 daily return series.",
        },
        {
            "Item": "annual_vol",
            "Description": "Annualized volatility using daily returns and sqrt(252).",
        },
        {
            "Item": "sharpe_no_rf",
            "Description": "Annual return divided by annual volatility. No risk-free-rate adjustment.",
        },
        {
            "Item": "max_drawdown",
            "Description": "Maximum drawdown of the Q5_minus_Q1 cumulative NAV curve.",
        },
        {
            "Item": "daily_win_rate",
            "Description": "Share of days where Q5_minus_Q1 daily return is positive.",
        },
    ]
)


# =============================================================================
# 5. Export Excel
# =============================================================================

with pd.ExcelWriter(Q5_COMPARE_EXCEL, engine="openpyxl") as writer:
    comparison_wide.to_excel(writer, sheet_name="Q5_Q1_Comparison", index=False)
    raw_q5_summary.to_excel(writer, sheet_name="Raw_Q5_Q1_Summary", index=False)
    q5_nav.to_excel(writer, sheet_name="Q5_Q1_Daily_NAV", index=False)
    methodology.to_excel(writer, sheet_name="Methodology", index=False)


# =============================================================================
# 6. Format Excel
# =============================================================================

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = load_workbook(Q5_COMPARE_EXCEL)

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)

percent_cols = {
    "orig_total_return",
    "ind_total_return",
    "diff_total_return",
    "orig_annual_return",
    "ind_annual_return",
    "diff_annual_return",
    "orig_annual_vol",
    "ind_annual_vol",
    "diff_annual_vol",
    "orig_max_drawdown",
    "ind_max_drawdown",
    "diff_max_drawdown",
    "orig_avg_daily_return",
    "ind_avg_daily_return",
    "diff_avg_daily_return",
    "orig_daily_win_rate",
    "ind_daily_win_rate",
    "diff_daily_win_rate",

    "total_return",
    "annual_return",
    "annual_vol",
    "max_drawdown",
    "avg_daily_return",
    "daily_win_rate",

    "q5_minus_q1_ret",
}

nav_cols = {
    "q5_minus_q1_nav",
}

numeric_4_cols = {
    "orig_sharpe_no_rf",
    "ind_sharpe_no_rf",
    "diff_sharpe_no_rf",
    "sharpe_no_rf",
}

integer_cols = {
    "horizon",
    "orig_n_days",
    "ind_neutral_n_days",
    "n_days",
}

date_cols = {
    "trade_day",
}

for ws in wb.worksheets:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    header_to_col = {
        str(ws.cell(row=1, column=col_idx).value): col_idx
        for col_idx in range(1, ws.max_column + 1)
    }

    for header, col_idx in header_to_col.items():
        if header in percent_cols:
            fmt = "0.00%"
        elif header in nav_cols:
            fmt = "0.0000"
        elif header in numeric_4_cols:
            fmt = "0.0000"
        elif header in integer_cols:
            fmt = "0"
        elif header in date_cols:
            fmt = "yyyy-mm-dd"
        else:
            fmt = None

        if fmt is not None:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = fmt

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0

        scan_rows = min(ws.max_row, 3000)

        for row_idx in range(1, scan_rows + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))

        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 32)

wb.save(Q5_COMPARE_EXCEL)

print(f"Saved Q5_minus_Q1 comparison Excel: {Q5_COMPARE_EXCEL}")
print("\nQ5_minus_Q1 comparison:")
print(comparison_wide.to_string(index=False))
