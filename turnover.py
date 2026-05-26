from __future__ import annotations

from pathlib import Path
from typing import Iterable, Sequence
import math
import re
import warnings

import numpy as np
import pandas as pd

try:
    from .config import *
except ImportError:  # allow running a file directly from this folder
    from config import *

try:
    from .core_base import *
    from .core_neutralization import *
    from .portfolio_five_group import *
    from .portfolio_industry_neutral import *
except ImportError:
    from core_base import *
    from core_neutralization import *
    from portfolio_five_group import *
    from portfolio_industry_neutral import *

# Teacher requirement:
# 统计之前五组分档中 Q1-Q5 每组的换手率。
#
# Main outputs:
# 1. 每组每个换仓周期的平均换手率
# 2. 每次换仓的明细换手率
# 3. 成分股层面的进入 / 退出 / 保留比例
# 4. Q2-Q5 相对 Q1 的换手率差异
#
# Notes:
# - Q1 = lowest signal
# - Q5 = highest signal
# - Main summary excludes the initial opening rebalance, because first opening from cash
#   will mechanically create 100% buy turnover and distort normal turnover.
# - It reports both:
#   1) weight turnover: based on portfolio weights
#   2) member turnover: based on stock membership changes

from pathlib import Path
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt


# =============================================================================
# CONFIG
# =============================================================================

TURNOVER_OUTDIR = OUTDIR / "five_group_turnover"
TURNOVER_OUTDIR.mkdir(parents=True, exist_ok=True)

TURNOVER_EXCEL = TURNOVER_OUTDIR / "five_group_turnover_statistics.xlsx"

TURNOVER_REBALANCE_PERIODS = GROUP_REBALANCE_PERIODS if "GROUP_REBALANCE_PERIODS" in globals() else [1, 5, 10, 20]
TURNOVER_N_GROUPS = N_GROUPS if "N_GROUPS" in globals() else 5
TURNOVER_MIN_STOCKS_TO_GROUP = MIN_STOCKS_TO_GROUP if "MIN_STOCKS_TO_GROUP" in globals() else 50

# Use the same signal mode as your five-group test.
# "current_backtest" follows SIGNAL_MODE.
# You can also set directly: "weighted" or "equal".
TURNOVER_SIGNAL_MODE = "current_backtest"

TURNOVER_SIGNAL_EQUAL_COL = globals().get("GROUP_SIGNAL_EQUAL_COL", "group_signal_equal_mad_z")
TURNOVER_SIGNAL_WEIGHTED_COL = globals().get("GROUP_SIGNAL_WEIGHTED_COL", "group_signal_weighted_mad_z")


# =============================================================================
# 1. Resolve signal panel and signal column
# =============================================================================

def get_signal_panel_for_turnover() -> pd.DataFrame:
    """
    Priority:
    1. Use five_group_result["signal_panel"] from the previous five-group test.
    2. Otherwise rebuild signal panel using build_group_test_signal_panel().
    """
    if (
        "five_group_result" in globals()
        and isinstance(five_group_result, dict)
        and "signal_panel" in five_group_result
    ):
        print("Using five_group_result['signal_panel'] from memory.")
        panel = five_group_result["signal_panel"].copy()
    else:
        if "build_group_test_signal_panel" not in globals():
            raise NameError(
                "Cannot find five_group_result or build_group_test_signal_panel(). "
                "Please run the five-group test chunk first."
            )

        print("five_group_result not found. Rebuilding signal panel ...")
        panel, _ = build_group_test_signal_panel()

    panel["trade_day"] = pd.to_datetime(panel["trade_day"]).dt.normalize()
    panel["code"] = panel["code"].astype(str)

    return panel.sort_values(["code", "trade_day"]).reset_index(drop=True)


def resolve_turnover_signal_col(signal_panel: pd.DataFrame) -> tuple[str, str]:
    """
    Return:
        signal_col, signal_mode_used
    """
    if TURNOVER_SIGNAL_MODE == "current_backtest":
        mode = globals().get("SIGNAL_MODE", "weighted")
    else:
        mode = TURNOVER_SIGNAL_MODE

    if mode == "weighted":
        signal_col = TURNOVER_SIGNAL_WEIGHTED_COL
    elif mode == "equal":
        signal_col = TURNOVER_SIGNAL_EQUAL_COL
    else:
        raise ValueError("TURNOVER_SIGNAL_MODE must be 'current_backtest', 'weighted', or 'equal'.")

    if signal_col not in signal_panel.columns:
        raise ValueError(
            f"Signal column {signal_col} not found in signal_panel. "
            f"Available columns are: {list(signal_panel.columns)}"
        )

    return signal_col, mode


# =============================================================================
# 2. Local helper functions
# =============================================================================

def _turnover_clean_weight_series(w: pd.Series) -> pd.Series:
    """
    Clean portfolio weight Series.

    Index = stock code
    Value = portfolio weight
    """
    if w is None or len(w) == 0:
        return pd.Series(dtype=float)

    out = pd.Series(w, dtype=float).copy()
    out.index = out.index.astype(str)

    out = out.replace([np.inf, -np.inf], np.nan).dropna()
    out = out[out.abs() > 1e-15]

    return out


def _turnover_normalize_long_only(w: pd.Series) -> pd.Series:
    """
    Normalize long-only weights to sum to 1.
    """
    w = _turnover_clean_weight_series(w)
    w = w[w > 0]

    s = w.sum()

    if not np.isfinite(s) or s <= 0:
        return pd.Series(dtype=float)

    return w / s


def _turnover_calc_weight_turnover(
    old_weights: pd.Series,
    new_weights: pd.Series,
) -> dict:
    """
    Calculate weight turnover from old portfolio to new target portfolio.

    buy_turnover:
        sum(max(new_weight - old_weight, 0))

    sell_turnover:
        sum(max(old_weight - new_weight, 0))

    turnover_two_way:
        buy_turnover + sell_turnover

    turnover_one_way:
        0.5 * turnover_two_way
        For normal fully invested long-only rebalances, buy ~= sell,
        so this is the usual one-way turnover.
    """
    old_w = _turnover_clean_weight_series(old_weights)
    new_w = _turnover_clean_weight_series(new_weights)

    all_codes = old_w.index.union(new_w.index)

    old_aligned = old_w.reindex(all_codes).fillna(0.0)
    new_aligned = new_w.reindex(all_codes).fillna(0.0)

    delta = new_aligned - old_aligned

    buy_turnover = float(delta.clip(lower=0.0).sum())
    sell_turnover = float((-delta.clip(upper=0.0)).sum())
    turnover_two_way = buy_turnover + sell_turnover
    turnover_one_way = 0.5 * turnover_two_way

    return {
        "buy_turnover": buy_turnover,
        "sell_turnover": sell_turnover,
        "turnover_two_way": turnover_two_way,
        "turnover_one_way": turnover_one_way,
    }


def _turnover_calc_member_change(
    old_weights: pd.Series,
    new_weights: pd.Series,
) -> dict:
    """
    Calculate membership change between two rebalance target portfolios.

    This ignores weight drift and only checks which stocks enter / exit the group.
    """
    old_w = _turnover_clean_weight_series(old_weights)
    new_w = _turnover_clean_weight_series(new_weights)

    old_set = set(old_w.index.astype(str))
    new_set = set(new_w.index.astype(str))

    old_n = len(old_set)
    new_n = len(new_set)

    if old_n == 0 and new_n == 0:
        return {
            "old_n": 0,
            "new_n": 0,
            "intersection_n": 0,
            "union_n": 0,
            "enter_n": 0,
            "exit_n": 0,
            "retention_ratio_old": np.nan,
            "enter_ratio_new": np.nan,
            "exit_ratio_old": np.nan,
            "member_turnover_ratio": np.nan,
            "jaccard_distance": np.nan,
        }

    intersection_n = len(old_set.intersection(new_set))
    union_n = len(old_set.union(new_set))
    enter_n = len(new_set - old_set)
    exit_n = len(old_set - new_set)

    retention_ratio_old = intersection_n / old_n if old_n > 0 else np.nan
    enter_ratio_new = enter_n / new_n if new_n > 0 else np.nan
    exit_ratio_old = exit_n / old_n if old_n > 0 else np.nan

    if old_n > 0 and new_n > 0:
        member_turnover_ratio = 0.5 * (enter_ratio_new + exit_ratio_old)
    else:
        member_turnover_ratio = np.nan

    jaccard_distance = 1.0 - intersection_n / union_n if union_n > 0 else np.nan

    return {
        "old_n": old_n,
        "new_n": new_n,
        "intersection_n": intersection_n,
        "union_n": union_n,
        "enter_n": enter_n,
        "exit_n": exit_n,
        "retention_ratio_old": retention_ratio_old,
        "enter_ratio_new": enter_ratio_new,
        "exit_ratio_old": exit_ratio_old,
        "member_turnover_ratio": member_turnover_ratio,
        "jaccard_distance": jaccard_distance,
    }


def _turnover_drift_weights_one_day(
    weights: pd.Series,
    ret_series: pd.Series,
) -> pd.Series:
    """
    Drift current weights by one-day stock returns.

    Missing stock returns are treated as 0 for weight-drift purpose.
    This preserves the existing position instead of dropping it.
    """
    w = _turnover_clean_weight_series(weights)

    if w.empty:
        return w

    r = pd.to_numeric(ret_series.reindex(w.index), errors="coerce")
    r = r.replace([np.inf, -np.inf], np.nan).fillna(0.0)

    port_ret = float((w * r).sum())

    if not np.isfinite(port_ret) or not np.isfinite(1.0 + port_ret) or (1.0 + port_ret) <= 0:
        return w

    drifted = w * (1.0 + r) / (1.0 + port_ret)

    return _turnover_normalize_long_only(drifted)


def _turnover_assign_groups_for_day(
    signal_day_df: pd.DataFrame,
    signal_col: str,
) -> pd.DataFrame:
    """
    Assign Q1-Q5 based on signal.

    Use your existing assign_quintile_groups_for_day() if available.
    Otherwise use the same fallback logic.
    """
    if "assign_quintile_groups_for_day" in globals():
        out = assign_quintile_groups_for_day(
            signal_day_df=signal_day_df,
            signal_col=signal_col,
            n_groups=TURNOVER_N_GROUPS,
            min_stocks=TURNOVER_MIN_STOCKS_TO_GROUP,
        )
        return out.copy()

    x = signal_day_df.copy()

    if not INCLUDE_SUSPENDED and "volume" in x.columns:
        x = x[x["volume"].fillna(0) > 0].copy()

    x = x.dropna(subset=[signal_col]).copy()

    if len(x) < TURNOVER_MIN_STOCKS_TO_GROUP:
        return pd.DataFrame(columns=["code", "group", signal_col])

    ranks = x[signal_col].rank(method="first")
    labels = [f"Q{i}" for i in range(1, TURNOVER_N_GROUPS + 1)]

    x["group"] = pd.qcut(
        ranks,
        q=TURNOVER_N_GROUPS,
        labels=labels,
    ).astype(str)

    return x[["code", "group", signal_col]].copy()


def _turnover_build_group_target_weights(
    group_df: pd.DataFrame,
) -> dict[str, pd.Series]:
    """
    Build equal-weight target portfolio for each Q group.
    """
    targets = {}

    for group in [f"Q{i}" for i in range(1, TURNOVER_N_GROUPS + 1)]:
        codes = (
            group_df.loc[group_df["group"] == group, "code"]
            .dropna()
            .astype(str)
            .drop_duplicates()
            .tolist()
        )

        if len(codes) == 0:
            targets[group] = pd.Series(dtype=float)
        else:
            targets[group] = pd.Series(1.0 / len(codes), index=codes, dtype=float)

    return targets


# =============================================================================
# 3. Main turnover calculation
# =============================================================================

def calculate_five_group_turnover() -> tuple[pd.DataFrame, pd.DataFrame, str, str]:
    """
    Calculate Q1-Q5 turnover for all rebalance periods.

    For each horizon h:
        - Use same rebalance schedule as previous five-group backtest.
        - Compute target Q1-Q5 equal-weight portfolios on each rebalance date.
        - Compare new target weights with previous drifted weights.
        - Also compare new members with previous target members.
    """
    panel = get_signal_panel_for_turnover()
    signal_col, signal_mode_used = resolve_turnover_signal_col(panel)

    print(f"\nTurnover signal mode: {signal_mode_used}")
    print(f"Turnover signal column: {signal_col}")

    if "vwap_price" not in panel.columns:
        raise ValueError("signal_panel must contain vwap_price to calculate weight drift.")

    panel = panel.sort_values(["code", "trade_day"]).reset_index(drop=True)

    panel["ret_vwap_1d"] = panel.groupby("code", sort=False)["vwap_price"].pct_change()
    panel["ret_vwap_1d"] = panel["ret_vwap_1d"].replace([np.inf, -np.inf], np.nan)

    trading_days = sorted(panel["trade_day"].dropna().unique())

    signal_by_day = {
        day: sub[["code", "trade_day", "volume", signal_col]].copy()
        for day, sub in panel[["code", "trade_day", "volume", signal_col]].groupby("trade_day")
    }

    ret_by_day = {
        day: sub.set_index("code")["ret_vwap_1d"]
        for day, sub in panel[["code", "trade_day", "ret_vwap_1d"]].groupby("trade_day")
    }

    records = []

    required_groups = [f"Q{i}" for i in range(1, TURNOVER_N_GROUPS + 1)]

    for h in TURNOVER_REBALANCE_PERIODS:
        print(f"\nCalculating group turnover, rebalance_period={h}d ...")

        current_weights = {g: pd.Series(dtype=float) for g in required_groups}
        previous_target_weights = {g: pd.Series(dtype=float) for g in required_groups}

        rebalance_id = 0
        valid_start_end = len(trading_days) - h

        for start_i in range(0, valid_start_end, h):
            signal_day = trading_days[start_i]

            if start_i + h + 1 >= len(trading_days):
                break

            entry_day = trading_days[start_i + 1]
            exit_day = trading_days[start_i + h + 1]

            signal_day_df = signal_by_day.get(signal_day)

            if signal_day_df is None or signal_day_df.empty:
                continue

            group_df = _turnover_assign_groups_for_day(
                signal_day_df=signal_day_df,
                signal_col=signal_col,
            )

            if group_df.empty:
                continue

            if not set(required_groups).issubset(set(group_df["group"].unique())):
                continue

            target_weights = _turnover_build_group_target_weights(group_df)

            rebalance_id += 1

            for group in required_groups:
                old_drifted_w = current_weights[group]
                old_target_w = previous_target_weights[group]
                new_target_w = target_weights[group]

                weight_turnover = _turnover_calc_weight_turnover(
                    old_weights=old_drifted_w,
                    new_weights=new_target_w,
                )

                member_change = _turnover_calc_member_change(
                    old_weights=old_target_w,
                    new_weights=new_target_w,
                )

                is_initial_rebalance = old_target_w.empty

                group_signal = group_df[group_df["group"] == group][signal_col]

                records.append(
                    {
                        "horizon": int(h),
                        "rebalance_id": rebalance_id,
                        "signal_day": signal_day,
                        "entry_day": entry_day,
                        "exit_day": exit_day,
                        "group": group,
                        "is_initial_rebalance": is_initial_rebalance,

                        "n_stocks": int(new_target_w.shape[0]),
                        "avg_signal": float(group_signal.mean()),
                        "median_signal": float(group_signal.median()),
                        "min_signal": float(group_signal.min()),
                        "max_signal": float(group_signal.max()),

                        # Weight turnover
                        "buy_turnover": weight_turnover["buy_turnover"],
                        "sell_turnover": weight_turnover["sell_turnover"],
                        "turnover_two_way": weight_turnover["turnover_two_way"],
                        "turnover_one_way": weight_turnover["turnover_one_way"],

                        # Membership turnover
                        "old_n": member_change["old_n"],
                        "new_n": member_change["new_n"],
                        "intersection_n": member_change["intersection_n"],
                        "union_n": member_change["union_n"],
                        "enter_n": member_change["enter_n"],
                        "exit_n": member_change["exit_n"],
                        "retention_ratio_old": member_change["retention_ratio_old"],
                        "enter_ratio_new": member_change["enter_ratio_new"],
                        "exit_ratio_old": member_change["exit_ratio_old"],
                        "member_turnover_ratio": member_change["member_turnover_ratio"],
                        "jaccard_distance": member_change["jaccard_distance"],
                    }
                )

                # After rebalance, current portfolio becomes target weights.
                current_weights[group] = new_target_w.copy()
                previous_target_weights[group] = new_target_w.copy()

            # Drift weights during the holding period.
            # This makes the next rebalance turnover closer to actual trading turnover.
            for interval_i in range(start_i + 1, start_i + h + 1):
                return_day = trading_days[interval_i + 1]
                ret_series = ret_by_day.get(return_day)

                if ret_series is None:
                    continue

                for group in required_groups:
                    current_weights[group] = _turnover_drift_weights_one_day(
                        weights=current_weights[group],
                        ret_series=ret_series,
                    )

    turnover_by_rebalance = pd.DataFrame(records)

    if turnover_by_rebalance.empty:
        raise ValueError("No turnover records generated. Please check signal coverage and grouping logic.")

    turnover_by_rebalance = turnover_by_rebalance.sort_values(
        ["horizon", "rebalance_id", "group"]
    ).reset_index(drop=True)

    return turnover_by_rebalance, panel, signal_col, signal_mode_used


# =============================================================================
# 4. Summary tables
# =============================================================================

def build_turnover_summary(turnover_by_rebalance: pd.DataFrame) -> pd.DataFrame:
    """
    Build group-level turnover summary.

    Main summary excludes initial rebalance.
    """
    work = turnover_by_rebalance.copy()
    work_non_initial = work[~work["is_initial_rebalance"].astype(bool)].copy()

    if work_non_initial.empty:
        print("Warning: no non-initial rebalances. Summary will use all rebalances.")
        work_non_initial = work.copy()

    summary = (
        work_non_initial
        .groupby(["horizon", "group"], as_index=False)
        .agg(
            rebalance_count=("rebalance_id", "count"),

            avg_n_stocks=("n_stocks", "mean"),
            median_n_stocks=("n_stocks", "median"),

            avg_buy_turnover=("buy_turnover", "mean"),
            avg_sell_turnover=("sell_turnover", "mean"),
            avg_turnover_two_way=("turnover_two_way", "mean"),
            avg_turnover_one_way=("turnover_one_way", "mean"),

            median_turnover_one_way=("turnover_one_way", "median"),
            std_turnover_one_way=("turnover_one_way", "std"),

            avg_member_turnover_ratio=("member_turnover_ratio", "mean"),
            median_member_turnover_ratio=("member_turnover_ratio", "median"),
            std_member_turnover_ratio=("member_turnover_ratio", "std"),

            avg_retention_ratio_old=("retention_ratio_old", "mean"),
            avg_jaccard_distance=("jaccard_distance", "mean"),

            avg_enter_n=("enter_n", "mean"),
            avg_exit_n=("exit_n", "mean"),
            avg_intersection_n=("intersection_n", "mean"),

            avg_signal=("avg_signal", "mean"),
            median_signal=("median_signal", "mean"),

            first_signal_day=("signal_day", "min"),
            last_signal_day=("signal_day", "max"),
        )
    )

    summary = summary.sort_values(["horizon", "group"]).reset_index(drop=True)

    return summary


def build_q1_relative_turnover_table(turnover_summary: pd.DataFrame) -> pd.DataFrame:
    """
    Compare Q2-Q5 turnover with Q1 turnover under the same horizon.
    """
    summary = turnover_summary.copy()

    q1 = summary[summary["group"] == "Q1"].copy()

    q1_cols = [
        "horizon",
        "avg_turnover_one_way",
        "avg_turnover_two_way",
        "avg_member_turnover_ratio",
        "avg_retention_ratio_old",
        "avg_jaccard_distance",
    ]

    q1 = q1[q1_cols].rename(
        columns={
            "avg_turnover_one_way": "q1_avg_turnover_one_way",
            "avg_turnover_two_way": "q1_avg_turnover_two_way",
            "avg_member_turnover_ratio": "q1_avg_member_turnover_ratio",
            "avg_retention_ratio_old": "q1_avg_retention_ratio_old",
            "avg_jaccard_distance": "q1_avg_jaccard_distance",
        }
    )

    out = summary.merge(q1, on="horizon", how="left")

    metrics = [
        ("avg_turnover_one_way", "q1_avg_turnover_one_way"),
        ("avg_turnover_two_way", "q1_avg_turnover_two_way"),
        ("avg_member_turnover_ratio", "q1_avg_member_turnover_ratio"),
        ("avg_retention_ratio_old", "q1_avg_retention_ratio_old"),
        ("avg_jaccard_distance", "q1_avg_jaccard_distance"),
    ]

    for metric, q1_metric in metrics:
        out[f"{metric}_minus_q1"] = out[metric] - out[q1_metric]
        out[f"{metric}_div_q1"] = np.where(
            out[q1_metric].abs() > 1e-12,
            out[metric] / out[q1_metric],
            np.nan,
        )

    out = out.sort_values(["horizon", "group"]).reset_index(drop=True)

    return out


def build_q1_vs_other_groups_summary(turnover_summary: pd.DataFrame) -> pd.DataFrame:
    """
    Summarize Q1 vs average of Q2-Q5.
    """
    records = []

    metrics = [
        "avg_turnover_one_way",
        "avg_turnover_two_way",
        "avg_member_turnover_ratio",
        "avg_retention_ratio_old",
        "avg_jaccard_distance",
    ]

    for h, sub in turnover_summary.groupby("horizon"):
        q1 = sub[sub["group"] == "Q1"]

        if q1.empty:
            continue

        others = sub[sub["group"].isin(["Q2", "Q3", "Q4", "Q5"])]

        if others.empty:
            continue

        row = {"horizon": h}

        for m in metrics:
            q1_val = float(q1.iloc[0][m])
            other_mean = float(others[m].mean())

            row[f"q1_{m}"] = q1_val
            row[f"q2_q5_avg_{m}"] = other_mean
            row[f"q2_q5_avg_minus_q1_{m}"] = other_mean - q1_val
            row[f"q2_q5_avg_div_q1_{m}"] = (
                other_mean / q1_val if abs(q1_val) > 1e-12 else np.nan
            )

        records.append(row)

    out = pd.DataFrame(records)

    if not out.empty:
        out = out.sort_values("horizon").reset_index(drop=True)

    return out


def build_turnover_performance_view(turnover_summary: pd.DataFrame) -> pd.DataFrame:
    """
    Optional table:
    merge turnover summary with original five-group performance summary.

    This helps inspect:
        Does Q1 have both different performance and different turnover?
    """
    if (
        "five_group_result" not in globals()
        or not isinstance(five_group_result, dict)
        or "performance_summary" not in five_group_result
    ):
        return pd.DataFrame()

    perf = five_group_result["performance_summary"].copy()

    if "portfolio" not in perf.columns:
        return pd.DataFrame()

    perf = perf[perf["portfolio"].isin([f"Q{i}" for i in range(1, TURNOVER_N_GROUPS + 1)])].copy()

    if perf.empty:
        return pd.DataFrame()

    merged = turnover_summary.merge(
        perf,
        left_on=["horizon", "group"],
        right_on=["horizon", "portfolio"],
        how="left",
        suffixes=("_turnover", "_performance"),
    )

    preferred_cols = [
        "horizon",
        "group",
        "avg_turnover_one_way",
        "avg_turnover_two_way",
        "avg_member_turnover_ratio",
        "avg_retention_ratio_old",
        "avg_jaccard_distance",
        "avg_n_stocks",
        "rebalance_count",

        "total_return",
        "annual_return",
        "annual_vol",
        "sharpe_no_rf",
        "max_drawdown",
        "avg_daily_return",
        "daily_win_rate",
    ]

    merged = merged[
        [c for c in preferred_cols if c in merged.columns]
        + [c for c in merged.columns if c not in preferred_cols]
    ]

    return merged.sort_values(["horizon", "group"]).reset_index(drop=True)


# =============================================================================
# 5. Plot
# =============================================================================

def plot_turnover_by_group(turnover_summary: pd.DataFrame):
    """
    Plot:
    1. Average one-way weight turnover by group
    2. Average member turnover by group
    """
    metrics = [
        (
            "avg_turnover_one_way",
            "Average One-way Weight Turnover",
            "avg_weight_turnover_by_group",
        ),
        (
            "avg_member_turnover_ratio",
            "Average Member Turnover Ratio",
            "avg_member_turnover_by_group",
        ),
        (
            "avg_retention_ratio_old",
            "Average Retention Ratio",
            "avg_retention_ratio_by_group",
        ),
        (
            "avg_jaccard_distance",
            "Average Jaccard Distance",
            "avg_jaccard_distance_by_group",
        ),
    ]

    for metric, title, filename_prefix in metrics:
        if metric not in turnover_summary.columns:
            continue

        for h, sub in turnover_summary.groupby("horizon"):
            sub = sub.sort_values("group").copy()

            fig, ax = plt.subplots(figsize=(9, 5))

            ax.bar(sub["group"], sub[metric])

            ax.set_title(f"{title}, rebalance every {h} trading day(s)")
            ax.set_xlabel("Group")
            ax.set_ylabel(metric)
            ax.grid(True, axis="y", alpha=0.3)

            fig.tight_layout()

            fig_path = TURNOVER_OUTDIR / f"{filename_prefix}_h{h}.png"
            fig.savefig(fig_path, dpi=300, bbox_inches="tight")
            plt.close(fig)

            print(f"Saved plot: {fig_path}")


# =============================================================================
# 6. Save Excel and format
# =============================================================================

def save_turnover_excel(
    turnover_summary: pd.DataFrame,
    turnover_by_rebalance: pd.DataFrame,
    q1_relative: pd.DataFrame,
    q1_vs_others: pd.DataFrame,
    turnover_performance_view: pd.DataFrame,
    signal_col: str,
    signal_mode_used: str,
):
    config = pd.DataFrame(
        [
            {"item": "signal_mode_used", "value": signal_mode_used},
            {"item": "signal_column", "value": signal_col},
            {"item": "rebalance_periods", "value": str(TURNOVER_REBALANCE_PERIODS)},
            {"item": "n_groups", "value": TURNOVER_N_GROUPS},
            {"item": "min_stocks_to_group", "value": TURNOVER_MIN_STOCKS_TO_GROUP},
            {"item": "main_summary", "value": "excludes initial opening rebalance"},
            {"item": "weight_turnover", "value": "calculated from previous drifted weights to new target weights"},
            {"item": "member_turnover", "value": "calculated from previous target members to new target members"},
            {"item": "turnover_one_way", "value": "0.5 * (buy_turnover + sell_turnover)"},
            {"item": "turnover_two_way", "value": "buy_turnover + sell_turnover"},
            {"item": "Q1", "value": "lowest signal"},
            {"item": "Q5", "value": "highest signal"},
        ]
    )

    methodology = pd.DataFrame(
        [
            {
                "Item": "Purpose",
                "Description": "Check whether Q1 has clearly different turnover / membership stability from Q2-Q5.",
            },
            {
                "Item": "Weight turnover",
                "Description": "At each rebalance, compare new equal-weight group portfolio with previous drifted portfolio weights.",
            },
            {
                "Item": "Member turnover",
                "Description": "Compares stock membership between two consecutive rebalance target portfolios.",
            },
            {
                "Item": "Retention ratio",
                "Description": "intersection_n / old_n. Higher value means the group is more stable.",
            },
            {
                "Item": "Member turnover ratio",
                "Description": "0.5 * (enter_n / new_n + exit_n / old_n). Higher value means group membership changes more.",
            },
            {
                "Item": "Jaccard distance",
                "Description": "1 - intersection_n / union_n. Higher value means two consecutive group member sets are more different.",
            },
            {
                "Item": "Initial rebalance",
                "Description": "Excluded from main summary because opening from cash mechanically generates high buy turnover.",
            },
        ]
    )

    with pd.ExcelWriter(TURNOVER_EXCEL, engine="openpyxl") as writer:
        turnover_summary.to_excel(writer, sheet_name="Turnover_Summary", index=False)
        q1_relative.to_excel(writer, sheet_name="Relative_To_Q1", index=False)
        q1_vs_others.to_excel(writer, sheet_name="Q1_vs_Q2_Q5", index=False)

        if turnover_performance_view is not None and not turnover_performance_view.empty:
            turnover_performance_view.to_excel(writer, sheet_name="Turnover_Performance", index=False)

        turnover_by_rebalance.to_excel(writer, sheet_name="Turnover_By_Rebalance", index=False)
        config.to_excel(writer, sheet_name="Config", index=False)
        methodology.to_excel(writer, sheet_name="Methodology", index=False)

    format_turnover_excel(TURNOVER_EXCEL)

    print(f"Saved turnover Excel: {TURNOVER_EXCEL}")


def format_turnover_excel(excel_path: Path):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(excel_path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    percent_cols = {
        "buy_turnover",
        "sell_turnover",
        "turnover_two_way",
        "turnover_one_way",

        "avg_buy_turnover",
        "avg_sell_turnover",
        "avg_turnover_two_way",
        "avg_turnover_one_way",
        "median_turnover_one_way",
        "std_turnover_one_way",

        "avg_member_turnover_ratio",
        "median_member_turnover_ratio",
        "std_member_turnover_ratio",
        "avg_retention_ratio_old",
        "avg_jaccard_distance",

        "retention_ratio_old",
        "enter_ratio_new",
        "exit_ratio_old",
        "member_turnover_ratio",
        "jaccard_distance",

        "q1_avg_turnover_one_way",
        "q1_avg_turnover_two_way",
        "q1_avg_member_turnover_ratio",
        "q1_avg_retention_ratio_old",
        "q1_avg_jaccard_distance",

        "avg_turnover_one_way_minus_q1",
        "avg_turnover_two_way_minus_q1",
        "avg_member_turnover_ratio_minus_q1",
        "avg_retention_ratio_old_minus_q1",
        "avg_jaccard_distance_minus_q1",

        "avg_turnover_one_way_div_q1",
        "avg_turnover_two_way_div_q1",
        "avg_member_turnover_ratio_div_q1",
        "avg_retention_ratio_old_div_q1",
        "avg_jaccard_distance_div_q1",

        "total_return",
        "annual_return",
        "annual_vol",
        "max_drawdown",
        "avg_daily_return",
        "daily_win_rate",
    }

    # Q1_vs_Q2_Q5 generated columns.
    for col in [
        "avg_turnover_one_way",
        "avg_turnover_two_way",
        "avg_member_turnover_ratio",
        "avg_retention_ratio_old",
        "avg_jaccard_distance",
    ]:
        percent_cols.update(
            {
                f"q1_{col}",
                f"q2_q5_avg_{col}",
                f"q2_q5_avg_minus_q1_{col}",
                f"q2_q5_avg_div_q1_{col}",
            }
        )

    numeric_4_cols = {
        "avg_signal",
        "median_signal",
        "min_signal",
        "max_signal",
        "sharpe_no_rf",
    }

    integer_cols = {
        "horizon",
        "rebalance_id",
        "rebalance_count",
        "n_stocks",
        "old_n",
        "new_n",
        "intersection_n",
        "union_n",
        "enter_n",
        "exit_n",
    }

    date_cols = {
        "signal_day",
        "entry_day",
        "exit_day",
        "first_signal_day",
        "last_signal_day",
    }

    for ws in wb.worksheets:
        if ws.max_row < 1 or ws.max_column < 1:
            continue

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        header_to_col = {
            str(ws.cell(row=1, column=col_idx).value): col_idx
            for col_idx in range(1, ws.max_column + 1)
        }

        for header, col_idx in header_to_col.items():
            if header in percent_cols:
                fmt = "0.00%"
            elif header in numeric_4_cols:
                fmt = "0.0000"
            elif header in integer_cols:
                fmt = "0"
            elif header in date_cols:
                fmt = "yyyy-mm-dd"
            else:
                fmt = None

            if fmt is not None:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = fmt

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            scan_rows = min(ws.max_row, 3000)

            for row_idx in range(1, scan_rows + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))

            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 36)

    wb.save(excel_path)


# =============================================================================
# 7. Run
# =============================================================================

def run_five_group_turnover_statistics():
    turnover_by_rebalance, panel, signal_col, signal_mode_used = calculate_five_group_turnover()

    turnover_summary = build_turnover_summary(turnover_by_rebalance)
    q1_relative = build_q1_relative_turnover_table(turnover_summary)
    q1_vs_others = build_q1_vs_other_groups_summary(turnover_summary)
    turnover_performance_view = build_turnover_performance_view(turnover_summary)

    # Save CSV
    turnover_by_rebalance.to_csv(
        TURNOVER_OUTDIR / "five_group_turnover_by_rebalance.csv",
        index=False,
        encoding="utf-8-sig",
    )

    turnover_summary.to_csv(
        TURNOVER_OUTDIR / "five_group_turnover_summary.csv",
        index=False,
        encoding="utf-8-sig",
    )

    q1_relative.to_csv(
        TURNOVER_OUTDIR / "five_group_turnover_relative_to_q1.csv",
        index=False,
        encoding="utf-8-sig",
    )

    q1_vs_others.to_csv(
        TURNOVER_OUTDIR / "five_group_turnover_q1_vs_q2_q5.csv",
        index=False,
        encoding="utf-8-sig",
    )

    if turnover_performance_view is not None and not turnover_performance_view.empty:
        turnover_performance_view.to_csv(
            TURNOVER_OUTDIR / "five_group_turnover_performance_view.csv",
            index=False,
            encoding="utf-8-sig",
        )

    # Save Excel
    save_turnover_excel(
        turnover_summary=turnover_summary,
        turnover_by_rebalance=turnover_by_rebalance,
        q1_relative=q1_relative,
        q1_vs_others=q1_vs_others,
        turnover_performance_view=turnover_performance_view,
        signal_col=signal_col,
        signal_mode_used=signal_mode_used,
    )

    # Plots
    plot_turnover_by_group(turnover_summary)

    print("\nTurnover summary:")
    show_cols = [
        "horizon",
        "group",
        "rebalance_count",
        "avg_n_stocks",
        "avg_turnover_one_way",
        "avg_turnover_two_way",
        "avg_member_turnover_ratio",
        "avg_retention_ratio_old",
        "avg_jaccard_distance",
    ]
    print(turnover_summary[[c for c in show_cols if c in turnover_summary.columns]].to_string(index=False))

    print("\nQ1 vs Q2-Q5 turnover comparison:")
    print(q1_vs_others.to_string(index=False))

    return {
        "turnover_by_rebalance": turnover_by_rebalance,
        "turnover_summary": turnover_summary,
        "q1_relative": q1_relative,
        "q1_vs_others": q1_vs_others,
        "turnover_performance_view": turnover_performance_view,
        "signal_panel": panel,
        "signal_col": signal_col,
        "signal_mode_used": signal_mode_used,
    }


if __name__ == "__main__":
    five_group_turnover_result = run_five_group_turnover_statistics()