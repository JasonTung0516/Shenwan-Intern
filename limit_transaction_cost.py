from __future__ import annotations

from pathlib import Path
from typing import Iterable, Sequence
import math
import re
import warnings

import numpy as np
import pandas as pd

try:
    from .config import *
except ImportError:  # allow running a file directly from this folder
    from config import *

try:
    from .core_base import *
    from .core_neutralization import *
    from .portfolio_five_group import *
    from .portfolio_industry_neutral import *
    from .transaction_cost import *
except ImportError:
    from core_base import *
    from core_neutralization import *
    from portfolio_five_group import *
    from portfolio_industry_neutral import *
    from transaction_cost import *

# Requirement:
# 1. Keep transaction cost:
#       buy cost  = 0.0005
#       sell cost = 0.001
# 2. Add limit-up / limit-down execution constraint on rebalance entry day:
#       if entry day is limit-up, cannot buy
#       if entry day is limit-down, cannot sell
# 3. Limit-up / limit-down is judged by comparing VWAP with limit-up / limit-down price:
#       np.isclose(vwap_price, limit_up_price)
#       np.isclose(vwap_price, limit_down_price)
# 4. If a trade is blocked:
#       old position remains unchanged for that stock
#       unfilled buy means cash position
#       unfilled sell means forced holding
# 5. Output standalone Excel.


# =============================================================================
# CONFIG
# =============================================================================

LIMIT_TC_OUTDIR = OUTDIR / "transaction_cost_limit_test"
LIMIT_TC_OUTDIR.mkdir(parents=True, exist_ok=True)

LIMIT_TC_EXCEL = LIMIT_TC_OUTDIR / "five_group_transaction_cost_limit_test.xlsx"

# Transaction cost
LIMIT_BUY_COST_RATE = 0.0005
LIMIT_SELL_COST_RATE = 0.001

# VWAP vs limit price tolerance.
# 5e-4 means 0.05%, useful because VWAP is calculated from amount / volume
# and may be slightly away from the theoretical limit price.
LIMIT_PRICE_RTOL = 5e-4
LIMIT_PRICE_ATOL = 1e-8

# If your parquet column names are not detected automatically,
# manually set them here, for example:
# LIMIT_UP_COL = "S_DQ_LIMIT"
# LIMIT_DOWN_COL = "S_DQ_STOPPING"
LIMIT_UP_COL = None
LIMIT_DOWN_COL = None

# Whether limit-up / limit-down price columns are raw prices.
# Usually True. If PRICE_MODE="adjusted", the code will multiply them by factor.
LIMIT_PRICE_IS_RAW = True

# For long-only Q1-Q5 portfolios:
# If sell is blocked, we do not allow the portfolio to become over-invested.
# Buy orders will be scaled down if available cash is insufficient.
ENFORCE_LONG_ONLY_CAPITAL_LIMIT = True

LIMIT_TC_PORTFOLIOS = [f"Q{i}" for i in range(1, N_GROUPS + 1)] + ["Q5_minus_Q1"]

RUN_ORIGINAL_EQUAL_WEIGHT_LIMIT_TC = True
RUN_INDUSTRY_NEUTRAL_LIMIT_TC = True

_LIMIT_PRICE_PANEL_CACHE = None


# =============================================================================
# 0. Check required functions from previous chunks
# =============================================================================

_required_for_limit_tc = [
    "find_input_files",
    "_get_parquet_columns",
    "prepare_data",
    "get_signal_panel_for_transaction_cost",
    "build_original_equal_weight_targets",
    "build_industry_neutral_targets",
    "_clean_weight_series",
    "calc_rebalance_turnover_and_cost",
    "calc_portfolio_return_and_drift_weights",
    "build_transaction_cost_nav_and_summary",
]

_missing_required = [name for name in _required_for_limit_tc if name not in globals()]

if _missing_required:
    raise NameError(
        "Please run the previous transaction-cost chunk first. "
        f"Missing required functions: {_missing_required}"
    )


# =============================================================================
# 1. Detect and load limit-up / limit-down price columns
# =============================================================================

LIMIT_UP_COL_CANDIDATES = [
    "limit_up",
    "up_limit",
    "limit_up_price",
    "up_limit_price",
    "high_limit",
    "upper_limit",
    "price_limit_up",
    "limit_price_up",
    "zt_price",
    "zhangting_price",
    "涨停价",
    "涨停价格",
    "S_DQ_LIMIT",
    "s_dq_limit",
]

LIMIT_DOWN_COL_CANDIDATES = [
    "limit_down",
    "down_limit",
    "limit_down_price",
    "down_limit_price",
    "low_limit",
    "lower_limit",
    "price_limit_down",
    "limit_price_down",
    "dt_price",
    "dieting_price",
    "跌停价",
    "跌停价格",
    "S_DQ_STOPPING",
    "s_dq_stopping",
]


def _find_first_existing_column(
    available_cols: list[str],
    candidates: list[str],
    manual_col: str | None = None,
) -> str | None:
    """
    Find first matching column name, case-insensitive.
    """
    if manual_col is not None:
        if manual_col in available_cols:
            return manual_col
        raise ValueError(
            f"Manually specified column {manual_col} not found. "
            f"Available columns include: {available_cols[:80]}"
        )

    lower_map = {str(c).lower(): c for c in available_cols}

    for c in candidates:
        if c in available_cols:
            return c

        c_lower = str(c).lower()
        if c_lower in lower_map:
            return lower_map[c_lower]

    return None


def detect_limit_price_columns_from_files(files: list[Path]) -> tuple[str, str]:
    """
    Detect limit-up and limit-down price columns from parquet schema.
    """
    all_cols = []

    for fp in files:
        cols = _get_parquet_columns(fp)

        if cols is not None:
            all_cols.extend(cols)

    all_cols = list(dict.fromkeys(all_cols))

    if not all_cols:
        raise ValueError(
            "Cannot read parquet schema columns. "
            "Please manually set LIMIT_UP_COL and LIMIT_DOWN_COL."
        )

    up_col = _find_first_existing_column(
        available_cols=all_cols,
        candidates=LIMIT_UP_COL_CANDIDATES,
        manual_col=LIMIT_UP_COL,
    )

    down_col = _find_first_existing_column(
        available_cols=all_cols,
        candidates=LIMIT_DOWN_COL_CANDIDATES,
        manual_col=LIMIT_DOWN_COL,
    )

    if up_col is None or down_col is None:
        print("\nCould not automatically detect limit-up / limit-down columns.")
        print("Available parquet columns:")
        print(all_cols)

        raise ValueError(
            "Please manually set LIMIT_UP_COL and LIMIT_DOWN_COL near the top of this chunk. "
            "Example: LIMIT_UP_COL = 'S_DQ_LIMIT', LIMIT_DOWN_COL = 'S_DQ_STOPPING'."
        )

    print(f"Detected limit-up column   : {up_col}")
    print(f"Detected limit-down column : {down_col}")

    return up_col, down_col


def read_raw_limit_price_data(
    files: list[Path],
    limit_up_col: str,
    limit_down_col: str,
) -> pd.DataFrame:
    """
    Read parquet files with necessary columns for limit detection.

    Need:
    - code
    - trade_day / datetime
    - close
    - volume
    - amount
    - factor
    - limit-up price column
    - limit-down price column
    """
    frames = []

    preferred_cols = [
        "code",
        "trade_day",
        "datetime",
        "close",
        "volume",
        "amount",
        "factor",
        limit_up_col,
        limit_down_col,
    ]

    for fp in files:
        print(f"Reading limit-price columns from {fp} ...")

        available_cols = _get_parquet_columns(fp)

        if available_cols is not None:
            cols = [c for c in preferred_cols if c in available_cols]
            df_part = pd.read_parquet(fp, columns=cols)
        else:
            df_part = pd.read_parquet(fp)
            cols = [c for c in preferred_cols if c in df_part.columns]
            df_part = df_part[cols].copy()

        missing = [c for c in [limit_up_col, limit_down_col] if c not in df_part.columns]
        if missing:
            raise ValueError(f"Missing limit columns in {fp}: {missing}")

        frames.append(df_part)

    return pd.concat(frames, ignore_index=True)


def build_limit_price_panel() -> pd.DataFrame:
    """
    Build daily stock panel with:
    - code
    - trade_day
    - vwap_price
    - limit_up_price
    - limit_down_price
    - is_limit_up
    - is_limit_down
    """
    files = find_input_files(
        DATA_DIR,
        years=YEARS,
        file_pattern=FILE_PATTERN,
        allow_missing=ALLOW_MISSING_FILES,
    )

    limit_up_col, limit_down_col = detect_limit_price_columns_from_files(files)

    raw = read_raw_limit_price_data(
        files=files,
        limit_up_col=limit_up_col,
        limit_down_col=limit_down_col,
    )

    # prepare_data keeps extra columns, and calculates vwap_price.
    df = prepare_data(
        raw,
        price_mode=PRICE_MODE,
        filter_a_share=FILTER_A_SHARE,
    )

    del raw

    if limit_up_col not in df.columns or limit_down_col not in df.columns:
        raise ValueError(
            f"After prepare_data(), cannot find {limit_up_col} / {limit_down_col}."
        )

    df[limit_up_col] = pd.to_numeric(df[limit_up_col], errors="coerce")
    df[limit_down_col] = pd.to_numeric(df[limit_down_col], errors="coerce")

    if (
        LIMIT_PRICE_IS_RAW
        and PRICE_MODE == "adjusted"
        and "factor" in df.columns
    ):
        factor_adj = pd.to_numeric(df["factor"], errors="coerce").where(
            pd.to_numeric(df["factor"], errors="coerce").notna(),
            1.0,
        )

        df["limit_up_price"] = df[limit_up_col] * factor_adj
        df["limit_down_price"] = df[limit_down_col] * factor_adj
    else:
        df["limit_up_price"] = df[limit_up_col]
        df["limit_down_price"] = df[limit_down_col]

    df["limit_up_price"] = df["limit_up_price"].replace([np.inf, -np.inf], np.nan)
    df["limit_down_price"] = df["limit_down_price"].replace([np.inf, -np.inf], np.nan)
    df["vwap_price"] = pd.to_numeric(df["vwap_price"], errors="coerce")

    valid_up = (
        df["vwap_price"].notna()
        & df["limit_up_price"].notna()
        & (df["vwap_price"] > 0)
        & (df["limit_up_price"] > 0)
    )

    valid_down = (
        df["vwap_price"].notna()
        & df["limit_down_price"].notna()
        & (df["vwap_price"] > 0)
        & (df["limit_down_price"] > 0)
    )

    # Limit-up:
    # Use np.isclose, and also allow VWAP slightly below limit price.
    df["is_limit_up"] = False
    df.loc[valid_up, "is_limit_up"] = (
        np.isclose(
            df.loc[valid_up, "vwap_price"],
            df.loc[valid_up, "limit_up_price"],
            rtol=LIMIT_PRICE_RTOL,
            atol=LIMIT_PRICE_ATOL,
        )
        | (
            df.loc[valid_up, "vwap_price"]
            >= df.loc[valid_up, "limit_up_price"] * (1.0 - LIMIT_PRICE_RTOL)
        )
    )

    # Limit-down:
    # Use np.isclose, and also allow VWAP slightly above limit-down price.
    df["is_limit_down"] = False
    df.loc[valid_down, "is_limit_down"] = (
        np.isclose(
            df.loc[valid_down, "vwap_price"],
            df.loc[valid_down, "limit_down_price"],
            rtol=LIMIT_PRICE_RTOL,
            atol=LIMIT_PRICE_ATOL,
        )
        | (
            df.loc[valid_down, "vwap_price"]
            <= df.loc[valid_down, "limit_down_price"] * (1.0 + LIMIT_PRICE_RTOL)
        )
    )

    out = df[
        [
            "code",
            "trade_day",
            "vwap_price",
            "limit_up_price",
            "limit_down_price",
            "is_limit_up",
            "is_limit_down",
        ]
    ].copy()

    out["trade_day"] = pd.to_datetime(out["trade_day"]).dt.normalize()
    out["code"] = out["code"].astype(str)

    out = out.drop_duplicates(["code", "trade_day"], keep="last")

    print("\nLimit detection diagnostics:")
    print(f"Rows with limit-up flag   : {out['is_limit_up'].sum():,}")
    print(f"Rows with limit-down flag : {out['is_limit_down'].sum():,}")
    print(f"Total rows                : {len(out):,}")

    return out


def get_limit_price_panel_cached() -> pd.DataFrame:
    """
    Cache limit price panel because reading parquet again can be slow.
    """
    global _LIMIT_PRICE_PANEL_CACHE

    if _LIMIT_PRICE_PANEL_CACHE is None:
        _LIMIT_PRICE_PANEL_CACHE = build_limit_price_panel()

    return _LIMIT_PRICE_PANEL_CACHE.copy()


def add_limit_status_to_signal_panel(signal_panel: pd.DataFrame) -> pd.DataFrame:
    """
    Merge limit-up / limit-down status into existing signal panel.
    """
    panel = signal_panel.copy()
    panel["trade_day"] = pd.to_datetime(panel["trade_day"]).dt.normalize()
    panel["code"] = panel["code"].astype(str)

    limit_panel = get_limit_price_panel_cached()

    drop_cols = [
        "limit_up_price",
        "limit_down_price",
        "is_limit_up",
        "is_limit_down",
    ]

    panel = panel.drop(columns=[c for c in drop_cols if c in panel.columns])

    panel = panel.merge(
        limit_panel[
            [
                "code",
                "trade_day",
                "limit_up_price",
                "limit_down_price",
                "is_limit_up",
                "is_limit_down",
            ]
        ],
        on=["code", "trade_day"],
        how="left",
    )

    panel["is_limit_up"] = panel["is_limit_up"].fillna(False).astype(bool)
    panel["is_limit_down"] = panel["is_limit_down"].fillna(False).astype(bool)

    return panel


# =============================================================================
# 2. Rebalance execution with limit-up / limit-down constraints
# =============================================================================

def execute_rebalance_with_limit_constraints(
    old_weights: pd.Series,
    desired_weights: pd.Series,
    entry_limit_status: pd.DataFrame | None,
    portfolio: str,
) -> tuple[pd.Series, dict]:
    """
    Convert desired target weights into actual executed weights.

    Rule:
        delta = desired_weight - old_weight

        delta > 0 means buy.
        If stock is limit-up on entry day, this buy cannot be executed.

        delta < 0 means sell.
        If stock is limit-down on entry day, this sell cannot be executed.

    Unexecuted trade:
        actual_weight remains old_weight.
    """
    old_w = _clean_weight_series(old_weights)
    desired_w = _clean_weight_series(desired_weights)

    all_codes = old_w.index.union(desired_w.index)

    old_aligned = old_w.reindex(all_codes).fillna(0.0)
    desired_aligned = desired_w.reindex(all_codes).fillna(0.0)

    desired_delta = desired_aligned - old_aligned

    if entry_limit_status is None or len(entry_limit_status) == 0:
        is_limit_up = pd.Series(False, index=all_codes)
        is_limit_down = pd.Series(False, index=all_codes)
    else:
        lim = entry_limit_status.copy()

        if "code" in lim.columns:
            lim["code"] = lim["code"].astype(str)
            lim = lim.set_index("code")

        lim.index = lim.index.astype(str)

        is_limit_up = lim.get("is_limit_up", pd.Series(False, index=lim.index))
        is_limit_down = lim.get("is_limit_down", pd.Series(False, index=lim.index))

        is_limit_up = is_limit_up.reindex(all_codes).fillna(False).astype(bool)
        is_limit_down = is_limit_down.reindex(all_codes).fillna(False).astype(bool)

    eps = 1e-15

    buy_blocked = (desired_delta > eps) & is_limit_up
    sell_blocked = (desired_delta < -eps) & is_limit_down

    allowed_delta = desired_delta.copy()
    allowed_delta.loc[buy_blocked | sell_blocked] = 0.0

    capital_scaled_buy_turnover = 0.0
    capital_scale = 1.0

    # Optional capital constraint for long-only Q1-Q5 portfolios.
    # If sell is blocked, we may not have enough cash to buy all desired new positions.
    # We scale down allowed buys to avoid over-investment.
    is_long_only = (
        (old_aligned.min() >= -eps)
        and (desired_aligned.min() >= -eps)
        and portfolio != "Q5_minus_Q1"
    )

    if ENFORCE_LONG_ONLY_CAPITAL_LIMIT and is_long_only:
        sell_delta = allowed_delta.clip(upper=0.0)
        buy_delta = allowed_delta.clip(lower=0.0)

        after_allowed_sells = old_aligned + sell_delta
        after_allowed_sells = after_allowed_sells.clip(lower=0.0)

        target_long_exposure = desired_aligned[desired_aligned > 0].sum()

        if not np.isfinite(target_long_exposure) or target_long_exposure <= 0:
            target_long_exposure = 1.0

        current_exposure_after_sells = after_allowed_sells.sum()
        available_buy_capacity = max(target_long_exposure - current_exposure_after_sells, 0.0)

        buy_sum = buy_delta.sum()

        if buy_sum > available_buy_capacity + eps:
            capital_scale = available_buy_capacity / buy_sum if buy_sum > 0 else 0.0
            scaled_buy_delta = buy_delta * capital_scale
            capital_scaled_buy_turnover = float(buy_sum - scaled_buy_delta.sum())
            allowed_delta = sell_delta + scaled_buy_delta

    actual_aligned = old_aligned + allowed_delta
    actual_aligned = actual_aligned.where(actual_aligned.abs() > eps, 0.0)

    actual_w = _clean_weight_series(actual_aligned)

    realized_delta = actual_aligned - old_aligned
    unexecuted_delta = desired_aligned - actual_aligned

    buy_turnover = float(realized_delta.clip(lower=0.0).sum())
    sell_turnover = float((-realized_delta.clip(upper=0.0)).sum())
    turnover = buy_turnover + sell_turnover

    transaction_cost = (
        buy_turnover * LIMIT_BUY_COST_RATE
        + sell_turnover * LIMIT_SELL_COST_RATE
    )

    failed_buy_turnover = float(unexecuted_delta.clip(lower=0.0).sum())
    failed_sell_turnover = float((-unexecuted_delta.clip(upper=0.0)).sum())

    info = {
        "buy_turnover": buy_turnover,
        "sell_turnover": sell_turnover,
        "turnover": turnover,
        "transaction_cost": float(transaction_cost),

        "failed_buy_turnover": failed_buy_turnover,
        "failed_sell_turnover": failed_sell_turnover,
        "failed_turnover": failed_buy_turnover + failed_sell_turnover,

        "buy_blocked_count": int(buy_blocked.sum()),
        "sell_blocked_count": int(sell_blocked.sum()),

        "capital_scaled_buy_turnover": capital_scaled_buy_turnover,
        "capital_scale": capital_scale,

        "desired_n_names": int((desired_aligned.abs() > eps).sum()),
        "actual_n_names": int((actual_aligned.abs() > eps).sum()),

        "desired_long_exposure": float(desired_aligned[desired_aligned > 0].sum()),
        "desired_short_exposure": float(desired_aligned[desired_aligned < 0].sum()),
        "desired_gross_exposure": float(desired_aligned.abs().sum()),
        "desired_net_exposure": float(desired_aligned.sum()),

        "actual_long_exposure": float(actual_aligned[actual_aligned > 0].sum()),
        "actual_short_exposure": float(actual_aligned[actual_aligned < 0].sum()),
        "actual_gross_exposure": float(actual_aligned.abs().sum()),
        "actual_net_exposure": float(actual_aligned.sum()),
    }

    return actual_w, info


# =============================================================================
# 3. Backtest with transaction cost + limit constraints
# =============================================================================

def run_one_limit_transaction_cost_backtest(
    signal_panel: pd.DataFrame,
    signal_col: str,
    method_name: str,
    target_builder,
    rebalance_periods: list[int] = GROUP_REBALANCE_PERIODS,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Run one backtest with:
    - transaction cost
    - limit-up cannot buy
    - limit-down cannot sell
    """
    panel = signal_panel.copy()
    panel["trade_day"] = pd.to_datetime(panel["trade_day"]).dt.normalize()
    panel["code"] = panel["code"].astype(str)

    panel = panel.sort_values(["code", "trade_day"]).reset_index(drop=True)

    panel["ret_vwap_1d"] = (
        panel.groupby("code", sort=False)["vwap_price"].pct_change()
    )
    panel["ret_vwap_1d"] = panel["ret_vwap_1d"].replace([np.inf, -np.inf], np.nan)

    trading_days = sorted(panel["trade_day"].dropna().unique())

    ret_by_day = {}

    for day, sub in panel[["code", "trade_day", "ret_vwap_1d"]].groupby("trade_day"):
        temp = sub.copy()
        temp["code"] = temp["code"].astype(str)
        ret_by_day[day] = temp.set_index("code")["ret_vwap_1d"]

    limit_by_day = {}

    for day, sub in panel[
        [
            "code",
            "trade_day",
            "is_limit_up",
            "is_limit_down",
            "limit_up_price",
            "limit_down_price",
        ]
    ].groupby("trade_day"):
        temp = sub.copy()
        temp["code"] = temp["code"].astype(str)
        limit_by_day[day] = temp.set_index("code")

    signal_cols = ["code", "trade_day", "volume", signal_col]

    # Needed only by industry-neutral target builder.
    for c in [
        "industry",
        "industry_valid_for_group",
        "mkt_cap",
        "is_limit_up",
        "is_limit_down",
        "limit_up_price",
        "limit_down_price",
    ]:
        if c in panel.columns and c not in signal_cols:
            signal_cols.append(c)

    signal_by_day = {
        day: sub[signal_cols].copy()
        for day, sub in panel[signal_cols].groupby("trade_day")
    }

    daily_records = []
    rebalance_records = []

    for h in rebalance_periods:
        print(
            f"\nRunning limit + transaction-cost backtest: "
            f"{method_name}, rebalance_period={h}d"
        )

        current_weights = {
            portfolio: pd.Series(dtype=float)
            for portfolio in LIMIT_TC_PORTFOLIOS
        }

        valid_start_end = len(trading_days) - h

        for start_i in range(0, valid_start_end, h):
            signal_day = trading_days[start_i]

            if start_i + h + 1 >= len(trading_days):
                break

            entry_day = trading_days[start_i + 1]
            exit_day = trading_days[start_i + h + 1]

            signal_day_df = signal_by_day.get(signal_day)

            if signal_day_df is None or signal_day_df.empty:
                continue

            desired_target_weights = target_builder(
                signal_day_df=signal_day_df,
                signal_col=signal_col,
            )

            if not desired_target_weights:
                continue

            if not set(LIMIT_TC_PORTFOLIOS).issubset(set(desired_target_weights.keys())):
                continue

            entry_limit_status = limit_by_day.get(entry_day, pd.DataFrame())

            rebalance_info = {}

            for portfolio in LIMIT_TC_PORTFOLIOS:
                old_w = current_weights.get(portfolio, pd.Series(dtype=float))
                desired_w = desired_target_weights[portfolio].copy()

                actual_w, exec_info = execute_rebalance_with_limit_constraints(
                    old_weights=old_w,
                    desired_weights=desired_w,
                    entry_limit_status=entry_limit_status,
                    portfolio=portfolio,
                )

                current_weights[portfolio] = actual_w.copy()
                rebalance_info[portfolio] = exec_info

                rebalance_records.append(
                    {
                        "method": method_name,
                        "horizon": h,
                        "signal_day": signal_day,
                        "entry_day": entry_day,
                        "exit_day": exit_day,
                        "portfolio": portfolio,

                        "desired_n_names": exec_info["desired_n_names"],
                        "actual_n_names": exec_info["actual_n_names"],

                        "desired_long_exposure": exec_info["desired_long_exposure"],
                        "desired_short_exposure": exec_info["desired_short_exposure"],
                        "desired_gross_exposure": exec_info["desired_gross_exposure"],
                        "desired_net_exposure": exec_info["desired_net_exposure"],

                        "actual_long_exposure": exec_info["actual_long_exposure"],
                        "actual_short_exposure": exec_info["actual_short_exposure"],
                        "actual_gross_exposure": exec_info["actual_gross_exposure"],
                        "actual_net_exposure": exec_info["actual_net_exposure"],

                        "buy_turnover": exec_info["buy_turnover"],
                        "sell_turnover": exec_info["sell_turnover"],
                        "turnover": exec_info["turnover"],
                        "transaction_cost": exec_info["transaction_cost"],

                        "failed_buy_turnover": exec_info["failed_buy_turnover"],
                        "failed_sell_turnover": exec_info["failed_sell_turnover"],
                        "failed_turnover": exec_info["failed_turnover"],

                        "buy_blocked_count": exec_info["buy_blocked_count"],
                        "sell_blocked_count": exec_info["sell_blocked_count"],

                        "capital_scaled_buy_turnover": exec_info["capital_scaled_buy_turnover"],
                        "capital_scale": exec_info["capital_scale"],

                        "buy_cost_rate": LIMIT_BUY_COST_RATE,
                        "sell_cost_rate": LIMIT_SELL_COST_RATE,
                    }
                )

            # Daily holding returns.
            # Transaction cost is charged on the first holding return day.
            for interval_i in range(start_i + 1, start_i + h + 1):
                return_day = trading_days[interval_i + 1]
                ret_series = ret_by_day.get(return_day)

                if ret_series is None:
                    continue

                is_first_holding_day = interval_i == start_i + 1

                for portfolio in LIMIT_TC_PORTFOLIOS:
                    current_w = _clean_weight_series(current_weights[portfolio])

                    # If all trades are blocked and position is cash, return = 0.
                    if current_w.empty:
                        gross_ret = 0.0
                        drifted_weights = current_w.copy()
                        n_ret_obs = 0
                    else:
                        gross_ret, drifted_weights, n_ret_obs = calc_portfolio_return_and_drift_weights(
                            current_weights=current_w,
                            ret_series=ret_series,
                        )

                    info = rebalance_info[portfolio]

                    tc = info["transaction_cost"] if is_first_holding_day else 0.0
                    buy_turnover = info["buy_turnover"] if is_first_holding_day else 0.0
                    sell_turnover = info["sell_turnover"] if is_first_holding_day else 0.0
                    turnover = info["turnover"] if is_first_holding_day else 0.0

                    failed_buy_turnover = (
                        info["failed_buy_turnover"] if is_first_holding_day else 0.0
                    )
                    failed_sell_turnover = (
                        info["failed_sell_turnover"] if is_first_holding_day else 0.0
                    )
                    failed_turnover = (
                        info["failed_turnover"] if is_first_holding_day else 0.0
                    )

                    buy_blocked_count = (
                        info["buy_blocked_count"] if is_first_holding_day else 0
                    )
                    sell_blocked_count = (
                        info["sell_blocked_count"] if is_first_holding_day else 0
                    )

                    if pd.notna(gross_ret):
                        net_ret = (1.0 - tc) * (1.0 + gross_ret) - 1.0
                    else:
                        net_ret = np.nan

                    daily_records.append(
                        {
                            "method": method_name,
                            "horizon": h,
                            "trade_day": return_day,
                            "signal_day": signal_day,
                            "entry_day": entry_day,
                            "exit_day": exit_day,
                            "portfolio": portfolio,

                            # gross_ret already includes limit-up / limit-down execution impact
                            # but before transaction cost.
                            "gross_ret": gross_ret,
                            "transaction_cost": tc,
                            "net_ret": net_ret,

                            "buy_turnover": buy_turnover,
                            "sell_turnover": sell_turnover,
                            "turnover": turnover,

                            "failed_buy_turnover": failed_buy_turnover,
                            "failed_sell_turnover": failed_sell_turnover,
                            "failed_turnover": failed_turnover,

                            "buy_blocked_count": buy_blocked_count,
                            "sell_blocked_count": sell_blocked_count,

                            "n_ret_obs": n_ret_obs,
                            "is_rebalance_cost_day": is_first_holding_day,
                        }
                    )

                    if pd.notna(gross_ret):
                        current_weights[portfolio] = drifted_weights.copy()

    daily_limit_tc_returns = pd.DataFrame(daily_records)
    rebalance_limit_tc = pd.DataFrame(rebalance_records)

    if daily_limit_tc_returns.empty:
        raise ValueError(
            f"No limit + transaction-cost returns generated for {method_name}. "
            "Please check limit-price columns and signal coverage."
        )

    return daily_limit_tc_returns, rebalance_limit_tc


# =============================================================================
# 4. Summary and comparison
# =============================================================================

def build_limit_event_summary(rebalance_limit_tc: pd.DataFrame) -> pd.DataFrame:
    """
    Summarize limit-up / limit-down execution failures.
    """
    if rebalance_limit_tc.empty:
        return pd.DataFrame()

    out = (
        rebalance_limit_tc
        .groupby(["method", "horizon", "portfolio"], as_index=False)
        .agg(
            rebalance_count=("signal_day", "count"),

            total_failed_buy_turnover=("failed_buy_turnover", "sum"),
            total_failed_sell_turnover=("failed_sell_turnover", "sum"),
            total_failed_turnover=("failed_turnover", "sum"),

            avg_failed_buy_turnover=("failed_buy_turnover", "mean"),
            avg_failed_sell_turnover=("failed_sell_turnover", "mean"),
            avg_failed_turnover=("failed_turnover", "mean"),

            total_buy_blocked_count=("buy_blocked_count", "sum"),
            total_sell_blocked_count=("sell_blocked_count", "sum"),
            avg_buy_blocked_count=("buy_blocked_count", "mean"),
            avg_sell_blocked_count=("sell_blocked_count", "mean"),

            avg_actual_gross_exposure=("actual_gross_exposure", "mean"),
            avg_actual_net_exposure=("actual_net_exposure", "mean"),
            avg_desired_gross_exposure=("desired_gross_exposure", "mean"),
            avg_desired_net_exposure=("desired_net_exposure", "mean"),

            avg_capital_scaled_buy_turnover=("capital_scaled_buy_turnover", "mean"),
        )
        .sort_values(["method", "horizon", "portfolio"])
        .reset_index(drop=True)
    )

    return out


def build_q5q1_comparison_with_cost_only(
    limit_q5q1_summary: pd.DataFrame,
) -> pd.DataFrame:
    """
    Compare:
    1. Transaction cost only result
    2. Limit + transaction cost result

    Uses transaction_cost_result from previous chunk if available.
    """
    if (
        "transaction_cost_result" not in globals()
        or not isinstance(transaction_cost_result, dict)
        or "q5q1_summary" not in transaction_cost_result
    ):
        print(
            "transaction_cost_result['q5q1_summary'] not found. "
            "Skipping comparison with transaction-cost-only result."
        )
        return pd.DataFrame()

    cost_only = transaction_cost_result["q5q1_summary"].copy()
    limit_tc = limit_q5q1_summary.copy()

    keep_metrics = [
        "total_return_net",
        "annual_return_net",
        "annual_vol_net",
        "sharpe_no_rf_net",
        "max_drawdown_net",
        "avg_daily_return_net",
        "daily_win_rate_net",
        "avg_turnover_on_rebalance_day",
        "avg_rebalance_day_transaction_cost",
    ]

    left_cols = ["method", "horizon", "portfolio"] + [
        c for c in keep_metrics if c in cost_only.columns
    ]

    right_cols = ["method", "horizon", "portfolio"] + [
        c for c in keep_metrics if c in limit_tc.columns
    ]

    merged = cost_only[left_cols].merge(
        limit_tc[right_cols],
        on=["method", "horizon", "portfolio"],
        how="inner",
        suffixes=("_cost_only", "_limit_tc"),
    )

    for m in keep_metrics:
        c1 = f"{m}_cost_only"
        c2 = f"{m}_limit_tc"

        if c1 in merged.columns and c2 in merged.columns:
            merged[f"{m}_decay_from_limit"] = merged[c2] - merged[c1]

    return merged.sort_values(["method", "horizon"]).reset_index(drop=True)


# =============================================================================
# 5. Excel formatting
# =============================================================================

def format_limit_tc_excel(excel_path: Path):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(excel_path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    percent_cols = {
        "total_transaction_cost",
        "avg_daily_transaction_cost",
        "avg_rebalance_day_transaction_cost",

        "avg_buy_turnover_on_rebalance_day",
        "avg_sell_turnover_on_rebalance_day",
        "avg_turnover_on_rebalance_day",

        "total_return_gross",
        "total_return_net",
        "total_return_decay",

        "annual_return_gross",
        "annual_return_net",
        "annual_return_decay",

        "annual_vol_gross",
        "annual_vol_net",
        "annual_vol_decay",

        "max_drawdown_gross",
        "max_drawdown_net",
        "max_drawdown_decay",

        "avg_daily_return_gross",
        "avg_daily_return_net",
        "avg_daily_return_decay",

        "daily_win_rate_gross",
        "daily_win_rate_net",
        "daily_win_rate_decay",

        "buy_turnover",
        "sell_turnover",
        "turnover",
        "transaction_cost",

        "failed_buy_turnover",
        "failed_sell_turnover",
        "failed_turnover",

        "total_failed_buy_turnover",
        "total_failed_sell_turnover",
        "total_failed_turnover",
        "avg_failed_buy_turnover",
        "avg_failed_sell_turnover",
        "avg_failed_turnover",

        "desired_long_exposure",
        "desired_short_exposure",
        "desired_gross_exposure",
        "desired_net_exposure",

        "actual_long_exposure",
        "actual_short_exposure",
        "actual_gross_exposure",
        "actual_net_exposure",

        "avg_actual_gross_exposure",
        "avg_actual_net_exposure",
        "avg_desired_gross_exposure",
        "avg_desired_net_exposure",

        "capital_scaled_buy_turnover",
        "avg_capital_scaled_buy_turnover",

        "buy_cost_rate",
        "sell_cost_rate",

        "gross_ret",
        "net_ret",

        "total_return_net_cost_only",
        "total_return_net_limit_tc",
        "total_return_net_decay_from_limit",

        "annual_return_net_cost_only",
        "annual_return_net_limit_tc",
        "annual_return_net_decay_from_limit",

        "annual_vol_net_cost_only",
        "annual_vol_net_limit_tc",
        "annual_vol_net_decay_from_limit",

        "max_drawdown_net_cost_only",
        "max_drawdown_net_limit_tc",
        "max_drawdown_net_decay_from_limit",

        "avg_daily_return_net_cost_only",
        "avg_daily_return_net_limit_tc",
        "avg_daily_return_net_decay_from_limit",

        "daily_win_rate_net_cost_only",
        "daily_win_rate_net_limit_tc",
        "daily_win_rate_net_decay_from_limit",

        "avg_turnover_on_rebalance_day_cost_only",
        "avg_turnover_on_rebalance_day_limit_tc",
        "avg_turnover_on_rebalance_day_decay_from_limit",

        "avg_rebalance_day_transaction_cost_cost_only",
        "avg_rebalance_day_transaction_cost_limit_tc",
        "avg_rebalance_day_transaction_cost_decay_from_limit",
    }

    numeric_4_cols = {
        "sharpe_no_rf_gross",
        "sharpe_no_rf_net",
        "sharpe_no_rf_decay",

        "gross_nav",
        "net_nav",

        "capital_scale",

        "sharpe_no_rf_net_cost_only",
        "sharpe_no_rf_net_limit_tc",
        "sharpe_no_rf_net_decay_from_limit",
    }

    integer_cols = {
        "horizon",
        "n_days",
        "rebalance_count",
        "desired_n_names",
        "actual_n_names",
        "buy_blocked_count",
        "sell_blocked_count",
        "total_buy_blocked_count",
        "total_sell_blocked_count",
        "n_ret_obs",
    }

    date_cols = {
        "trade_day",
        "signal_day",
        "entry_day",
        "exit_day",
    }

    for ws in wb.worksheets:
        if ws.max_row < 1 or ws.max_column < 1:
            continue

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        header_to_col = {
            str(ws.cell(row=1, column=col_idx).value): col_idx
            for col_idx in range(1, ws.max_column + 1)
        }

        for header, col_idx in header_to_col.items():
            if header in percent_cols:
                fmt = "0.00%"
            elif header in numeric_4_cols:
                fmt = "0.0000"
            elif header in integer_cols:
                fmt = "0"
            elif header in date_cols:
                fmt = "yyyy-mm-dd"
            else:
                fmt = None

            if fmt is not None:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = fmt

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            scan_rows = min(ws.max_row, 3000)

            for row_idx in range(1, scan_rows + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))

            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 34)

    wb.save(excel_path)


# =============================================================================
# 6. Run full limit + transaction cost test
# =============================================================================

def run_limit_transaction_cost_tests():
    all_daily = []
    all_rebalance = []

    if SIGNAL_MODE == "weighted":
        signal_col = GROUP_SIGNAL_WEIGHTED_COL
    elif SIGNAL_MODE == "equal":
        signal_col = GROUP_SIGNAL_EQUAL_COL
    else:
        raise ValueError("SIGNAL_MODE must be 'weighted' or 'equal'.")

    if RUN_ORIGINAL_EQUAL_WEIGHT_LIMIT_TC:
        original_signal_panel, _ = get_signal_panel_for_transaction_cost(
            method_name="Original equal-weight"
        )

        original_signal_panel = add_limit_status_to_signal_panel(original_signal_panel)

        original_daily, original_rebalance = run_one_limit_transaction_cost_backtest(
            signal_panel=original_signal_panel,
            signal_col=signal_col,
            method_name="Original equal-weight",
            target_builder=build_original_equal_weight_targets,
            rebalance_periods=GROUP_REBALANCE_PERIODS,
        )

        all_daily.append(original_daily)
        all_rebalance.append(original_rebalance)

    if RUN_INDUSTRY_NEUTRAL_LIMIT_TC:
        ind_signal_panel, _ = get_signal_panel_for_transaction_cost(
            method_name="Industry-neutral weight"
        )

        ind_signal_panel = add_limit_status_to_signal_panel(ind_signal_panel)

        industry_daily, industry_rebalance = run_one_limit_transaction_cost_backtest(
            signal_panel=ind_signal_panel,
            signal_col=signal_col,
            method_name="Industry-neutral weight",
            target_builder=build_industry_neutral_targets,
            rebalance_periods=GROUP_REBALANCE_PERIODS,
        )

        all_daily.append(industry_daily)
        all_rebalance.append(industry_rebalance)

    daily_limit_tc_returns = pd.concat(all_daily, ignore_index=True)
    rebalance_limit_tc = pd.concat(all_rebalance, ignore_index=True)

    nav_df, summary_df = build_transaction_cost_nav_and_summary(daily_limit_tc_returns)

    q5q1_summary = (
        summary_df[summary_df["portfolio"] == "Q5_minus_Q1"]
        .sort_values(["method", "horizon"])
        .reset_index(drop=True)
    )

    q_group_summary = (
        summary_df[summary_df["portfolio"] != "Q5_minus_Q1"]
        .sort_values(["method", "horizon", "portfolio"])
        .reset_index(drop=True)
    )

    limit_event_summary = build_limit_event_summary(rebalance_limit_tc)

    q5q1_compare_with_cost_only = build_q5q1_comparison_with_cost_only(q5q1_summary)

    methodology = pd.DataFrame(
        [
            {
                "Item": "Limit-up rule",
                "Description": "If a stock is limit-up on the rebalance entry day, buy trades are blocked.",
            },
            {
                "Item": "Limit-down rule",
                "Description": "If a stock is limit-down on the rebalance entry day, sell trades are blocked.",
            },
            {
                "Item": "Limit detection",
                "Description": f"Use np.isclose(VWAP, limit price), rtol={LIMIT_PRICE_RTOL}, atol={LIMIT_PRICE_ATOL}. Also allow VWAP to be slightly inside the limit price.",
            },
            {
                "Item": "Blocked buy",
                "Description": "The target buy is not executed. If old weight is zero, the stock becomes cash with zero return.",
            },
            {
                "Item": "Blocked sell",
                "Description": "The target sell is not executed. The old position continues to be held and earns the next holding-period return.",
            },
            {
                "Item": "Transaction cost",
                "Description": f"Executed buy turnover * {LIMIT_BUY_COST_RATE} + executed sell turnover * {LIMIT_SELL_COST_RATE}. Blocked trades do not pay transaction cost.",
            },
            {
                "Item": "Gross return",
                "Description": "Return after limit-up / limit-down execution constraint, before transaction cost.",
            },
            {
                "Item": "Net return",
                "Description": "Return after limit-up / limit-down execution constraint and transaction cost.",
            },
            {
                "Item": "Long-only capital rule",
                "Description": "For Q1-Q5 long-only portfolios, if sell orders are blocked, buy orders are scaled down if needed to avoid over-investment.",
            },
        ]
    )

    config = pd.DataFrame(
        [
            {"item": "signal_mode", "value": SIGNAL_MODE},
            {"item": "signal_column", "value": signal_col},
            {"item": "rebalance_periods", "value": str(GROUP_REBALANCE_PERIODS)},
            {"item": "buy_cost_rate", "value": LIMIT_BUY_COST_RATE},
            {"item": "sell_cost_rate", "value": LIMIT_SELL_COST_RATE},
            {"item": "limit_price_rtol", "value": LIMIT_PRICE_RTOL},
            {"item": "limit_price_atol", "value": LIMIT_PRICE_ATOL},
            {"item": "limit_price_is_raw", "value": LIMIT_PRICE_IS_RAW},
            {"item": "price_mode", "value": PRICE_MODE},
            {"item": "enforce_long_only_capital_limit", "value": ENFORCE_LONG_ONLY_CAPITAL_LIMIT},
            {"item": "run_original_equal_weight", "value": RUN_ORIGINAL_EQUAL_WEIGHT_LIMIT_TC},
            {"item": "run_industry_neutral", "value": RUN_INDUSTRY_NEUTRAL_LIMIT_TC},
            {"item": "industry_weight_method", "value": globals().get("INDUSTRY_WEIGHT_METHOD", np.nan)},
            {"item": "use_common_industries_only", "value": globals().get("USE_COMMON_INDUSTRIES_ONLY", np.nan)},
        ]
    )

    # Save CSV
    summary_df.to_csv(
        LIMIT_TC_OUTDIR / "limit_tc_performance_summary_all.csv",
        index=False,
        encoding="utf-8-sig",
    )

    q5q1_summary.to_csv(
        LIMIT_TC_OUTDIR / "limit_tc_q5_minus_q1_summary.csv",
        index=False,
        encoding="utf-8-sig",
    )

    nav_df.to_csv(
        LIMIT_TC_OUTDIR / "limit_tc_daily_nav.csv",
        index=False,
        encoding="utf-8-sig",
    )

    rebalance_limit_tc.to_csv(
        LIMIT_TC_OUTDIR / "limit_tc_rebalance_execution.csv",
        index=False,
        encoding="utf-8-sig",
    )

    limit_event_summary.to_csv(
        LIMIT_TC_OUTDIR / "limit_tc_limit_event_summary.csv",
        index=False,
        encoding="utf-8-sig",
    )

    if not q5q1_compare_with_cost_only.empty:
        q5q1_compare_with_cost_only.to_csv(
            LIMIT_TC_OUTDIR / "limit_tc_q5q1_compare_with_cost_only.csv",
            index=False,
            encoding="utf-8-sig",
        )

    # Save Excel
    with pd.ExcelWriter(LIMIT_TC_EXCEL, engine="openpyxl") as writer:
        q5q1_summary.to_excel(writer, sheet_name="Q5_Q1_Summary", index=False)
        q5q1_compare_with_cost_only.to_excel(writer, sheet_name="Compare_With_Cost_Only", index=False)
        limit_event_summary.to_excel(writer, sheet_name="Limit_Event_Summary", index=False)
        summary_df.to_excel(writer, sheet_name="All_Portfolio_Summary", index=False)
        q_group_summary.to_excel(writer, sheet_name="Q1_Q5_Summary", index=False)
        rebalance_limit_tc.to_excel(writer, sheet_name="Rebalance_Execution", index=False)
        nav_df.to_excel(writer, sheet_name="Daily_NAV_Returns", index=False)
        config.to_excel(writer, sheet_name="Config", index=False)
        methodology.to_excel(writer, sheet_name="Methodology", index=False)

    format_limit_tc_excel(LIMIT_TC_EXCEL)

    print(f"\nSaved limit + transaction cost Excel: {LIMIT_TC_EXCEL}")

    print("\nQ5_minus_Q1 limit + transaction cost summary:")
    show_cols = [
        "method",
        "horizon",
        "annual_return_gross",
        "annual_return_net",
        "annual_return_decay",
        "sharpe_no_rf_gross",
        "sharpe_no_rf_net",
        "sharpe_no_rf_decay",
        "max_drawdown_gross",
        "max_drawdown_net",
        "avg_turnover_on_rebalance_day",
        "avg_rebalance_day_transaction_cost",
    ]

    print(q5q1_summary[[c for c in show_cols if c in q5q1_summary.columns]].to_string(index=False))

    if not q5q1_compare_with_cost_only.empty:
        print("\nQ5_minus_Q1 comparison: cost-only vs limit + cost:")
        compare_cols = [
            "method",
            "horizon",
            "annual_return_net_cost_only",
            "annual_return_net_limit_tc",
            "annual_return_net_decay_from_limit",
            "sharpe_no_rf_net_cost_only",
            "sharpe_no_rf_net_limit_tc",
            "sharpe_no_rf_net_decay_from_limit",
            "max_drawdown_net_cost_only",
            "max_drawdown_net_limit_tc",
            "max_drawdown_net_decay_from_limit",
        ]
        print(
            q5q1_compare_with_cost_only[
                [c for c in compare_cols if c in q5q1_compare_with_cost_only.columns]
            ].to_string(index=False)
        )

    return {
        "daily_limit_tc_returns": daily_limit_tc_returns,
        "rebalance_limit_tc": rebalance_limit_tc,
        "nav": nav_df,
        "summary": summary_df,
        "q5q1_summary": q5q1_summary,
        "limit_event_summary": limit_event_summary,
        "q5q1_compare_with_cost_only": q5q1_compare_with_cost_only,
    }


# 7. Run

if __name__ == "__main__":
    limit_transaction_cost_result = run_limit_transaction_cost_tests()