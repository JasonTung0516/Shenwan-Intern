from __future__ import annotations

from pathlib import Path
from typing import Iterable, Sequence
import math
import re
import warnings

import numpy as np
import pandas as pd

try:
    from .config import *
except ImportError:  # allow running a file directly from this folder
    from config import *

try:
    from .core_base import *
    from .core_neutralization import *
except ImportError:
    from core_base import *
    from core_neutralization import *

# This script only reads existing CSV files and creates an organized Excel file.
# It does NOT rerun MA / EMA / RSI / MACD / IC calculations.

from pathlib import Path
import re



# =============================================================================
# 1. CONFIG: change paths here
# =============================================================================

OUTDIR = Path("/Users/zhaoshengdong/Desktop/ic_output_full_2005_2026")

SUMMARY_CSV = OUTDIR / "technical_factor_ic_summary_winsor_industry_size_neutral_2005_2026.csv"
DAILY_CSV = OUTDIR / "technical_factor_daily_ic_winsor_industry_size_neutral_2005_2026.csv"

EXCEL_PATH = OUTDIR / "technical_factor_ic_organized_winsor_industry_size_neutral_2005_2026.xlsx"

# Set False if the daily IC sheet is too large or not needed.
INCLUDE_DAILY_RAW = True

# 2. Sheet grouping rules
SHEET_ORDER = [
    "MA_5",
    "MA_10",
    "MA_20",
    "MA_60",
    "MA_Cross",
    "EMA_5",
    "EMA_10",
    "EMA_20",
    "EMA_60",
    "EMA_Cross",
    "RSI",
    "MACD",
    "Other",
]


def factor_group(factor: str) -> str:
    """
    Map factor names into technical indicator groups.
    Compatible with both my factor names and Codex-style factor names.
    """
    factor = str(factor)

    # My version: close_ma5_gap, close_ma10_gap, ...
    m = re.match(r"^close_ma(\d+)_gap$", factor)
    if m:
        return f"MA_{m.group(1)}"

    # Codex version: close_over_ma_5, close_over_ma_10, ...
    m = re.match(r"^close_over_ma_(\d+)$", factor)
    if m:
        return f"MA_{m.group(1)}"

    # MA cross factors: ma5_ma20_gap, ma10_ma20_gap
    if re.match(r"^ma\d+_ma\d+_gap$", factor):
        return "MA_Cross"

    # My version: close_ema5_gap, close_ema10_gap, ...
    m = re.match(r"^close_ema(\d+)_gap$", factor)
    if m:
        return f"EMA_{m.group(1)}"

    # EMA cross factors
    if factor in ["ema5_ema20_gap", "ema10_ema20_gap"]:
        return "EMA_Cross"

    # EMA12 / EMA26 is treated as MACD-related
    if factor in ["ema12_ema26_gap", "ema_12_over_26"]:
        return "MACD"

    # RSI
    if factor.startswith("rsi"):
        return "RSI"

    # MACD
    if factor.startswith("macd"):
        return "MACD"

    return "Other"


def factor_label(factor: str) -> str:
    """
    Human-readable factor labels.
    """
    factor = str(factor)

    label_map = {
        # MA
        "close_ma5_gap": "Close / MA5 - 1",
        "close_ma10_gap": "Close / MA10 - 1",
        "close_ma20_gap": "Close / MA20 - 1",
        "close_ma60_gap": "Close / MA60 - 1",
        "ma5_ma20_gap": "MA5 / MA20 - 1",
        "ma10_ma20_gap": "MA10 / MA20 - 1",

        # EMA
        "close_ema5_gap": "Close / EMA5 - 1",
        "close_ema10_gap": "Close / EMA10 - 1",
        "close_ema20_gap": "Close / EMA20 - 1",
        "close_ema60_gap": "Close / EMA60 - 1",
        "ema5_ema20_gap": "EMA5 / EMA20 - 1",
        "ema10_ema20_gap": "EMA10 / EMA20 - 1",

        # RSI
        "rsi14": "RSI 14",
        "rsi14_centered": "RSI 14 - 50",

        # MACD
        "macd_dif_pct": "MACD DIF / Price",
        "macd_dea_pct": "MACD DEA / Price",
        "macd_hist_pct": "MACD Hist / Price",
        "ema12_ema26_gap": "EMA12 / EMA26 - 1",

        # Codex-style names
        "close_over_ma_5": "Close / MA5 - 1",
        "close_over_ma_10": "Close / MA10 - 1",
        "close_over_ma_20": "Close / MA20 - 1",
        "close_over_ma_60": "Close / MA60 - 1",
        "ema_12_over_26": "EMA12 / EMA26 - 1",
        "rsi_14_centered": "RSI 14 - 50",
        "macd_dif": "MACD DIF",
        "macd_dea": "MACD DEA",
        "macd_hist": "MACD Hist",
    }

    return label_map.get(factor, factor)


# =============================================================================
# 3. Read CSV files
# =============================================================================

if not SUMMARY_CSV.exists():
    raise FileNotFoundError(f"Summary CSV not found: {SUMMARY_CSV}")

summary = pd.read_csv(SUMMARY_CSV)

daily = None
if INCLUDE_DAILY_RAW:
    if DAILY_CSV.exists():
        daily = pd.read_csv(DAILY_CSV)
    else:
        print(f"Warning: daily CSV not found, skipping Daily_IC_Raw sheet: {DAILY_CSV}")
        INCLUDE_DAILY_RAW = False


if daily is not None and (
    "ic_nw_icir" not in summary.columns
    or "rank_ic_nw_icir" not in summary.columns
):
    print("Computing Newey-West ICIR columns from daily IC CSV ...")

    nw_icir_df = compute_newey_west_icir_table(daily)

    summary = summary.merge(
        nw_icir_df,
        on=["factor", "horizon"],
        how="left",
    )

    if "icir" in summary.columns and "ic_nw_icir" in summary.columns:
        summary["icir_nw_diff"] = summary["ic_nw_icir"] - summary["icir"]
        summary["icir_nw_change_pct"] = np.where(
            summary["icir"].abs() > 1e-12,
            summary["ic_nw_icir"] / summary["icir"] - 1,
            np.nan,
        )

    if "rank_icir" in summary.columns and "rank_ic_nw_icir" in summary.columns:
        summary["rank_icir_nw_diff"] = summary["rank_ic_nw_icir"] - summary["rank_icir"]
        summary["rank_icir_nw_change_pct"] = np.where(
            summary["rank_icir"].abs() > 1e-12,
            summary["rank_ic_nw_icir"] / summary["rank_icir"] - 1,
            np.nan,
        )


# 4. Prepare summary table
summary_work = summary.copy()

# In case old summary does not have horizon, assume 5-day.
if "horizon" not in summary_work.columns:
    summary_work["horizon"] = 5

# Compatibility with old non-winsorized / Codex-style summary
if "rank_ic_mean" not in summary_work.columns and "ic_mean" in summary_work.columns:
    summary_work["rank_ic_mean"] = summary_work["ic_mean"]

if "rank_ic_std" not in summary_work.columns and "ic_std" in summary_work.columns:
    summary_work["rank_ic_std"] = summary_work["ic_std"]

if "rank_icir" not in summary_work.columns:
    if "ic_ir" in summary_work.columns:
        summary_work["rank_icir"] = summary_work["ic_ir"]
    elif {"rank_ic_mean", "rank_ic_std"}.issubset(summary_work.columns):
        summary_work["rank_icir"] = summary_work["rank_ic_mean"] / summary_work["rank_ic_std"]

if "rank_ic_positive_ratio" not in summary_work.columns and "positive_ratio" in summary_work.columns:
    summary_work["rank_ic_positive_ratio"] = summary_work["positive_ratio"]

if "sample_days" not in summary_work.columns and "count" in summary_work.columns:
    summary_work["sample_days"] = summary_work["count"]

if "icir" not in summary_work.columns and {"ic_mean", "ic_std"}.issubset(summary_work.columns):
    summary_work["icir"] = summary_work["ic_mean"] / summary_work["ic_std"]

if "winsor_icir" not in summary_work.columns and {"winsor_ic_mean", "winsor_ic_std"}.issubset(summary_work.columns):
    summary_work["winsor_icir"] = summary_work["winsor_ic_mean"] / summary_work["winsor_ic_std"]

# Add indicator_group and factor_label
for col in ["indicator_group", "factor_label"]:
    if col in summary_work.columns:
        summary_work = summary_work.drop(columns=[col])

summary_work.insert(1, "indicator_group", summary_work["factor"].map(factor_group))
summary_work.insert(2, "factor_label", summary_work["factor"].map(factor_label))

summary_work["indicator_group"] = pd.Categorical(
    summary_work["indicator_group"],
    categories=SHEET_ORDER,
    ordered=True,
)

# Convert date columns
for col in ["start_day", "end_day"]:
    if col in summary_work.columns:
        summary_work[col] = pd.to_datetime(summary_work[col], errors="coerce")

# Column order
preferred_cols = [
    "indicator_group",
    "factor",
    "factor_label",
    "horizon",

    # Raw Pearson IC
    "ic_mean",
    "ic_std",
    "icir",
    "ic_nw_std",
    "ic_nw_icir",
    "icir_nw_diff",
    "icir_nw_change_pct",
    "ic_positive_ratio",

    # Winsorized Pearson IC
    "winsor_ic_mean",
    "winsor_ic_std",
    "winsor_icir",
    "winsor_ic_positive_ratio",

    # Rank IC
    "rank_ic_mean",
    "rank_ic_std",
    "rank_icir",
    "rank_ic_nw_std",
    "rank_ic_nw_icir",
    "rank_icir_nw_diff",
    "rank_icir_nw_change_pct",
    "rank_ic_positive_ratio",

    # Raw IC vs Rank IC comparison
    "raw_rank_abs_gap_mean",
    "winsor_rank_abs_gap_mean",
    "gap_reduction_mean",
    "raw_rank_sign_match_ratio",
    "winsor_rank_sign_match_ratio",

    # Neutral IC
    "neutral_ic_mean",
    "neutral_ic_std",
    "neutral_icir",
    "neutral_ic_positive_ratio",

    "neutral_winsor_ic_mean",
    "neutral_winsor_ic_std",
    "neutral_winsor_icir",
    "neutral_winsor_ic_positive_ratio",

    "neutral_rank_ic_mean",
    "neutral_rank_ic_std",
    "neutral_rank_icir",
    "neutral_rank_ic_positive_ratio",

    "neutral_rank_abs_gap_mean",
    "neutral_winsor_rank_abs_gap_mean",
    "neutral_gap_reduction_mean",
    "neutral_rank_sign_match_ratio",
    "neutral_winsor_rank_sign_match_ratio",

    # Other diagnostics
    "abs_rank_ic_mean",
    "neutral_abs_rank_ic_mean",
    "sample_days",
    "neutral_sample_days",
    "avg_obs_per_day",
    "avg_neutral_obs_per_day",
    "nw_lag",
    "nw_n_obs",
    "start_day",
    "end_day",
    "t_stat",
]

summary_work = summary_work[
    [c for c in preferred_cols if c in summary_work.columns]
    + [c for c in summary_work.columns if c not in preferred_cols]
]

# Sorting
sort_col = "rank_icir" if "rank_icir" in summary_work.columns else "ic_mean"

summary_work = summary_work.sort_values(
    ["indicator_group", "horizon", sort_col],
    ascending=[True, True, False],
).reset_index(drop=True)


# 5. Create Overview sheet

overview = (
    summary_work
    .sort_values(["indicator_group", "horizon", sort_col], ascending=[True, True, False])
    .groupby(["indicator_group", "horizon"], group_keys=False, observed=False)
    .head(3)
    .reset_index(drop=True)
)

overview_cols = [
    "indicator_group",
    "factor",
    "factor_label",
    "horizon",

    "ic_mean",
    "rank_ic_mean",

    # raw ICIR vs NW ICIR
    "icir",
    "ic_nw_icir",
    "icir_nw_diff",
    "icir_nw_change_pct",

    # raw Rank ICIR vs NW Rank ICIR
    "rank_icir",
    "rank_ic_nw_icir",
    "rank_icir_nw_diff",
    "rank_icir_nw_change_pct",

    # winsorized ICIR, keep for reference
    "winsor_icir",

    "rank_ic_positive_ratio",
    "sample_days",
    "avg_obs_per_day",
    "nw_lag",
    "nw_n_obs",
    "start_day",
    "end_day",
]

overview = overview[[c for c in overview_cols if c in overview.columns]]


# 6. Create Winsor_Check sheet
winsor_check_cols = [
    "indicator_group",
    "factor",
    "factor_label",
    "horizon",

    "ic_mean",
    "winsor_ic_mean",
    "rank_ic_mean",

    "icir",
    "winsor_icir",
    "rank_icir",

    "raw_rank_abs_gap_mean",
    "winsor_rank_abs_gap_mean",
    "gap_reduction_mean",

    "raw_rank_sign_match_ratio",
    "winsor_rank_sign_match_ratio",

    "ic_positive_ratio",
    "winsor_ic_positive_ratio",
    "rank_ic_positive_ratio",

    "sample_days",
    "avg_obs_per_day",
    "start_day",
    "end_day",
]

# Create Neutral_Check sheet
neutral_check_cols = [
    "indicator_group",
    "factor",
    "factor_label",
    "horizon",

    "ic_mean",
    "winsor_ic_mean",
    "rank_ic_mean",

    "neutral_ic_mean",
    "neutral_winsor_ic_mean",
    "neutral_rank_ic_mean",

    "icir",
    "winsor_icir",
    "rank_icir",

    "neutral_icir",
    "neutral_winsor_icir",
    "neutral_rank_icir",

    "rank_ic_positive_ratio",
    "neutral_rank_ic_positive_ratio",

    "sample_days",
    "neutral_sample_days",
    "avg_obs_per_day",
    "avg_neutral_obs_per_day",

    "start_day",
    "end_day",
]

neutral_check = summary_work[
    [c for c in neutral_check_cols if c in summary_work.columns]
].copy()

if "neutral_rank_icir" in neutral_check.columns:
    neutral_check = neutral_check.sort_values(
        ["horizon", "neutral_rank_icir"],
        ascending=[True, False],
    ).reset_index(drop=True)

winsor_check = summary_work[[c for c in winsor_check_cols if c in summary_work.columns]].copy()

if "gap_reduction_mean" in winsor_check.columns:
    winsor_check = winsor_check.sort_values(
        ["horizon", "gap_reduction_mean"],
        ascending=[True, False],
    ).reset_index(drop=True)
else:
    winsor_check = winsor_check.sort_values(
        ["horizon", sort_col],
        ascending=[True, False],
    ).reset_index(drop=True)

# =============================================================================
# Create NW_ICIR_Check sheet
# raw ICIR vs Newey-West ICIR for ordinary IC and Rank IC
# =============================================================================

nw_check_cols = [
    "indicator_group",
    "factor",
    "factor_label",
    "horizon",

    # ordinary IC comparison
    "ic_mean",
    "ic_std",
    "icir",
    "ic_nw_std",
    "ic_nw_icir",
    "icir_nw_diff",
    "icir_nw_change_pct",

    # Rank IC comparison
    "rank_ic_mean",
    "rank_ic_std",
    "rank_icir",
    "rank_ic_nw_std",
    "rank_ic_nw_icir",
    "rank_icir_nw_diff",
    "rank_icir_nw_change_pct",

    "nw_lag",
    "nw_n_obs",
    "sample_days",
    "start_day",
    "end_day",
]

nw_check = summary_work[
    [c for c in nw_check_cols if c in summary_work.columns]
].copy()

if "rank_icir_nw_change_pct" in nw_check.columns:
    nw_check = nw_check.sort_values(
        ["horizon", "rank_icir_nw_change_pct"],
        ascending=[True, True],
    ).reset_index(drop=True)
elif "icir_nw_change_pct" in nw_check.columns:
    nw_check = nw_check.sort_values(
        ["horizon", "icir_nw_change_pct"],
        ascending=[True, True],
    ).reset_index(drop=True)
    
# =============================================================================
# Create NW_Horizon_Summary sheet
# One row per horizon: 1, 5, 10, 20
# =============================================================================

nw_horizon_summary = (
    summary_work
    .groupby("horizon", as_index=False)
    .agg(
        factor_count=("factor", "count"),

        avg_raw_icir=("icir", "mean"),
        avg_nw_icir=("ic_nw_icir", "mean"),
        avg_icir_nw_diff=("icir_nw_diff", "mean"),
        avg_icir_nw_change_pct=("icir_nw_change_pct", "mean"),

        avg_raw_rank_icir=("rank_icir", "mean"),
        avg_nw_rank_icir=("rank_ic_nw_icir", "mean"),
        avg_rank_icir_nw_diff=("rank_icir_nw_diff", "mean"),
        avg_rank_icir_nw_change_pct=("rank_icir_nw_change_pct", "mean"),

        median_raw_icir=("icir", "median"),
        median_nw_icir=("ic_nw_icir", "median"),
        median_raw_rank_icir=("rank_icir", "median"),
        median_nw_rank_icir=("rank_ic_nw_icir", "median"),
    )
    .sort_values("horizon")
    .reset_index(drop=True)
)



# 7. Create Methodology sheet
methodology = pd.DataFrame(
    [
        {
            "Item": "Raw Pearson IC",
            "Description": "Linear correlation between factor values and forward returns.",
        },
        {
            "Item": "Winsorized Pearson IC",
            "Description": "Pearson IC after replacing the most extreme 1% and 99% values by the corresponding quantiles within each daily cross-section.",
        },
        {
            "Item": "Rank IC",
            "Description": "Spearman rank correlation between factor ranks and forward return ranks.",
        },
        {
            "Item": "raw_rank_abs_gap_mean",
            "Description": "Average absolute difference between raw Pearson IC and Rank IC.",
        },
        {
            "Item": "winsor_rank_abs_gap_mean",
            "Description": "Average absolute difference between winsorized Pearson IC and Rank IC.",
        },
        {
            "Item": "gap_reduction_mean",
            "Description": "raw_rank_abs_gap_mean minus winsor_rank_abs_gap_mean. Positive value means winsorized IC is closer to Rank IC.",
        },
        {
            "Item": "raw_rank_sign_match_ratio",
            "Description": "Share of days where raw Pearson IC and Rank IC have the same sign.",
        },
        {
            "Item": "winsor_rank_sign_match_ratio",
            "Description": "Share of days where winsorized Pearson IC and Rank IC have the same sign.",
        },
    ]
)


# 8. Prepare daily IC sheet

daily_work = None

if INCLUDE_DAILY_RAW and daily is not None:
    daily_work = daily.copy()

    # My version: long format, has factor column
    if "factor" in daily_work.columns:
        for col in ["indicator_group", "factor_label"]:
            if col in daily_work.columns:
                daily_work = daily_work.drop(columns=[col])

        daily_work.insert(1, "indicator_group", daily_work["factor"].map(factor_group))
        daily_work.insert(2, "factor_label", daily_work["factor"].map(factor_label))

        if "trade_day" in daily_work.columns:
            daily_work["trade_day"] = pd.to_datetime(daily_work["trade_day"], errors="coerce")

        daily_preferred_cols = [
            "trade_day",
            "indicator_group",
            "factor",
            "factor_label",
            "horizon",
            "ic",
            "winsor_ic",
            "rank_ic",
            "n_obs",
        ]

        daily_work = daily_work[
            [c for c in daily_preferred_cols if c in daily_work.columns]
            + [c for c in daily_work.columns if c not in daily_preferred_cols]
        ]

        sort_cols = [c for c in ["trade_day", "indicator_group", "horizon", "factor"] if c in daily_work.columns]
        daily_work = daily_work.sort_values(sort_cols).reset_index(drop=True)

    # Codex version: wide format, each factor is one column
    else:
        first_col = daily_work.columns[0]

        if str(first_col).startswith("Unnamed"):
            daily_work = daily_work.rename(columns={first_col: "trade_day"})

        if "trade_day" not in daily_work.columns and "date" in daily_work.columns:
            daily_work = daily_work.rename(columns={"date": "trade_day"})

        if "trade_day" not in daily_work.columns:
            raise ValueError("Cannot find trade_day/date column in daily CSV.")

        daily_work["trade_day"] = pd.to_datetime(daily_work["trade_day"], errors="coerce")

        factor_cols = [c for c in daily_work.columns if c != "trade_day"]

        daily_work = daily_work.melt(
            id_vars="trade_day",
            value_vars=factor_cols,
            var_name="factor",
            value_name="rank_ic",
        )

        daily_work["horizon"] = 5
        daily_work.insert(1, "indicator_group", daily_work["factor"].map(factor_group))
        daily_work.insert(2, "factor_label", daily_work["factor"].map(factor_label))

        daily_work = daily_work[
            ["trade_day", "indicator_group", "factor", "factor_label", "horizon", "rank_ic"]
        ].sort_values(
            ["trade_day", "indicator_group", "factor"]
        ).reset_index(drop=True)


# 9. Write Excel
def safe_sheet_name(name: str) -> str:
    """
    Excel sheet name max length is 31.
    Certain characters are invalid.
    """
    name = str(name)
    name = re.sub(r"[\[\]\:\*\?\/\\]", "_", name)
    return name[:31]


def write_df(writer, df: pd.DataFrame, sheet_name: str):
    """
    Write DataFrame to Excel.
    If rows exceed Excel limit, split into multiple sheets.
    """
    sheet_name = safe_sheet_name(sheet_name)

    max_excel_rows = 1_048_576
    max_data_rows = max_excel_rows - 1

    if len(df) <= max_data_rows:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        for i, start in enumerate(range(0, len(df), max_data_rows), start=1):
            part = df.iloc[start:start + max_data_rows]
            part_sheet_name = safe_sheet_name(f"{sheet_name}_{i}")
            part.to_excel(writer, sheet_name=part_sheet_name, index=False)


with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl") as writer:
    write_df(writer, methodology, "Methodology")
    write_df(writer, overview, "Overview")
    write_df(writer, winsor_check, "Winsor_Check")
    write_df(writer, neutral_check, "Neutral_Check")
    write_df(writer, nw_check, "NW_Check")
    write_df(writer, summary_work, "All_Summary")

    for group in SHEET_ORDER:
        part = summary_work[summary_work["indicator_group"].astype(str) == group].copy()
        if not part.empty:
            write_df(writer, part, group)

    if INCLUDE_DAILY_RAW and daily_work is not None:
        write_df(writer, daily_work, "Daily_IC_Raw")


# 10. Format Excel
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = load_workbook(EXCEL_PATH)

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)

good_fill = PatternFill("solid", fgColor="E2F0D9")
bad_fill = PatternFill("solid", fgColor="FCE4D6")

numeric_4_cols = {
    "ic_mean",
    "ic_std",
    "icir",
    "ic_nw_std",
    "ic_nw_icir",
    "icir_nw_diff",

    "winsor_ic_mean",
    "winsor_ic_std",
    "winsor_icir",

    "rank_ic_mean",
    "rank_ic_std",
    "rank_icir",
    "rank_ic_nw_std",
    "rank_ic_nw_icir",
    "rank_icir_nw_diff",

    "neutral_ic_mean",
    "neutral_ic_std",
    "neutral_icir",

    "neutral_winsor_ic_mean",
    "neutral_winsor_ic_std",
    "neutral_winsor_icir",

    "neutral_rank_ic_mean",
    "neutral_rank_ic_std",
    "neutral_rank_icir",

    "raw_rank_abs_gap_mean",
    "winsor_rank_abs_gap_mean",
    "gap_reduction_mean",

    "neutral_rank_abs_gap_mean",
    "neutral_winsor_rank_abs_gap_mean",
    "neutral_gap_reduction_mean",

    "abs_rank_ic_mean",
    "neutral_abs_rank_ic_mean",
    "avg_obs_per_day",
    "avg_neutral_obs_per_day",

    "ic",
    "winsor_ic",
    "rank_ic",
    "neutral_ic",
    "neutral_winsor_ic",
    "neutral_rank_ic",

    "avg_raw_icir",
    "avg_nw_icir",
    "avg_icir_nw_diff",
    "avg_raw_rank_icir",
    "avg_nw_rank_icir",
    "avg_rank_icir_nw_diff",
    "median_raw_icir",
    "median_nw_icir",
    "median_raw_rank_icir",
    "median_nw_rank_icir",
}

percent_cols = {
    "ic_positive_ratio",
    "winsor_ic_positive_ratio",
    "rank_ic_positive_ratio",
    "positive_ratio",
    "raw_rank_sign_match_ratio",
    "winsor_rank_sign_match_ratio",

    "neutral_ic_positive_ratio",
    "neutral_winsor_ic_positive_ratio",
    "neutral_rank_ic_positive_ratio",
    "neutral_rank_sign_match_ratio",
    "neutral_winsor_rank_sign_match_ratio",

    "icir_nw_change_pct",
    "rank_icir_nw_change_pct",
    "avg_icir_nw_change_pct",
    "avg_rank_icir_nw_change_pct",
}

integer_cols = {
    "sample_days",
    "neutral_sample_days",
    "n_obs",
    "neutral_n_obs",
    "nw_lag",
    "nw_n_obs",
    "factor_count",
    "count",
    "horizon",
}

date_cols = {
    "trade_day",
    "start_day",
    "end_day",
}

for ws in wb.worksheets:
    if ws.max_row < 1 or ws.max_column < 1:
        continue

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Header style
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Map header to column index
    header_to_col = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        header_to_col[str(header)] = col_idx

    # Column widths
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0

        # Limit scanning to first 3000 rows for speed
        scan_rows = min(ws.max_row, 3000)
        for row_idx in range(1, scan_rows + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))

        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 30)

    # Number formats by column
    for header, col_idx in header_to_col.items():
        if header in numeric_4_cols:
            fmt = "0.0000"
        elif header in percent_cols:
            fmt = "0.00%"
        elif header in integer_cols:
            fmt = "0"
        elif header in date_cols:
            fmt = "yyyy-mm-dd"
        else:
            fmt = None

        if fmt is not None:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = fmt

    # Highlight gap_reduction_mean:
    # positive = winsorized IC closer to Rank IC
    if "gap_reduction_mean" in header_to_col:
        col_idx = header_to_col["gap_reduction_mean"]
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                value = float(cell.value)
            except (TypeError, ValueError):
                continue

            if value > 0:
                cell.fill = good_fill
            elif value < 0:
                cell.fill = bad_fill

wb.save(EXCEL_PATH)

print(f"Saved organized Excel: {EXCEL_PATH}")
