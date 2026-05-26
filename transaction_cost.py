from __future__ import annotations

from pathlib import Path
from typing import Iterable, Sequence
import math
import re
import warnings

import numpy as np
import pandas as pd

try:
    from .config import *
except ImportError:  # allow running a file directly from this folder
    from config import *

try:
    from .core_base import *
    from .core_neutralization import *
    from .portfolio_five_group import *
    from .portfolio_industry_neutral import *
except ImportError:
    from core_base import *
    from core_neutralization import *
    from portfolio_five_group import *
    from portfolio_industry_neutral import *

# Requirement:
# Buy cost = 0.0005
# Sell cost = 0.001
# Cost is calculated by actual rebalance turnover:
#   buy_turnover  = sum(max(new_weight - old_weight, 0))
#   sell_turnover = sum(max(old_weight - new_weight, 0))
#   cost = buy_turnover * buy_cost + sell_turnover * sell_cost
#
# This chunk compares gross vs net performance after transaction costs.
# It supports:
#   1. Original equal-weight five-group test
#   2. Industry-neutral five-group test
#
# It outputs one standalone Excel.

# =============================================================================
# CONFIG
# =============================================================================

TC_OUTDIR = OUTDIR / "transaction_cost_test"
TC_OUTDIR.mkdir(parents=True, exist_ok=True)

TC_EXCEL = TC_OUTDIR / "five_group_transaction_cost_comparison.xlsx"

BUY_COST_RATE = 0.0005
SELL_COST_RATE = 0.001

TC_PORTFOLIOS = [f"Q{i}" for i in range(1, N_GROUPS + 1)] + ["Q5_minus_Q1"]

# Run both versions.
RUN_ORIGINAL_EQUAL_WEIGHT_TC = True
RUN_INDUSTRY_NEUTRAL_TC = True


# =============================================================================
# 1. Basic helpers
# =============================================================================

def _clean_weight_series(w: pd.Series) -> pd.Series:
    """
    Clean portfolio weight Series.

    Index = stock code
    Value = portfolio weight

    Positive weight = long position.
    Negative weight = short position.
    """
    if w is None or len(w) == 0:
        return pd.Series(dtype=float)

    out = pd.Series(w, dtype=float).copy()
    out.index = out.index.astype(str)

    out = out.replace([np.inf, -np.inf], np.nan).dropna()
    out = out[out.abs() > 1e-15]

    return out


def _normalize_long_only_weights(w: pd.Series) -> pd.Series:
    """
    Normalize long-only weights to sum to 1.
    """
    w = _clean_weight_series(w)
    w = w[w > 0]

    s = w.sum()

    if not np.isfinite(s) or s <= 0:
        return pd.Series(dtype=float)

    return w / s


def _normalize_signed_long_short_weights(w: pd.Series) -> pd.Series:
    """
    Normalize signed long-short weights.

    If both long and short sides exist:
        long side sums to +1
        short side sums to -1

    If only long side exists:
        long side sums to +1

    If only short side exists:
        short side sums to -1
    """
    w = _clean_weight_series(w)

    if w.empty:
        return w

    pos = w[w > 0].copy()
    neg = w[w < 0].copy()

    pieces = []

    if not pos.empty:
        pos_sum = pos.sum()
        if np.isfinite(pos_sum) and pos_sum > 0:
            pieces.append(pos / pos_sum)

    if not neg.empty:
        neg_abs_sum = -neg.sum()
        if np.isfinite(neg_abs_sum) and neg_abs_sum > 0:
            pieces.append(neg / neg_abs_sum)

    if not pieces:
        return pd.Series(dtype=float)

    out = pd.concat(pieces)
    out = out.groupby(out.index).sum()
    out = out[out.abs() > 1e-15]

    return out


def calc_rebalance_turnover_and_cost(
    old_weights: pd.Series,
    new_weights: pd.Series,
    buy_cost_rate: float = BUY_COST_RATE,
    sell_cost_rate: float = SELL_COST_RATE,
) -> dict:
    """
    Calculate actual trading proportion and transaction cost.

    For signed weights:
        positive delta means buy
        negative delta means sell

    This works for:
        long-only Q1-Q5 portfolios
        signed Q5_minus_Q1 long-short portfolio
    """
    old_w = _clean_weight_series(old_weights)
    new_w = _clean_weight_series(new_weights)

    all_codes = old_w.index.union(new_w.index)

    old_aligned = old_w.reindex(all_codes).fillna(0.0)
    new_aligned = new_w.reindex(all_codes).fillna(0.0)

    delta = new_aligned - old_aligned

    buy_turnover = float(delta.clip(lower=0).sum())
    sell_turnover = float((-delta.clip(upper=0)).sum())

    turnover = buy_turnover + sell_turnover

    transaction_cost = (
        buy_turnover * buy_cost_rate
        + sell_turnover * sell_cost_rate
    )

    return {
        "buy_turnover": buy_turnover,
        "sell_turnover": sell_turnover,
        "turnover": turnover,
        "transaction_cost": float(transaction_cost),
    }


def _handle_missing_returns_preserve_exposure(
    weights: pd.Series,
    returns: pd.Series,
) -> tuple[pd.Series, pd.Series]:
    """
    Align current weights and stock returns.

    If some stocks have missing returns, drop them and rescale the remaining
    same-side weights to preserve the current long and short exposure.
    """
    w = _clean_weight_series(weights)

    if w.empty:
        return pd.Series(dtype=float), pd.Series(dtype=float)

    r = pd.to_numeric(returns.reindex(w.index), errors="coerce")
    r = r.replace([np.inf, -np.inf], np.nan)

    finite = r.notna()

    if finite.sum() == 0:
        return pd.Series(dtype=float), pd.Series(dtype=float)

    w_used = w.loc[finite].copy()
    r_used = r.loc[finite].copy()

    # Preserve current long exposure.
    old_pos_sum = w[w > 0].sum()
    used_pos_sum = w_used[w_used > 0].sum()

    if np.isfinite(old_pos_sum) and old_pos_sum > 0 and np.isfinite(used_pos_sum) and used_pos_sum > 0:
        w_used.loc[w_used > 0] *= old_pos_sum / used_pos_sum

    # Preserve current short exposure.
    old_neg_abs_sum = -w[w < 0].sum()
    used_neg_abs_sum = -w_used[w_used < 0].sum()

    if (
        np.isfinite(old_neg_abs_sum)
        and old_neg_abs_sum > 0
        and np.isfinite(used_neg_abs_sum)
        and used_neg_abs_sum > 0
    ):
        w_used.loc[w_used < 0] *= old_neg_abs_sum / used_neg_abs_sum

    return w_used, r_used


def calc_portfolio_return_and_drift_weights(
    current_weights: pd.Series,
    ret_series: pd.Series,
) -> tuple[float, pd.Series, int]:
    """
    Calculate portfolio gross return and drifted weights after one day.

    Portfolio return:
        r_p = sum_i weight_i * r_i

    New drifted weight:
        weight_i_new = weight_i * (1 + r_i) / (1 + r_p)
    """
    w_used, r_used = _handle_missing_returns_preserve_exposure(
        weights=current_weights,
        returns=ret_series,
    )

    if w_used.empty or r_used.empty:
        return np.nan, current_weights.copy(), 0

    gross_ret = float((w_used * r_used).sum())

    if not np.isfinite(gross_ret) or not np.isfinite(1.0 + gross_ret) or (1.0 + gross_ret) <= 0:
        return gross_ret, w_used.copy(), int(len(w_used))

    drifted_weights = w_used * (1.0 + r_used) / (1.0 + gross_ret)
    drifted_weights = _clean_weight_series(drifted_weights)

    return gross_ret, drifted_weights, int(len(w_used))


# =============================================================================
# 2. Target weight builders
# =============================================================================

def _add_q5_minus_q1_target_weights(targets: dict[str, pd.Series]) -> dict[str, pd.Series]:
    """
    Add signed long-short target:

        Q5_minus_Q1 = long Q5 and short Q1

    Q5 side sums to +1.
    Q1 side sums to -1.
    """
    out = dict(targets)

    if "Q5" in targets and "Q1" in targets:
        spread_w = targets["Q5"].add(-targets["Q1"], fill_value=0.0)
        out["Q5_minus_Q1"] = _normalize_signed_long_short_weights(spread_w)

    return out


def build_original_equal_weight_targets(
    signal_day_df: pd.DataFrame,
    signal_col: str,
) -> dict[str, pd.Series]:
    """
    Build target stock weights for the original five-group test.

    Q1-Q5:
        equal-weight within each quintile

    Q5_minus_Q1:
        long Q5, short Q1
    """
    group_df = assign_quintile_groups_for_day(
        signal_day_df=signal_day_df,
        signal_col=signal_col,
        n_groups=N_GROUPS,
        min_stocks=MIN_STOCKS_TO_GROUP,
    )

    if group_df.empty:
        return {}

    targets = {}

    for group in [f"Q{i}" for i in range(1, N_GROUPS + 1)]:
        codes = (
            group_df.loc[group_df["group"] == group, "code"]
            .astype(str)
            .drop_duplicates()
            .tolist()
        )

        if len(codes) == 0:
            continue

        w = pd.Series(1.0 / len(codes), index=codes, dtype=float)
        targets[group] = _normalize_long_only_weights(w)

    targets = _add_q5_minus_q1_target_weights(targets)

    return targets


def build_industry_neutral_targets(
    signal_day_df: pd.DataFrame,
    signal_col: str,
) -> dict[str, pd.Series]:
    """
    Build target stock weights for the industry-neutral five-group test.

    Q1-Q5:
        within industry equal-weight
        across industries use same target industry weights

    Q5_minus_Q1:
        long industry-neutral Q5, short industry-neutral Q1
    """
    group_df = assign_quintile_groups_for_day_industry_neutral(
        signal_day_df=signal_day_df,
        signal_col=signal_col,
        n_groups=N_GROUPS,
        min_stocks=MIN_STOCKS_TO_GROUP,
    )

    if group_df.empty:
        return {}

    required_groups = [f"Q{i}" for i in range(1, N_GROUPS + 1)]

    if not set(required_groups).issubset(set(group_df["group"].unique())):
        return {}

    target_industry_weight = get_target_industry_weights_for_day(
        signal_day_df=signal_day_df,
        signal_col=signal_col,
        method=INDUSTRY_WEIGHT_METHOD,
    )

    if target_industry_weight.empty:
        return {}

    if USE_COMMON_INDUSTRIES_ONLY:
        common_industries = set(target_industry_weight.index.astype(str))

        for group in required_groups:
            inds = set(group_df.loc[group_df["group"] == group, "industry"].astype(str))
            common_industries = common_industries.intersection(inds)

        common_industries = sorted(common_industries)

        if len(common_industries) < MIN_INDUSTRIES_FOR_REBALANCE:
            return {}

        target_industry_weight = target_industry_weight.loc[common_industries]
        target_industry_weight = target_industry_weight / target_industry_weight.sum()

        group_df = group_df[group_df["industry"].isin(common_industries)].copy()

    targets = {}

    for group in required_groups:
        gtmp = group_df[group_df["group"] == group].copy()

        pieces = []

        for industry, industry_w in target_industry_weight.items():
            codes = (
                gtmp.loc[gtmp["industry"] == industry, "code"]
                .astype(str)
                .drop_duplicates()
                .tolist()
            )

            if len(codes) == 0:
                continue

            one_industry_weight = pd.Series(
                float(industry_w) / len(codes),
                index=codes,
                dtype=float,
            )

            pieces.append(one_industry_weight)

        if pieces:
            w = pd.concat(pieces)
            w = w.groupby(w.index).sum()
            targets[group] = _normalize_long_only_weights(w)

    targets = _add_q5_minus_q1_target_weights(targets)

    return targets


# =============================================================================
# 3. Backtest with transaction cost
# =============================================================================

def get_signal_panel_for_transaction_cost(
    method_name: str,
) -> tuple[pd.DataFrame, pd.DataFrame | None]:
    """
    Reuse existing signal panels if available.

    For original equal-weight:
        use five_group_result["signal_panel"] if available.

    For industry-neutral:
        use industry_neutral_five_group_result["signal_panel"] if available.
        Otherwise add industry info to the original signal panel.
    """
    method_name = str(method_name)

    if method_name == "Original equal-weight":
        if (
            "five_group_result" in globals()
            and isinstance(five_group_result, dict)
            and "signal_panel" in five_group_result
        ):
            return five_group_result["signal_panel"].copy(), None

        signal_panel, weights = build_group_test_signal_panel()
        return signal_panel, weights

    if method_name == "Industry-neutral weight":
        if (
            "industry_neutral_five_group_result" in globals()
            and isinstance(industry_neutral_five_group_result, dict)
            and "signal_panel" in industry_neutral_five_group_result
        ):
            return industry_neutral_five_group_result["signal_panel"].copy(), None

        if (
            "five_group_result" in globals()
            and isinstance(five_group_result, dict)
            and "signal_panel" in five_group_result
        ):
            base_signal_panel = five_group_result["signal_panel"].copy()
            signal_panel_industry, industry_cols = add_industry_to_group_signal_panel(base_signal_panel)
            return signal_panel_industry, None

        base_signal_panel, weights = build_group_test_signal_panel()
        signal_panel_industry, industry_cols = add_industry_to_group_signal_panel(base_signal_panel)
        return signal_panel_industry, weights

    raise ValueError(f"Unknown method_name: {method_name}")


def run_one_transaction_cost_backtest(
    signal_panel: pd.DataFrame,
    signal_col: str,
    method_name: str,
    target_builder,
    rebalance_periods: list[int] = GROUP_REBALANCE_PERIODS,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Run one transaction-cost-aware backtest.

    Output:
        daily_tc_returns:
            daily gross/net returns, transaction cost, turnover

        rebalance_tc:
            rebalance-level turnover and transaction cost
    """
    panel = signal_panel.copy()
    panel["trade_day"] = pd.to_datetime(panel["trade_day"]).dt.normalize()
    panel["code"] = panel["code"].astype(str)

    panel = panel.sort_values(["code", "trade_day"]).reset_index(drop=True)

    panel["ret_vwap_1d"] = (
        panel.groupby("code", sort=False)["vwap_price"].pct_change()
    )
    panel["ret_vwap_1d"] = panel["ret_vwap_1d"].replace([np.inf, -np.inf], np.nan)

    trading_days = sorted(panel["trade_day"].dropna().unique())

    ret_by_day = {}

    for day, sub in panel[["code", "trade_day", "ret_vwap_1d"]].groupby("trade_day"):
        temp = sub.copy()
        temp["code"] = temp["code"].astype(str)
        ret_by_day[day] = temp.set_index("code")["ret_vwap_1d"]

    signal_cols = ["code", "trade_day", "volume", signal_col]

    # Needed only by industry-neutral target builder.
    for c in ["industry", "industry_valid_for_group", "mkt_cap"]:
        if c in panel.columns and c not in signal_cols:
            signal_cols.append(c)

    signal_by_day = {
        day: sub[signal_cols].copy()
        for day, sub in panel[signal_cols].groupby("trade_day")
    }

    daily_records = []
    rebalance_records = []

    for h in rebalance_periods:
        print(f"\nRunning transaction-cost backtest: {method_name}, rebalance_period={h}d")

        current_weights = {
            portfolio: pd.Series(dtype=float)
            for portfolio in TC_PORTFOLIOS
        }

        valid_start_end = len(trading_days) - h

        for start_i in range(0, valid_start_end, h):
            signal_day = trading_days[start_i]

            if start_i + h + 1 >= len(trading_days):
                break

            entry_day = trading_days[start_i + 1]
            exit_day = trading_days[start_i + h + 1]

            signal_day_df = signal_by_day.get(signal_day)

            if signal_day_df is None or signal_day_df.empty:
                continue

            target_weights = target_builder(
                signal_day_df=signal_day_df,
                signal_col=signal_col,
            )

            if not target_weights:
                continue

            # Require all Q1-Q5 and Q5_minus_Q1.
            if not set(TC_PORTFOLIOS).issubset(set(target_weights.keys())):
                continue

            rebalance_cost = {}

            for portfolio in TC_PORTFOLIOS:
                old_w = current_weights.get(portfolio, pd.Series(dtype=float))
                new_w = target_weights[portfolio].copy()

                cost_info = calc_rebalance_turnover_and_cost(
                    old_weights=old_w,
                    new_weights=new_w,
                    buy_cost_rate=BUY_COST_RATE,
                    sell_cost_rate=SELL_COST_RATE,
                )

                rebalance_cost[portfolio] = cost_info

                current_weights[portfolio] = new_w.copy()

                pos_exposure = float(new_w[new_w > 0].sum())
                neg_exposure = float(new_w[new_w < 0].sum())
                gross_exposure = float(new_w.abs().sum())
                net_exposure = float(new_w.sum())

                rebalance_records.append(
                    {
                        "method": method_name,
                        "horizon": h,
                        "signal_day": signal_day,
                        "entry_day": entry_day,
                        "exit_day": exit_day,
                        "portfolio": portfolio,
                        "n_names": int(new_w.shape[0]),
                        "long_names": int((new_w > 0).sum()),
                        "short_names": int((new_w < 0).sum()),
                        "long_exposure": pos_exposure,
                        "short_exposure": neg_exposure,
                        "gross_exposure": gross_exposure,
                        "net_exposure": net_exposure,
                        "buy_turnover": cost_info["buy_turnover"],
                        "sell_turnover": cost_info["sell_turnover"],
                        "turnover": cost_info["turnover"],
                        "transaction_cost": cost_info["transaction_cost"],
                        "buy_cost_rate": BUY_COST_RATE,
                        "sell_cost_rate": SELL_COST_RATE,
                    }
                )

            # Daily holding returns.
            # Cost is charged on the first holding return day after entry.
            for interval_i in range(start_i + 1, start_i + h + 1):
                return_day = trading_days[interval_i + 1]
                ret_series = ret_by_day.get(return_day)

                if ret_series is None:
                    continue

                is_first_holding_day = interval_i == start_i + 1

                for portfolio in TC_PORTFOLIOS:
                    gross_ret, drifted_weights, n_ret_obs = calc_portfolio_return_and_drift_weights(
                        current_weights=current_weights[portfolio],
                        ret_series=ret_series,
                    )

                    tc = (
                        rebalance_cost[portfolio]["transaction_cost"]
                        if is_first_holding_day
                        else 0.0
                    )

                    buy_turnover = (
                        rebalance_cost[portfolio]["buy_turnover"]
                        if is_first_holding_day
                        else 0.0
                    )

                    sell_turnover = (
                        rebalance_cost[portfolio]["sell_turnover"]
                        if is_first_holding_day
                        else 0.0
                    )

                    turnover = (
                        rebalance_cost[portfolio]["turnover"]
                        if is_first_holding_day
                        else 0.0
                    )

                    if pd.notna(gross_ret):
                        # Cost is paid before the holding-period return.
                        # For small transaction costs, this is close to gross_ret - tc.
                        net_ret = (1.0 - tc) * (1.0 + gross_ret) - 1.0
                    else:
                        net_ret = np.nan

                    daily_records.append(
                        {
                            "method": method_name,
                            "horizon": h,
                            "trade_day": return_day,
                            "signal_day": signal_day,
                            "entry_day": entry_day,
                            "exit_day": exit_day,
                            "portfolio": portfolio,
                            "gross_ret": gross_ret,
                            "transaction_cost": tc,
                            "net_ret": net_ret,
                            "buy_turnover": buy_turnover,
                            "sell_turnover": sell_turnover,
                            "turnover": turnover,
                            "n_ret_obs": n_ret_obs,
                            "is_rebalance_cost_day": is_first_holding_day,
                        }
                    )

                    if pd.notna(gross_ret):
                        current_weights[portfolio] = drifted_weights.copy()

    daily_tc_returns = pd.DataFrame(daily_records)
    rebalance_tc = pd.DataFrame(rebalance_records)

    return daily_tc_returns, rebalance_tc


# =============================================================================
# 4. Performance summary
# =============================================================================

def calc_max_drawdown_from_nav(nav: pd.Series) -> float:
    nav = pd.Series(nav).dropna()

    if nav.empty:
        return np.nan

    running_max = nav.cummax()
    dd = nav / running_max - 1.0

    return float(dd.min())


def calc_perf_metrics_from_ret(ret: pd.Series) -> dict:
    ret = pd.Series(ret).replace([np.inf, -np.inf], np.nan).dropna()

    if ret.empty:
        return {
            "n_days": 0,
            "total_return": np.nan,
            "annual_return": np.nan,
            "annual_vol": np.nan,
            "sharpe_no_rf": np.nan,
            "max_drawdown": np.nan,
            "avg_daily_return": np.nan,
            "daily_win_rate": np.nan,
        }

    nav = (1.0 + ret).cumprod()
    n_days = int(ret.shape[0])

    total_return = float(nav.iloc[-1] - 1.0)
    annual_return = float(nav.iloc[-1] ** (252.0 / n_days) - 1.0)
    annual_vol = float(ret.std(ddof=1) * np.sqrt(252.0))
    sharpe_no_rf = annual_return / annual_vol if annual_vol > 0 else np.nan
    max_drawdown = calc_max_drawdown_from_nav(nav)

    return {
        "n_days": n_days,
        "total_return": total_return,
        "annual_return": annual_return,
        "annual_vol": annual_vol,
        "sharpe_no_rf": sharpe_no_rf,
        "max_drawdown": max_drawdown,
        "avg_daily_return": float(ret.mean()),
        "daily_win_rate": float((ret > 0).mean()),
    }


def build_transaction_cost_nav_and_summary(
    daily_tc_returns: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Build NAV and performance summary from daily gross/net returns.
    """
    nav_records = []
    summary_records = []

    metric_names = [
        "total_return",
        "annual_return",
        "annual_vol",
        "sharpe_no_rf",
        "max_drawdown",
        "avg_daily_return",
        "daily_win_rate",
    ]

    group_cols = ["method", "horizon", "portfolio"]

    for (method, h, portfolio), sub in daily_tc_returns.groupby(group_cols):
        sub = sub.sort_values("trade_day").copy()

        gross_ret = sub["gross_ret"].replace([np.inf, -np.inf], np.nan)
        net_ret = sub["net_ret"].replace([np.inf, -np.inf], np.nan)

        gross_nav = (1.0 + gross_ret.fillna(0.0)).cumprod()
        net_nav = (1.0 + net_ret.fillna(0.0)).cumprod()

        temp_nav = pd.DataFrame(
            {
                "method": method,
                "horizon": h,
                "portfolio": portfolio,
                "trade_day": sub["trade_day"].values,
                "gross_ret": gross_ret.values,
                "net_ret": net_ret.values,
                "transaction_cost": sub["transaction_cost"].values,
                "buy_turnover": sub["buy_turnover"].values,
                "sell_turnover": sub["sell_turnover"].values,
                "turnover": sub["turnover"].values,
                "gross_nav": gross_nav.values,
                "net_nav": net_nav.values,
            }
        )

        nav_records.append(temp_nav)

        gross_metrics = calc_perf_metrics_from_ret(gross_ret)
        net_metrics = calc_perf_metrics_from_ret(net_ret)

        row = {
            "method": method,
            "horizon": h,
            "portfolio": portfolio,
            "n_days": net_metrics["n_days"],
            "total_transaction_cost": float(sub["transaction_cost"].sum()),
            "avg_daily_transaction_cost": float(sub["transaction_cost"].mean()),
            "avg_rebalance_day_transaction_cost": float(
                sub.loc[sub["is_rebalance_cost_day"], "transaction_cost"].mean()
            ),
            "avg_buy_turnover_on_rebalance_day": float(
                sub.loc[sub["is_rebalance_cost_day"], "buy_turnover"].mean()
            ),
            "avg_sell_turnover_on_rebalance_day": float(
                sub.loc[sub["is_rebalance_cost_day"], "sell_turnover"].mean()
            ),
            "avg_turnover_on_rebalance_day": float(
                sub.loc[sub["is_rebalance_cost_day"], "turnover"].mean()
            ),
        }

        for m in metric_names:
            row[f"{m}_gross"] = gross_metrics[m]
            row[f"{m}_net"] = net_metrics[m]
            row[f"{m}_decay"] = net_metrics[m] - gross_metrics[m]

        summary_records.append(row)

    nav_df = pd.concat(nav_records, ignore_index=True) if nav_records else pd.DataFrame()
    summary_df = pd.DataFrame(summary_records)

    if not summary_df.empty:
        summary_df = summary_df.sort_values(
            ["method", "horizon", "portfolio"],
            ascending=[True, True, True],
        ).reset_index(drop=True)

    return nav_df, summary_df


# =============================================================================
# 5. Run both transaction-cost tests
# =============================================================================

def run_transaction_cost_tests():
    all_daily = []
    all_rebalance = []

    if SIGNAL_MODE == "weighted":
        signal_col = GROUP_SIGNAL_WEIGHTED_COL
    elif SIGNAL_MODE == "equal":
        signal_col = GROUP_SIGNAL_EQUAL_COL
    else:
        raise ValueError("SIGNAL_MODE must be 'weighted' or 'equal'.")

    if RUN_ORIGINAL_EQUAL_WEIGHT_TC:
        original_signal_panel, _ = get_signal_panel_for_transaction_cost(
            method_name="Original equal-weight"
        )

        original_daily, original_rebalance = run_one_transaction_cost_backtest(
            signal_panel=original_signal_panel,
            signal_col=signal_col,
            method_name="Original equal-weight",
            target_builder=build_original_equal_weight_targets,
            rebalance_periods=GROUP_REBALANCE_PERIODS,
        )

        all_daily.append(original_daily)
        all_rebalance.append(original_rebalance)

    if RUN_INDUSTRY_NEUTRAL_TC:
        ind_signal_panel, _ = get_signal_panel_for_transaction_cost(
            method_name="Industry-neutral weight"
        )

        industry_daily, industry_rebalance = run_one_transaction_cost_backtest(
            signal_panel=ind_signal_panel,
            signal_col=signal_col,
            method_name="Industry-neutral weight",
            target_builder=build_industry_neutral_targets,
            rebalance_periods=GROUP_REBALANCE_PERIODS,
        )

        all_daily.append(industry_daily)
        all_rebalance.append(industry_rebalance)

    daily_tc_returns = pd.concat(all_daily, ignore_index=True)
    rebalance_tc = pd.concat(all_rebalance, ignore_index=True)

    nav_df, summary_df = build_transaction_cost_nav_and_summary(daily_tc_returns)

    q5q1_summary = (
        summary_df[summary_df["portfolio"] == "Q5_minus_Q1"]
        .sort_values(["method", "horizon"])
        .reset_index(drop=True)
    )

    q_group_summary = (
        summary_df[summary_df["portfolio"] != "Q5_minus_Q1"]
        .sort_values(["method", "horizon", "portfolio"])
        .reset_index(drop=True)
    )

    methodology = pd.DataFrame(
        [
            {
                "Item": "Buy cost",
                "Description": f"{BUY_COST_RATE:.6f}. Applied to positive rebalance delta.",
            },
            {
                "Item": "Sell cost",
                "Description": f"{SELL_COST_RATE:.6f}. Applied to negative rebalance delta.",
            },
            {
                "Item": "Buy turnover",
                "Description": "sum(max(new_weight - old_weight, 0)) at each rebalance.",
            },
            {
                "Item": "Sell turnover",
                "Description": "sum(max(old_weight - new_weight, 0)) at each rebalance.",
            },
            {
                "Item": "Transaction cost",
                "Description": "buy_turnover * buy_cost + sell_turnover * sell_cost.",
            },
            {
                "Item": "Weight drift",
                "Description": "Between rebalances, weights drift with daily VWAP returns before the next turnover calculation.",
            },
            {
                "Item": "Q1-Q5",
                "Description": "Long-only quintile portfolios.",
            },
            {
                "Item": "Q5_minus_Q1",
                "Description": "Signed long-short portfolio: long Q5 and short Q1. Opening Q5 is buy turnover; opening Q1 short is sell turnover.",
            },
            {
                "Item": "Net return",
                "Description": "For each cost day, net_ret = (1 - transaction_cost) * (1 + gross_ret) - 1.",
            },
            {
                "Item": "Execution assumption",
                "Description": "Same as previous group tests: signal at t, rebalance/enter at VWAP[t+1].",
            },
        ]
    )

    config = pd.DataFrame(
        [
            {"item": "signal_mode", "value": SIGNAL_MODE},
            {"item": "signal_column", "value": signal_col},
            {"item": "rebalance_periods", "value": str(GROUP_REBALANCE_PERIODS)},
            {"item": "buy_cost_rate", "value": BUY_COST_RATE},
            {"item": "sell_cost_rate", "value": SELL_COST_RATE},
            {"item": "run_original_equal_weight", "value": RUN_ORIGINAL_EQUAL_WEIGHT_TC},
            {"item": "run_industry_neutral", "value": RUN_INDUSTRY_NEUTRAL_TC},
            {"item": "industry_weight_method", "value": globals().get("INDUSTRY_WEIGHT_METHOD", np.nan)},
            {"item": "use_common_industries_only", "value": globals().get("USE_COMMON_INDUSTRIES_ONLY", np.nan)},
        ]
    )

    # Save CSV
    summary_df.to_csv(
        TC_OUTDIR / "transaction_cost_performance_summary_all.csv",
        index=False,
        encoding="utf-8-sig",
    )

    q5q1_summary.to_csv(
        TC_OUTDIR / "transaction_cost_q5_minus_q1_summary.csv",
        index=False,
        encoding="utf-8-sig",
    )

    nav_df.to_csv(
        TC_OUTDIR / "transaction_cost_daily_nav.csv",
        index=False,
        encoding="utf-8-sig",
    )

    rebalance_tc.to_csv(
        TC_OUTDIR / "transaction_cost_rebalance_turnover.csv",
        index=False,
        encoding="utf-8-sig",
    )

    # Save Excel
    with pd.ExcelWriter(TC_EXCEL, engine="openpyxl") as writer:
        q5q1_summary.to_excel(writer, sheet_name="Q5_Q1_Summary", index=False)
        summary_df.to_excel(writer, sheet_name="All_Portfolio_Summary", index=False)
        q_group_summary.to_excel(writer, sheet_name="Q1_Q5_Summary", index=False)
        rebalance_tc.to_excel(writer, sheet_name="Rebalance_Turnover", index=False)
        nav_df.to_excel(writer, sheet_name="Daily_NAV_Returns", index=False)
        config.to_excel(writer, sheet_name="Config", index=False)
        methodology.to_excel(writer, sheet_name="Methodology", index=False)

    format_transaction_cost_excel(TC_EXCEL)

    print(f"\nSaved transaction cost Excel: {TC_EXCEL}")

    print("\nQ5_minus_Q1 gross vs net summary:")
    print(
        q5q1_summary[
            [
                "method",
                "horizon",
                "annual_return_gross",
                "annual_return_net",
                "annual_return_decay",
                "sharpe_no_rf_gross",
                "sharpe_no_rf_net",
                "sharpe_no_rf_decay",
                "max_drawdown_gross",
                "max_drawdown_net",
                "avg_turnover_on_rebalance_day",
                "avg_rebalance_day_transaction_cost",
            ]
        ].to_string(index=False)
    )

    return {
        "daily_tc_returns": daily_tc_returns,
        "rebalance_tc": rebalance_tc,
        "nav": nav_df,
        "summary": summary_df,
        "q5q1_summary": q5q1_summary,
    }


# =============================================================================
# 6. Excel formatting
# =============================================================================

def format_transaction_cost_excel(excel_path: Path):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(excel_path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    percent_cols = {
        "total_transaction_cost",
        "avg_daily_transaction_cost",
        "avg_rebalance_day_transaction_cost",
        "avg_buy_turnover_on_rebalance_day",
        "avg_sell_turnover_on_rebalance_day",
        "avg_turnover_on_rebalance_day",

        "total_return_gross",
        "total_return_net",
        "total_return_decay",
        "annual_return_gross",
        "annual_return_net",
        "annual_return_decay",
        "annual_vol_gross",
        "annual_vol_net",
        "annual_vol_decay",
        "max_drawdown_gross",
        "max_drawdown_net",
        "max_drawdown_decay",
        "avg_daily_return_gross",
        "avg_daily_return_net",
        "avg_daily_return_decay",
        "daily_win_rate_gross",
        "daily_win_rate_net",
        "daily_win_rate_decay",

        "long_exposure",
        "short_exposure",
        "gross_exposure",
        "net_exposure",
        "buy_turnover",
        "sell_turnover",
        "turnover",
        "transaction_cost",
        "buy_cost_rate",
        "sell_cost_rate",

        "gross_ret",
        "net_ret",
    }

    numeric_4_cols = {
        "sharpe_no_rf_gross",
        "sharpe_no_rf_net",
        "sharpe_no_rf_decay",
        "gross_nav",
        "net_nav",
    }

    integer_cols = {
        "horizon",
        "n_days",
        "n_names",
        "long_names",
        "short_names",
        "n_ret_obs",
    }

    date_cols = {
        "trade_day",
        "signal_day",
        "entry_day",
        "exit_day",
    }

    for ws in wb.worksheets:
        if ws.max_row < 1 or ws.max_column < 1:
            continue

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        header_to_col = {
            str(ws.cell(row=1, column=col_idx).value): col_idx
            for col_idx in range(1, ws.max_column + 1)
        }

        for header, col_idx in header_to_col.items():
            if header in percent_cols:
                fmt = "0.00%"
            elif header in numeric_4_cols:
                fmt = "0.0000"
            elif header in integer_cols:
                fmt = "0"
            elif header in date_cols:
                fmt = "yyyy-mm-dd"
            else:
                fmt = None

            if fmt is not None:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = fmt

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            scan_rows = min(ws.max_row, 3000)

            for row_idx in range(1, scan_rows + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))

            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 32)

    wb.save(excel_path)


# 7. Run

if __name__ == "__main__":
    transaction_cost_result = run_transaction_cost_tests()