from __future__ import annotations

from pathlib import Path
from typing import Iterable, Sequence
import math
import re
import warnings

import numpy as np
import pandas as pd

try:
    from .config import *
except ImportError:  # allow running a file directly from this folder
    from config import *

try:
    from .core_base import *
    from .core_neutralization import *
    from .composite_weighted import *
except ImportError:
    from core_base import *
    from core_neutralization import *
    from composite_weighted import *

# Teacher requirement:
# 1. Use the integrated composite signal used for backtest.
# 2. Test IC decay for holding horizons 1, 2, ..., 30.
# 3. Plot holding horizon vs IC decay curve.
# 4. Export standalone Excel and PNG charts.



# Test horizons 1 to 60 trading days.
DECAY_MAX_HORIZON = 60
IC_DECAY_HORIZONS = list(range(1, DECAY_MAX_HORIZON + 1))

IC_DECAY_OUTDIR = OUTDIR / f"composite_signal_ic_decay_1_{DECAY_MAX_HORIZON}"
IC_DECAY_OUTDIR.mkdir(parents=True, exist_ok=True)

IC_DECAY_EXCEL = IC_DECAY_OUTDIR / f"composite_signal_ic_decay_1_{DECAY_MAX_HORIZON}.xlsx"

IC_DECAY_CURVE_PNG = IC_DECAY_OUTDIR / "composite_signal_rank_ic_decay_curve.png"
IC_DECAY_ICIR_PNG = IC_DECAY_OUTDIR / "composite_signal_rank_icir_decay_curve.png"
IC_DECAY_POSITIVE_RATIO_PNG = IC_DECAY_OUTDIR / "composite_signal_positive_ratio_decay_curve.png"

# Signal choice:
# "equal"            = group_signal_equal_mad_z, 平权后的因子
# "weighted"         = group_signal_weighted_mad_z
# "current_backtest" = follow current SIGNAL_MODE
DECAY_SIGNAL_MODE = "current_backtest"

DECAY_FACTOR_COL = "composite_signal_for_ic_decay"

# Keep this True if you want to show industry + size neutralized IC decay too.
DO_DECAY_INDUSTRY_SIZE_NEUTRAL_IC = True

# Main IC column used in decay diagnostics.
# "auto" prefers neutral_rank_ic_mean if available, otherwise rank_ic_mean.
DECAY_MAIN_IC_COL = "auto"
USE_COMMON_DECAY_SAMPLE = True
DECAY_NW_LAG_RULE = "h+10"
def choose_decay_nw_lag(horizon: int, n_obs: int, lag_rule: str = DECAY_NW_LAG_RULE) -> int:
    """
    Choose Newey-West lag for IC decay analysis.

    Recommended main rule:
        "h" = lag equals forward-return horizon.

    Robustness:
        "h+10"
        "fixed20"
        "fixed30"
    """
    h = int(horizon)

    if lag_rule == "h":
        L = h
    elif lag_rule == "h+10":
        L = h + 10
    elif lag_rule.startswith("fixed"):
        L = int(lag_rule.replace("fixed", ""))
    else:
        raise ValueError(
            "lag_rule must be one of: 'h', 'h+10', 'fixed20', 'fixed30'"
        )

    return int(min(max(L, 1), n_obs - 1))


def newey_west_std_icir_decay(
    ic_series: pd.Series,
    horizon: int,
    lag_rule: str = DECAY_NW_LAG_RULE,
) -> dict:
    """
    Newey-West corrected std and ICIR for IC decay only.
    This does not affect earlier factor selection code.
    """
    x = pd.Series(ic_series).dropna().astype(float).to_numpy()
    n = len(x)

    if n < 2:
        return {
            "nw_std": np.nan,
            "nw_icir": np.nan,
            "nw_lag": np.nan,
            "nw_n_obs": n,
        }

    mu = np.mean(x)
    xc = x - mu

    L = choose_decay_nw_lag(horizon=horizon, n_obs=n, lag_rule=lag_rule)

    nw_var = np.dot(xc, xc) / n

    for lag in range(1, L + 1):
        gamma_l = np.dot(xc[lag:], xc[:-lag]) / n
        weight = 1.0 - lag / (L + 1.0)
        nw_var += 2.0 * weight * gamma_l

    if not np.isfinite(nw_var) or nw_var <= 0:
        nw_std = np.nan
        nw_icir = np.nan
    else:
        nw_std = np.sqrt(nw_var)
        nw_icir = mu / nw_std if nw_std > 0 else np.nan

    return {
        "nw_std": nw_std,
        "nw_icir": nw_icir,
        "nw_lag": L,
        "nw_n_obs": n,
    }


def compute_newey_west_icir_table_all_decay(
    daily: pd.DataFrame,
    lag_rule: str = DECAY_NW_LAG_RULE,
) -> pd.DataFrame:
    """
    Compute Newey-West adjusted ICIR for decay analysis.
    Uses decay-specific lag rule.
    """
    ic_col_map = {
        "ic": "ic_nw",
        "winsor_ic": "winsor_ic_nw",
        "rank_ic": "rank_ic_nw",
        "neutral_ic": "neutral_ic_nw",
        "neutral_winsor_ic": "neutral_winsor_ic_nw",
        "neutral_rank_ic": "neutral_rank_ic_nw",
    }

    records = []

    for (factor, horizon), grp in daily.groupby(["factor", "horizon"], sort=True):
        horizon = int(horizon)

        rec = {
            "factor": factor,
            "horizon": horizon,
            "decay_nw_lag_rule": lag_rule,
            "nw_lag": np.nan,
            "nw_n_obs": np.nan,
        }

        rank_stats = None
        first_valid_stats = None

        for ic_col, prefix in ic_col_map.items():
            if ic_col not in grp.columns:
                continue

            stats = newey_west_std_icir_decay(
                ic_series=grp[ic_col],
                horizon=horizon,
                lag_rule=lag_rule,
            )

            rec[f"{prefix}_std"] = stats["nw_std"]
            rec[f"{prefix}_icir"] = stats["nw_icir"]
            rec[f"{prefix}_lag"] = stats["nw_lag"]
            rec[f"{prefix}_n_obs"] = stats["nw_n_obs"]

            if first_valid_stats is None and pd.notna(stats["nw_icir"]):
                first_valid_stats = stats

            if ic_col == "rank_ic":
                rank_stats = stats

        main_stats = rank_stats or first_valid_stats

        if main_stats is not None:
            rec["nw_lag"] = main_stats["nw_lag"]
            rec["nw_n_obs"] = main_stats["nw_n_obs"]

        records.append(rec)

    return pd.DataFrame(records)



def decay_horizon_label() -> str:
    return f"1-{max(IC_DECAY_HORIZONS)} Trading Days"


def decay_xticks() -> list[int]:
    """
    Avoid showing 60 crowded x-axis labels.
    For 1-60, show 1, 5, 10, ..., 60.
    """
    max_h = max(IC_DECAY_HORIZONS)

    if max_h <= 30:
        return IC_DECAY_HORIZONS

    ticks = [1] + list(range(5, max_h + 1, 5))

    if max_h not in ticks:
        ticks.append(max_h)

    return ticks

# =============================================================================
# 1. Resolve signal column
# =============================================================================

def resolve_decay_signal_col(signal_panel: pd.DataFrame) -> tuple[str, str, str]:
    """
    Decide which composite signal column to use.

    Returns:
        signal_col, signal_mode_used, signal_label
    """
    equal_col = globals().get("GROUP_SIGNAL_EQUAL_COL", "group_signal_equal_mad_z")
    weighted_col = globals().get("GROUP_SIGNAL_WEIGHTED_COL", "group_signal_weighted_mad_z")

    if DECAY_SIGNAL_MODE == "current_backtest":
        mode = globals().get("SIGNAL_MODE", "weighted")
    else:
        mode = DECAY_SIGNAL_MODE

    if mode == "equal":
        signal_col = equal_col
        signal_label = "Equal-weight composite signal, MAD-z"

    elif mode == "weighted":
        signal_col = weighted_col
        signal_label = "Abs(NW Rank ICIR)-weighted composite signal, MAD-z"

    else:
        raise ValueError(
            "DECAY_SIGNAL_MODE must be 'equal', 'weighted', or 'current_backtest'."
        )

    if signal_col not in signal_panel.columns:
        raise ValueError(
            f"Signal column {signal_col} not found in signal_panel. "
            f"Available columns are: {list(signal_panel.columns)}"
        )

    return signal_col, mode, signal_label


def get_signal_panel_for_ic_decay() -> tuple[pd.DataFrame, pd.DataFrame | None]:
    """
    Reuse existing signal panel if available.
    Otherwise rebuild signal panel using build_group_test_signal_panel().
    """
    if (
        "five_group_result" in globals()
        and isinstance(five_group_result, dict)
        and "signal_panel" in five_group_result
    ):
        print("Using five_group_result['signal_panel'] from memory.")
        return five_group_result["signal_panel"].copy(), None

    if "build_group_test_signal_panel" not in globals():
        raise NameError(
            "build_group_test_signal_panel() is not defined. "
            "Please run the five-group signal construction chunk first."
        )

    print("five_group_result not found. Rebuilding signal panel ...")
    signal_panel, weights = build_group_test_signal_panel()
    return signal_panel, weights


# =============================================================================
# 2. Build IC input panel
# =============================================================================

def build_composite_signal_ic_decay_panel() -> tuple[pd.DataFrame, str, str, list[str]]:
    """
    Build panel for 1-30 day IC decay test.

    Output panel includes:
        code
        trade_day
        price
        vwap_price
        volume
        DECAY_FACTOR_COL
        fwd_ret_1d ... fwd_ret_30d
        optional industry + size exposure columns
    """
    signal_panel, signal_weights = get_signal_panel_for_ic_decay()

    signal_panel = signal_panel.copy()
    signal_panel["trade_day"] = pd.to_datetime(signal_panel["trade_day"]).dt.normalize()
    signal_panel["code"] = signal_panel["code"].astype(str).str.upper()

    signal_col, signal_mode_used, signal_label = resolve_decay_signal_col(signal_panel)

    print(f"Using signal mode: {signal_mode_used}")
    print(f"Using signal column: {signal_col}")
    print(f"Signal label: {signal_label}")

    required_cols = ["code", "trade_day", "price", "vwap_price", "volume", signal_col]
    missing = [c for c in required_cols if c not in signal_panel.columns]

    if missing:
        raise ValueError(f"signal_panel missing required columns: {missing}")

    df = signal_panel[required_cols].copy()

    df[DECAY_FACTOR_COL] = pd.to_numeric(df[signal_col], errors="coerce")

    # Drop original signal column after copying into DECAY_FACTOR_COL.
    # This avoids long column names in later output.
    df = df.drop(columns=[signal_col])

    exposure_cols = []

    if DO_DECAY_INDUSTRY_SIZE_NEUTRAL_IC:
        print("Adding industry + size exposures for neutral IC decay ...")

        df, exposure_cols, industry_cols = add_industry_size_exposures(
            df,
            industry_xlsx=INDUSTRY_XLSX,
            market_cap_xlsx=MARKET_CAP_XLSX,
            base_date=MARKET_CAP_BASE_DATE,
            price_col="price",
        )

        keep_cols = [
            "code",
            "trade_day",
            "price",
            "vwap_price",
            "volume",
            DECAY_FACTOR_COL,
        ] + exposure_cols

        if "industry_valid" in df.columns:
            keep_cols.append("industry_valid")

        keep_cols = list(dict.fromkeys([c for c in keep_cols if c in df.columns]))
        df = df[keep_cols].copy()

    df = add_forward_returns(
        df,
        price_col="price",
        return_price_col="vwap_price",
        horizons=IC_DECAY_HORIZONS,
        use_next_day_vwap=USE_NEXT_DAY_VWAP_RETURN,
    )
    if USE_COMMON_DECAY_SAMPLE:
        max_h = max(IC_DECAY_HORIZONS)
        max_ret_col = f"fwd_ret_{max_h}d"

        before_rows = len(df)
        before_days = df["trade_day"].nunique()

        df = df[df[max_ret_col].notna()].copy()

        after_rows = len(df)
        after_days = df["trade_day"].nunique()

    print(
        f"Common decay sample enabled using {max_ret_col}: "
        f"{before_rows:,} rows / {before_days:,} days -> "
        f"{after_rows:,} rows / {after_days:,} days"
    )

    return df, signal_mode_used, signal_label, exposure_cols


# =============================================================================
# 3. Compute 1-30 day IC decay
# =============================================================================

def compute_composite_signal_ic_decay() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, str, str]:
    """
    Compute IC decay for composite signal over 1-30 holding horizons.
    """
    df_ic, signal_mode_used, signal_label, exposure_cols = build_composite_signal_ic_decay_panel()

    decay_summary, decay_daily = compute_ic_summary_fast(
        df_ic,
        factor_cols=[DECAY_FACTOR_COL],
        horizons=IC_DECAY_HORIZONS,
        min_obs=MIN_OBS,
        tradable_only=not INCLUDE_SUSPENDED,
        exposure_cols=exposure_cols if DO_DECAY_INDUSTRY_SIZE_NEUTRAL_IC else None,
        neutralize_return_too=NEUTRALIZE_RETURN_TOO,
    )

    if decay_summary.empty:
        raise ValueError(
            "decay_summary is empty. Please check signal coverage, forward returns, and MIN_OBS."
        )
        
        
    decay_nw = compute_newey_west_icir_table_all_decay(
    daily=decay_daily,
    lag_rule=DECAY_NW_LAG_RULE,
)

# Avoid duplicated columns if rerun in the same notebook session.
    nw_cols_to_remove = [
    c for c in decay_nw.columns
    if c in decay_summary.columns and c not in ["factor", "horizon"]
]

    decay_summary = decay_summary.drop(columns=nw_cols_to_remove, errors="ignore")

    decay_summary = decay_summary.merge(
    decay_nw,
    on=["factor", "horizon"],
    how="left",
)

# Remove raw-vs-NW comparison columns if they already exist from an older run.
    raw_vs_nw_cols = [
    "icir_nw_diff",
    "rank_icir_nw_diff",
    "icir_nw_change_pct",
    "rank_icir_nw_change_pct",
]

    decay_summary = decay_summary.drop(columns=raw_vs_nw_cols, errors="ignore")

    decay_summary["signal_mode"] = signal_mode_used
    decay_summary["signal_label"] = signal_label
    decay_summary["factor_label"] = signal_label

    decay_summary = decay_summary.sort_values("horizon").reset_index(drop=True)
    decay_daily = decay_daily.sort_values(["horizon", "trade_day"]).reset_index(drop=True)

    diagnostics = build_ic_decay_diagnostics(
        decay_summary=decay_summary,
        signal_mode_used=signal_mode_used,
        signal_label=signal_label,
    )

    return decay_summary, decay_daily, diagnostics, signal_mode_used, signal_label


# =============================================================================
# 4. Diagnostics
# =============================================================================

def choose_main_ic_col(decay_summary: pd.DataFrame) -> str:
    """
    Choose main IC column for interpretation.
    """
    if DECAY_MAIN_IC_COL != "auto":
        if DECAY_MAIN_IC_COL not in decay_summary.columns:
            raise ValueError(f"DECAY_MAIN_IC_COL={DECAY_MAIN_IC_COL} not found.")
        return DECAY_MAIN_IC_COL

    if (
        "neutral_rank_ic_mean" in decay_summary.columns
        and decay_summary["neutral_rank_ic_mean"].notna().any()
    ):
        return "neutral_rank_ic_mean"

    return "rank_ic_mean"


def build_ic_decay_diagnostics(
    decay_summary: pd.DataFrame,
    signal_mode_used: str,
    signal_label: str,
) -> pd.DataFrame:
    """
    Summarize effective horizon scale.

    Main logic:
        peak horizon = horizon with max abs(main IC)
        half-life horizon = first horizon after peak where abs(IC) <= 50% of peak
        80% horizon = max horizon where abs(IC) >= 80% of peak
        sign reversal horizon = first horizon where IC has opposite sign to peak IC
    """
    main_col = choose_main_ic_col(decay_summary)

    s = decay_summary[["horizon", main_col]].dropna().copy()

    if s.empty:
        return pd.DataFrame(
            [
                {
                    "signal_mode": signal_mode_used,
                    "signal_label": signal_label,
                    "main_ic_col": main_col,
                    "note": "No valid IC values.",
                }
            ]
        )

    s["abs_ic"] = s[main_col].abs()

    peak_row = s.loc[s["abs_ic"].idxmax()]
    peak_horizon = int(peak_row["horizon"])
    peak_ic = float(peak_row[main_col])
    peak_abs_ic = float(abs(peak_ic))

    if peak_abs_ic > 0:
        half_threshold = 0.5 * peak_abs_ic
        threshold_80 = 0.8 * peak_abs_ic
    else:
        half_threshold = np.nan
        threshold_80 = np.nan

    after_peak = s[s["horizon"] > peak_horizon].copy()

    half_life_horizon = np.nan

    if peak_abs_ic > 0 and not after_peak.empty:
        tmp_half = after_peak[after_peak["abs_ic"] <= half_threshold]
        if not tmp_half.empty:
            half_life_horizon = int(tmp_half.iloc[0]["horizon"])

    effective_80_horizon = np.nan

    if peak_abs_ic > 0:
        tmp_80 = s[s["abs_ic"] >= threshold_80]
        if not tmp_80.empty:
            effective_80_horizon = int(tmp_80["horizon"].max())

    sign_reversal_horizon = np.nan
    peak_sign = np.sign(peak_ic)

    if peak_sign != 0:
        tmp_rev = s[(s["horizon"] > peak_horizon) & (np.sign(s[main_col]) == -peak_sign)]
        if not tmp_rev.empty:
            sign_reversal_horizon = int(tmp_rev.iloc[0]["horizon"])

    # Also record simple best horizons by other metrics.
    rank_ic_peak_h = np.nan
    neutral_rank_ic_peak_h = np.nan
    nw_rank_icir_peak_h = np.nan

    if "rank_ic_mean" in decay_summary.columns and decay_summary["rank_ic_mean"].notna().any():
        rank_ic_peak_h = int(
            decay_summary.loc[decay_summary["rank_ic_mean"].abs().idxmax(), "horizon"]
        )

    if (
        "neutral_rank_ic_mean" in decay_summary.columns
        and decay_summary["neutral_rank_ic_mean"].notna().any()
    ):
        neutral_rank_ic_peak_h = int(
            decay_summary.loc[decay_summary["neutral_rank_ic_mean"].abs().idxmax(), "horizon"]
        )

    if (
        "rank_ic_nw_icir" in decay_summary.columns
        and decay_summary["rank_ic_nw_icir"].notna().any()
    ):
        nw_rank_icir_peak_h = int(
            decay_summary.loc[decay_summary["rank_ic_nw_icir"].abs().idxmax(), "horizon"]
        )

    diagnostics = pd.DataFrame(
        [
            {
                "signal_mode": signal_mode_used,
                "signal_label": signal_label,
                "main_ic_col": main_col,

                "peak_horizon_by_abs_main_ic": peak_horizon,
                "peak_main_ic": peak_ic,
                "peak_abs_main_ic": peak_abs_ic,

                "half_life_horizon_after_peak": half_life_horizon,
                "effective_80pct_horizon": effective_80_horizon,
                "sign_reversal_horizon_after_peak": sign_reversal_horizon,

                "rank_ic_peak_horizon_by_abs_mean": rank_ic_peak_h,
                "neutral_rank_ic_peak_horizon_by_abs_mean": neutral_rank_ic_peak_h,
                "nw_rank_icir_peak_horizon_by_abs": nw_rank_icir_peak_h,
            }
        ]
    )

    return diagnostics


# =============================================================================
# 5. Plot IC decay curves
# =============================================================================

def plot_ic_decay_curves(decay_summary: pd.DataFrame):
    """
    Save IC decay plots.

    Revised for 1-60:
    - Keep IC mean decay plot.
    - For ICIR, only show Newey-West adjusted ICIR.
    - Do NOT show raw ICIR.
    - Use dynamic horizon title and cleaner x-axis ticks.
    """
    import matplotlib.pyplot as plt

    x = decay_summary["horizon"].astype(int)
    horizon_label = decay_horizon_label()
    x_ticks = decay_xticks()

    # -------------------------------------------------------------------------
    # Plot 1: IC mean decay
    # -------------------------------------------------------------------------
    fig, ax = plt.subplots(figsize=(12, 7))

    if "rank_ic_mean" in decay_summary.columns:
        ax.plot(
            x,
            decay_summary["rank_ic_mean"],
            marker="o",
            label="Rank IC mean",
        )

    if (
        "neutral_rank_ic_mean" in decay_summary.columns
        and decay_summary["neutral_rank_ic_mean"].notna().any()
    ):
        ax.plot(
            x,
            decay_summary["neutral_rank_ic_mean"],
            marker="o",
            label="Neutral Rank IC mean",
        )

    if "winsor_ic_mean" in decay_summary.columns:
        ax.plot(
            x,
            decay_summary["winsor_ic_mean"],
            marker="o",
            label="Winsorized Pearson IC mean",
        )

    ax.axhline(0.0, linewidth=1)
    ax.set_title(f"Composite Signal IC Mean Decay, Holding Horizon {horizon_label}")
    ax.set_xlabel("Holding horizon / forward return horizon, trading days")
    ax.set_ylabel("Mean daily IC")
    ax.set_xticks(x_ticks)
    ax.grid(True, alpha=0.3)
    ax.legend()
    fig.tight_layout()
    fig.savefig(IC_DECAY_CURVE_PNG, dpi=300, bbox_inches="tight")
    plt.close(fig)

    print(f"Saved IC mean decay curve: {IC_DECAY_CURVE_PNG}")

    # -------------------------------------------------------------------------
    # Plot 2: Newey-West adjusted ICIR only
    # -------------------------------------------------------------------------
    fig, ax = plt.subplots(figsize=(12, 7))

    plotted = False

    nw_icir_plot_cols = [
        ("rank_ic_nw_icir", "NW-adjusted Rank ICIR"),
        ("neutral_rank_ic_nw_icir", "NW-adjusted Neutral Rank ICIR"),
    ]

    for col, label in nw_icir_plot_cols:
        if col in decay_summary.columns and decay_summary[col].notna().any():
            ax.plot(
                x,
                decay_summary[col],
                marker="o",
                label=label,
            )
            plotted = True

    if plotted:
        ax.axhline(0.0, linewidth=1)
        ax.set_title(f"Composite Signal Newey-West Adjusted ICIR Decay, Holding Horizon {horizon_label}")
        ax.set_xlabel("Holding horizon / forward return horizon, trading days")
        ax.set_ylabel("NW-adjusted ICIR")
        ax.set_xticks(x_ticks)
        ax.grid(True, alpha=0.3)
        ax.legend()
        fig.tight_layout()
        fig.savefig(IC_DECAY_ICIR_PNG, dpi=300, bbox_inches="tight")
        plt.close(fig)

        print(f"Saved NW-adjusted ICIR decay curve: {IC_DECAY_ICIR_PNG}")
    else:
        plt.close(fig)
        print("No Newey-West adjusted ICIR columns found. ICIR plot skipped.")

    # -------------------------------------------------------------------------
    # Plot 3: Positive ratio decay
    # -------------------------------------------------------------------------
    fig, ax = plt.subplots(figsize=(12, 7))

    if "rank_ic_positive_ratio" in decay_summary.columns:
        ax.plot(
            x,
            decay_summary["rank_ic_positive_ratio"],
            marker="o",
            label="Rank IC positive ratio",
        )

    if (
        "neutral_rank_ic_positive_ratio" in decay_summary.columns
        and decay_summary["neutral_rank_ic_positive_ratio"].notna().any()
    ):
        ax.plot(
            x,
            decay_summary["neutral_rank_ic_positive_ratio"],
            marker="o",
            label="Neutral Rank IC positive ratio",
        )

    ax.axhline(0.5, linewidth=1)
    ax.set_title(f"Composite Signal IC Positive Ratio Decay, Holding Horizon {horizon_label}")
    ax.set_xlabel("Holding horizon / forward return horizon, trading days")
    ax.set_ylabel("Positive ratio")
    ax.set_xticks(x_ticks)
    ax.grid(True, alpha=0.3)
    ax.legend()
    fig.tight_layout()
    fig.savefig(IC_DECAY_POSITIVE_RATIO_PNG, dpi=300, bbox_inches="tight")
    plt.close(fig)

    print(f"Saved positive ratio decay curve: {IC_DECAY_POSITIVE_RATIO_PNG}")

# =============================================================================
# 6. Export Excel
# =============================================================================
EXPORT_DAILY_IC_INPUT_FOR_AUDIT = False


def make_adjusted_only_decay_summary(decay_summary: pd.DataFrame) -> pd.DataFrame:
    """
    Create final display table.

    Rule:
    - Keep IC mean / positive ratio because they are not ICIR.
    - For ICIR, only keep Newey-West adjusted ICIR.
    - Remove raw ICIR and raw-vs-NW comparison columns from final display.
    """
    display_cols = [
        # Identity
        "signal_mode",
        "signal_label",
        "factor",
        "factor_label",
        "horizon",

        # IC mean decay
        "ic_mean",
        "winsor_ic_mean",
        "rank_ic_mean",
        "neutral_ic_mean",
        "neutral_winsor_ic_mean",
        "neutral_rank_ic_mean",

        # Positive ratio
        "ic_positive_ratio",
        "winsor_ic_positive_ratio",
        "rank_ic_positive_ratio",
        "neutral_ic_positive_ratio",
        "neutral_winsor_ic_positive_ratio",
        "neutral_rank_ic_positive_ratio",

        # Newey-West adjusted ICIR only
        "ic_nw_icir",
        "winsor_ic_nw_icir",
        "rank_ic_nw_icir",
        "neutral_ic_nw_icir",
        "neutral_winsor_ic_nw_icir",
        "neutral_rank_ic_nw_icir",

        # Newey-West adjusted std
        "ic_nw_std",
        "winsor_ic_nw_std",
        "rank_ic_nw_std",
        "neutral_ic_nw_std",
        "neutral_winsor_ic_nw_std",
        "neutral_rank_ic_nw_std",

        # Diagnostics
        "sample_days",
        "neutral_sample_days",
        "avg_obs_per_day",
        "avg_neutral_obs_per_day",
        "nw_lag",
        "nw_n_obs",
        "start_day",
        "end_day",
    ]

    out = decay_summary[[c for c in display_cols if c in decay_summary.columns]].copy()

    return out

def write_ic_decay_excel(
    decay_summary: pd.DataFrame,
    decay_daily: pd.DataFrame,
    diagnostics: pd.DataFrame,
    signal_mode_used: str,
    signal_label: str,
):
    """
    Export IC decay summary, daily IC, diagnostics, config, and methodology to Excel.
    """
    config = pd.DataFrame(
        [
            {"item": "signal_mode_used", "value": signal_mode_used},
            {"item": "signal_label", "value": signal_label},
            {"item": "factor_column_in_test", "value": DECAY_FACTOR_COL},
            {"item": "horizons", "value": str(IC_DECAY_HORIZONS)},
            {"item": "return_definition", "value": "VWAP[t+h+1] / VWAP[t+1] - 1"},
            {"item": "price_mode", "value": PRICE_MODE},
            {"item": "tradable_filter", "value": "volume > 0" if not INCLUDE_SUSPENDED else "include suspended"},
            {"item": "winsor_q", "value": WINSOR_Q},
            {"item": "winsorize_factor", "value": WINSORIZE_FACTOR},
            {"item": "winsorize_return", "value": WINSORIZE_RETURN},
            {"item": "do_industry_size_neutral_ic", "value": DO_DECAY_INDUSTRY_SIZE_NEUTRAL_IC},
            {"item": "neutralize_return_too", "value": NEUTRALIZE_RETURN_TOO},
            {"item": "min_obs", "value": MIN_OBS},
            {"item": "years", "value": str(YEARS)},
        ]
    )

    methodology = pd.DataFrame(
    [
        {
            "Item": "Goal",
            "Description": (
                f"Measure how the composite signal IC decays as forward return "
                f"horizon increases from 1 to {max(IC_DECAY_HORIZONS)} trading days."
            ),
        },
        {
            "Item": "Signal",
            "Description": (
                "The integrated composite signal used in the five-group backtest. "
                "Default here is the equal-weight MAD-z composite signal."
            ),
        },
        {
            "Item": "Forward return",
            "Description": (
                "For horizon h, return is VWAP[t+h+1] / VWAP[t+1] - 1. "
                "This avoids trading at same-day close."
            ),
        },
        {
            "Item": "Rank IC",
            "Description": "Spearman correlation between cross-sectional signal ranks and forward return ranks.",
        },
        {
            "Item": "Neutral Rank IC",
            "Description": "Rank IC after residualizing the signal by industry dummies and log market cap.",
        },
        {
            "Item": "NW Rank ICIR",
            "Description": "Newey-West corrected Rank ICIR using lag = horizon + 10.",
        },
        {
            "Item": "Main interpretation",
            "Description": (
                "If Rank IC remains positive at longer horizons, the signal has slower decay. "
                "If NW-adjusted ICIR declines as horizon increases, the statistical stability weakens "
                "after accounting for overlapping forward returns."
            ),
        },
    ]
)

    summary_display = make_adjusted_only_decay_summary(decay_summary)

    with pd.ExcelWriter(IC_DECAY_EXCEL, engine="openpyxl") as writer:
        summary_display.to_excel(writer, sheet_name="IC_Decay_Summary", index=False)
        diagnostics.to_excel(writer, sheet_name="Decay_Diagnostics", index=False)

    # Daily IC is the input series for Newey-West.
    # It is useful for audit, but not part of final adjusted-only display.
        if EXPORT_DAILY_IC_INPUT_FOR_AUDIT:
            decay_daily.to_excel(writer, sheet_name="Daily_IC_Input_For_NW", index=False)

        config.to_excel(writer, sheet_name="Config", index=False)
        methodology.to_excel(writer, sheet_name="Methodology", index=False)

    format_ic_decay_excel(IC_DECAY_EXCEL)

    print(f"Saved IC decay Excel: {IC_DECAY_EXCEL}")


def format_ic_decay_excel(excel_path: Path):
    """
    Format Excel workbook and optionally insert saved charts.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(excel_path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    percent_cols = {
        "ic_positive_ratio",
        "winsor_ic_positive_ratio",
        "rank_ic_positive_ratio",
        "neutral_ic_positive_ratio",
        "neutral_winsor_ic_positive_ratio",
        "neutral_rank_ic_positive_ratio",
        "raw_rank_sign_match_ratio",
        "winsor_rank_sign_match_ratio",
        "neutral_rank_sign_match_ratio",
        "neutral_winsor_rank_sign_match_ratio",
        "icir_nw_change_pct",
        "rank_icir_nw_change_pct",
    }

    integer_cols = {
        "horizon",
        "sample_days",
        "neutral_sample_days",
        "nw_lag",
        "nw_n_obs",
        "n_obs",
        "neutral_n_obs",
        "peak_horizon_by_abs_main_ic",
        "half_life_horizon_after_peak",
        "effective_80pct_horizon",
        "sign_reversal_horizon_after_peak",
        "rank_ic_peak_horizon_by_abs_mean",
        "neutral_rank_ic_peak_horizon_by_abs_mean",
        "nw_rank_icir_peak_horizon_by_abs",
    }

    date_cols = {
        "trade_day",
        "start_day",
        "end_day",
    }

    numeric_4_keywords = [
        "ic",
        "icir",
        "std",
        "mean",
        "gap",
        "obs",
        "main_ic",
    ]

    for ws in wb.worksheets:
        if ws.max_row < 1 or ws.max_column < 1:
            continue

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        header_to_col = {
            str(ws.cell(row=1, column=col_idx).value): col_idx
            for col_idx in range(1, ws.max_column + 1)
        }

        for header, col_idx in header_to_col.items():
            header_lower = header.lower()

            if header in percent_cols:
                fmt = "0.00%"
            elif header in integer_cols:
                fmt = "0"
            elif header in date_cols:
                fmt = "yyyy-mm-dd"
            elif any(k in header_lower for k in numeric_4_keywords):
                fmt = "0.0000"
            else:
                fmt = None

            if fmt is not None:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = fmt

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            scan_rows = min(ws.max_row, 3000)

            for row_idx in range(1, scan_rows + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))

            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 34)

    # Insert charts into a separate sheet if possible.
    try:
        from openpyxl.drawing.image import Image as XLImage

        if "Charts" in wb.sheetnames:
            del wb["Charts"]

        ws_chart = wb.create_sheet("Charts")

        chart_files = [
            ("A1", IC_DECAY_CURVE_PNG),
            ("A38", IC_DECAY_ICIR_PNG),
            ("A75", IC_DECAY_POSITIVE_RATIO_PNG),
        ]

        for anchor, img_path in chart_files:
            if Path(img_path).exists():
                img = XLImage(str(img_path))
                img.anchor = anchor
                ws_chart.add_image(img)

        ws_chart.sheet_view.showGridLines = False

    except Exception as e:
        print(f"Warning: charts were not inserted into Excel. Reason: {e}")

    wb.save(excel_path)


# =============================================================================
# 7. Run
# =============================================================================

def run_composite_signal_ic_decay_1_30():
    decay_summary, decay_daily, diagnostics, signal_mode_used, signal_label = (
        compute_composite_signal_ic_decay()
    )

    plot_ic_decay_curves(decay_summary)

    write_ic_decay_excel(
        decay_summary=decay_summary,
        decay_daily=decay_daily,
        diagnostics=diagnostics,
        signal_mode_used=signal_mode_used,
        signal_label=signal_label,
    )

    print("\nIC decay summary:")
    show_cols = [
    "horizon",

    # IC mean decay
    "rank_ic_mean",
    "neutral_rank_ic_mean",
    "winsor_ic_mean",

    # Final adjusted ICIR only
    "rank_ic_nw_icir",
    "ic_nw_icir",
    "neutral_rank_ic_nw_icir",
    "winsor_ic_nw_icir",

    # Direction stability
    "rank_ic_positive_ratio",
    "neutral_rank_ic_positive_ratio",

    # Sample diagnostics
    "sample_days",
    "avg_obs_per_day",
    "nw_lag",
    "nw_n_obs",
]

    print(decay_summary[[c for c in show_cols if c in decay_summary.columns]].to_string(index=False))


    print("\nDecay diagnostics:")
    print(diagnostics.to_string(index=False))

    return {
        "decay_summary": decay_summary,
        "decay_daily": decay_daily,
        "diagnostics": diagnostics,
        "signal_mode_used": signal_mode_used,
        "signal_label": signal_label,
    }


if __name__ == "__main__":
    ic_decay_result = run_composite_signal_ic_decay_1_30()