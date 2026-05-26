from __future__ import annotations

from pathlib import Path
from typing import Iterable, Sequence
import math
import re
import warnings

import numpy as np
import pandas as pd

try:
    from .config import *
except ImportError:  # allow running a file directly from this folder
    from config import *

try:
    from .core_base import *
    from .core_neutralization import *
    from .portfolio_five_group import *
    from .limit_transaction_cost import *
except ImportError:
    from core_base import *
    from core_neutralization import *
    from portfolio_five_group import *
    from limit_transaction_cost import *

# Teacher requirement:
# 1. Change Q1-Q5 grouping to Q1-Q10.
# 2. Check whether only the lowest layer is clearly different.
# 3. Within each layer, calculate limit-up / limit-down caused trade failure ratios.
#
# Main interpretation:
# - Q1 = lowest signal
# - Q10 = highest signal
# - If Q1 has high limit-down / failed-sell ratio, it supports the hypothesis that
#   the lowest layer's return separation is related to limit-down illiquidity.
# - If Q10 has high limit-up / failed-buy ratio, it means high-signal names may be
#   harder to buy after signals appear.

from pathlib import Path
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt


# =============================================================================
# CONFIG
# =============================================================================

DECILE_OUTDIR = OUTDIR / "decile_q1_q10_test_limit_failure"
DECILE_OUTDIR.mkdir(parents=True, exist_ok=True)

DECILE_EXCEL = DECILE_OUTDIR / "decile_q1_q10_test_limit_failure.xlsx"

DECILE_N_GROUPS = 10
DECILE_GROUPS = [f"Q{i}" for i in range(1, DECILE_N_GROUPS + 1)]

# Use same rebalance periods as previous five-group test.
DECILE_REBALANCE_PERIODS = GROUP_REBALANCE_PERIODS if "GROUP_REBALANCE_PERIODS" in globals() else [1, 5, 10, 20]

# Use same signal mode as your current five-group backtest.
# "current_backtest" follows global SIGNAL_MODE.
# You can also set directly: "weighted" or "equal".
DECILE_SIGNAL_MODE = "current_backtest"

DECILE_SIGNAL_EQUAL_COL = globals().get("GROUP_SIGNAL_EQUAL_COL", "group_signal_equal_mad_z")
DECILE_SIGNAL_WEIGHTED_COL = globals().get("GROUP_SIGNAL_WEIGHTED_COL", "group_signal_weighted_mad_z")

# Need enough stocks to split into 10 groups.
DECILE_MIN_STOCKS_TO_GROUP = max(
    globals().get("MIN_STOCKS_TO_GROUP", 50),
    DECILE_N_GROUPS * 10,
)

# Limit price detection settings.
# If auto detection fails, manually set:
# DECILE_LIMIT_UP_COL = "S_DQ_LIMIT"
# DECILE_LIMIT_DOWN_COL = "S_DQ_STOPPING"
DECILE_LIMIT_UP_COL = globals().get("LIMIT_UP_COL", None)
DECILE_LIMIT_DOWN_COL = globals().get("LIMIT_DOWN_COL", None)

DECILE_LIMIT_PRICE_RTOL = globals().get("LIMIT_PRICE_RTOL", 5e-4)
DECILE_LIMIT_PRICE_ATOL = globals().get("LIMIT_PRICE_ATOL", 1e-8)
DECILE_LIMIT_PRICE_IS_RAW = globals().get("LIMIT_PRICE_IS_RAW", True)

# For long-only Q1-Q10 portfolios, if sell is blocked, we scale down buys to avoid over-investment.
DECILE_ENFORCE_LONG_ONLY_CAPITAL_LIMIT = True

# Main summary excludes initial opening rebalance by default.
DECILE_EXCLUDE_INITIAL_REBALANCE_IN_FAILURE_SUMMARY = True


# =============================================================================
# 1. Resolve signal panel and signal column
# =============================================================================

def get_signal_panel_for_decile_test() -> pd.DataFrame:
    """
    Priority:
    1. Reuse five_group_result["signal_panel"] from previous five-group test.
    2. Otherwise rebuild signal panel using build_group_test_signal_panel().
    """
    if (
        "five_group_result" in globals()
        and isinstance(five_group_result, dict)
        and "signal_panel" in five_group_result
    ):
        print("Using five_group_result['signal_panel'] from memory.")
        panel = five_group_result["signal_panel"].copy()
    else:
        if "build_group_test_signal_panel" not in globals():
            raise NameError(
                "Cannot find five_group_result or build_group_test_signal_panel(). "
                "Please run the five-group signal construction chunk first."
            )

        print("five_group_result not found. Rebuilding signal panel ...")
        panel, _ = build_group_test_signal_panel()

    panel["trade_day"] = pd.to_datetime(panel["trade_day"]).dt.normalize()
    panel["code"] = panel["code"].astype(str).str.upper()

    return panel.sort_values(["code", "trade_day"]).reset_index(drop=True)


def resolve_decile_signal_col(signal_panel: pd.DataFrame) -> tuple[str, str]:
    """
    Return:
        signal_col, signal_mode_used
    """
    if DECILE_SIGNAL_MODE == "current_backtest":
        mode = globals().get("SIGNAL_MODE", "weighted")
    else:
        mode = DECILE_SIGNAL_MODE

    if mode == "weighted":
        signal_col = DECILE_SIGNAL_WEIGHTED_COL
    elif mode == "equal":
        signal_col = DECILE_SIGNAL_EQUAL_COL
    else:
        raise ValueError("DECILE_SIGNAL_MODE must be 'current_backtest', 'weighted', or 'equal'.")

    if signal_col not in signal_panel.columns:
        raise ValueError(
            f"Signal column {signal_col} not found in signal_panel. "
            f"Available columns are: {list(signal_panel.columns)}"
        )

    return signal_col, mode


# =============================================================================
# 2. Limit-up / limit-down status helper
# =============================================================================

DECILE_LIMIT_UP_COL_CANDIDATES = [
    "limit_up",
    "up_limit",
    "limit_up_price",
    "up_limit_price",
    "high_limit",
    "upper_limit",
    "price_limit_up",
    "limit_price_up",
    "zt_price",
    "zhangting_price",
    "涨停价",
    "涨停价格",
    "S_DQ_LIMIT",
    "s_dq_limit",
]

DECILE_LIMIT_DOWN_COL_CANDIDATES = [
    "limit_down",
    "down_limit",
    "limit_down_price",
    "down_limit_price",
    "low_limit",
    "lower_limit",
    "price_limit_down",
    "limit_price_down",
    "dt_price",
    "dieting_price",
    "跌停价",
    "跌停价格",
    "S_DQ_STOPPING",
    "s_dq_stopping",
]


def _decile_find_first_existing_column(
    available_cols: list[str],
    candidates: list[str],
    manual_col: str | None = None,
) -> str | None:
    """
    Find first matching column name, case-insensitive.
    """
    if manual_col is not None:
        if manual_col in available_cols:
            return manual_col
        raise ValueError(
            f"Manually specified column {manual_col} not found. "
            f"Available columns include: {available_cols[:80]}"
        )

    lower_map = {str(c).lower(): c for c in available_cols}

    for c in candidates:
        if c in available_cols:
            return c

        c_lower = str(c).lower()
        if c_lower in lower_map:
            return lower_map[c_lower]

    return None


def _decile_detect_limit_price_columns(files: list[Path]) -> tuple[str, str]:
    """
    Detect limit-up and limit-down price columns from parquet schema.
    """
    all_cols = []

    for fp in files:
        cols = _get_parquet_columns(fp)
        if cols is not None:
            all_cols.extend(cols)

    all_cols = list(dict.fromkeys(all_cols))

    if not all_cols:
        raise ValueError(
            "Cannot read parquet schema columns. "
            "Please manually set DECILE_LIMIT_UP_COL and DECILE_LIMIT_DOWN_COL."
        )

    up_col = _decile_find_first_existing_column(
        available_cols=all_cols,
        candidates=DECILE_LIMIT_UP_COL_CANDIDATES,
        manual_col=DECILE_LIMIT_UP_COL,
    )

    down_col = _decile_find_first_existing_column(
        available_cols=all_cols,
        candidates=DECILE_LIMIT_DOWN_COL_CANDIDATES,
        manual_col=DECILE_LIMIT_DOWN_COL,
    )

    if up_col is None or down_col is None:
        print("\nCould not automatically detect limit-up / limit-down columns.")
        print("Available parquet columns:")
        print(all_cols)

        raise ValueError(
            "Please manually set DECILE_LIMIT_UP_COL and DECILE_LIMIT_DOWN_COL near the top of this chunk. "
            "Example: DECILE_LIMIT_UP_COL = 'S_DQ_LIMIT', "
            "DECILE_LIMIT_DOWN_COL = 'S_DQ_STOPPING'."
        )

    print(f"Detected limit-up column   : {up_col}")
    print(f"Detected limit-down column : {down_col}")

    return up_col, down_col


def _decile_read_raw_limit_price_data(
    files: list[Path],
    limit_up_col: str,
    limit_down_col: str,
) -> pd.DataFrame:
    """
    Read parquet files with necessary columns for limit detection.
    """
    frames = []

    preferred_cols = [
        "code",
        "trade_day",
        "datetime",
        "close",
        "volume",
        "amount",
        "factor",
        limit_up_col,
        limit_down_col,
    ]

    for fp in files:
        print(f"Reading limit-price columns from {fp} ...")

        available_cols = _get_parquet_columns(fp)

        if available_cols is not None:
            cols = [c for c in preferred_cols if c in available_cols]
            df_part = pd.read_parquet(fp, columns=cols)
        else:
            df_part = pd.read_parquet(fp)
            cols = [c for c in preferred_cols if c in df_part.columns]
            df_part = df_part[cols].copy()

        missing = [c for c in [limit_up_col, limit_down_col] if c not in df_part.columns]
        if missing:
            raise ValueError(f"Missing limit columns in {fp}: {missing}")

        frames.append(df_part)

    return pd.concat(frames, ignore_index=True)


_DECILE_LIMIT_PANEL_CACHE = None


def build_decile_limit_price_panel() -> pd.DataFrame:
    """
    Build daily stock panel with:
    - code
    - trade_day
    - limit_up_price
    - limit_down_price
    - is_limit_up
    - is_limit_down
    """
    files = find_input_files(
        DATA_DIR,
        years=YEARS,
        file_pattern=FILE_PATTERN,
        allow_missing=ALLOW_MISSING_FILES,
    )

    limit_up_col, limit_down_col = _decile_detect_limit_price_columns(files)

    raw = _decile_read_raw_limit_price_data(
        files=files,
        limit_up_col=limit_up_col,
        limit_down_col=limit_down_col,
    )

    # prepare_data keeps extra columns and calculates vwap_price.
    df = prepare_data(
        raw,
        price_mode=PRICE_MODE,
        filter_a_share=FILTER_A_SHARE,
    )

    del raw

    if limit_up_col not in df.columns or limit_down_col not in df.columns:
        raise ValueError(
            f"After prepare_data(), cannot find {limit_up_col} / {limit_down_col}."
        )

    df[limit_up_col] = pd.to_numeric(df[limit_up_col], errors="coerce")
    df[limit_down_col] = pd.to_numeric(df[limit_down_col], errors="coerce")

    if (
        DECILE_LIMIT_PRICE_IS_RAW
        and PRICE_MODE == "adjusted"
        and "factor" in df.columns
    ):
        factor_adj = pd.to_numeric(df["factor"], errors="coerce").where(
            pd.to_numeric(df["factor"], errors="coerce").notna(),
            1.0,
        )

        df["limit_up_price"] = df[limit_up_col] * factor_adj
        df["limit_down_price"] = df[limit_down_col] * factor_adj
    else:
        df["limit_up_price"] = df[limit_up_col]
        df["limit_down_price"] = df[limit_down_col]

    df["vwap_price"] = pd.to_numeric(df["vwap_price"], errors="coerce")
    df["limit_up_price"] = pd.to_numeric(df["limit_up_price"], errors="coerce")
    df["limit_down_price"] = pd.to_numeric(df["limit_down_price"], errors="coerce")

    valid_up = (
        df["vwap_price"].notna()
        & df["limit_up_price"].notna()
        & (df["vwap_price"] > 0)
        & (df["limit_up_price"] > 0)
    )

    valid_down = (
        df["vwap_price"].notna()
        & df["limit_down_price"].notna()
        & (df["vwap_price"] > 0)
        & (df["limit_down_price"] > 0)
    )

    df["is_limit_up"] = False
    df.loc[valid_up, "is_limit_up"] = (
        np.isclose(
            df.loc[valid_up, "vwap_price"],
            df.loc[valid_up, "limit_up_price"],
            rtol=DECILE_LIMIT_PRICE_RTOL,
            atol=DECILE_LIMIT_PRICE_ATOL,
        )
        | (
            df.loc[valid_up, "vwap_price"]
            >= df.loc[valid_up, "limit_up_price"] * (1.0 - DECILE_LIMIT_PRICE_RTOL)
        )
    )

    df["is_limit_down"] = False
    df.loc[valid_down, "is_limit_down"] = (
        np.isclose(
            df.loc[valid_down, "vwap_price"],
            df.loc[valid_down, "limit_down_price"],
            rtol=DECILE_LIMIT_PRICE_RTOL,
            atol=DECILE_LIMIT_PRICE_ATOL,
        )
        | (
            df.loc[valid_down, "vwap_price"]
            <= df.loc[valid_down, "limit_down_price"] * (1.0 + DECILE_LIMIT_PRICE_RTOL)
        )
    )

    out = df[
        [
            "code",
            "trade_day",
            "limit_up_price",
            "limit_down_price",
            "is_limit_up",
            "is_limit_down",
        ]
    ].copy()

    out["trade_day"] = pd.to_datetime(out["trade_day"]).dt.normalize()
    out["code"] = out["code"].astype(str).str.upper()

    out = out.drop_duplicates(["code", "trade_day"], keep="last")

    print("\nLimit detection diagnostics:")
    print(f"Rows with limit-up flag   : {out['is_limit_up'].sum():,}")
    print(f"Rows with limit-down flag : {out['is_limit_down'].sum():,}")
    print(f"Total rows                : {len(out):,}")

    return out


def get_decile_limit_price_panel_cached() -> pd.DataFrame:
    global _DECILE_LIMIT_PANEL_CACHE

    if _DECILE_LIMIT_PANEL_CACHE is None:
        _DECILE_LIMIT_PANEL_CACHE = build_decile_limit_price_panel()

    return _DECILE_LIMIT_PANEL_CACHE.copy()


def add_limit_status_to_decile_signal_panel(signal_panel: pd.DataFrame) -> pd.DataFrame:
    """
    Add limit-up / limit-down status into signal panel.

    If your earlier limit chunk already defined add_limit_status_to_signal_panel(),
    this function uses that directly.
    """
    if "add_limit_status_to_signal_panel" in globals():
        print("Using existing add_limit_status_to_signal_panel() from previous limit chunk.")
        return add_limit_status_to_signal_panel(signal_panel)

    panel = signal_panel.copy()
    panel["trade_day"] = pd.to_datetime(panel["trade_day"]).dt.normalize()
    panel["code"] = panel["code"].astype(str).str.upper()

    limit_panel = get_decile_limit_price_panel_cached()

    drop_cols = [
        "limit_up_price",
        "limit_down_price",
        "is_limit_up",
        "is_limit_down",
    ]

    panel = panel.drop(columns=[c for c in drop_cols if c in panel.columns])

    panel = panel.merge(
        limit_panel,
        on=["code", "trade_day"],
        how="left",
    )

    panel["is_limit_up"] = panel["is_limit_up"].fillna(False).astype(bool)
    panel["is_limit_down"] = panel["is_limit_down"].fillna(False).astype(bool)

    return panel


# =============================================================================
# 3. Decile grouping and target weights
# =============================================================================

def assign_decile_groups_for_day(
    signal_day_df: pd.DataFrame,
    signal_col: str,
    n_groups: int = DECILE_N_GROUPS,
    min_stocks: int = DECILE_MIN_STOCKS_TO_GROUP,
) -> pd.DataFrame:
    """
    Assign Q1-Q10 based on signal ranking.

    Q1 = lowest signal
    Q10 = highest signal
    """
    x = signal_day_df.copy()

    if not INCLUDE_SUSPENDED and "volume" in x.columns:
        x = x[x["volume"].fillna(0) > 0].copy()

    x = x.dropna(subset=[signal_col]).copy()

    if len(x) < min_stocks:
        return pd.DataFrame(columns=["code", "group", signal_col])

    ranks = x[signal_col].rank(method="first")
    labels = [f"Q{i}" for i in range(1, n_groups + 1)]

    x["group"] = pd.qcut(
        ranks,
        q=n_groups,
        labels=labels,
    ).astype(str)

    return x[["code", "group", signal_col]].copy()


def build_decile_target_weights(group_df: pd.DataFrame) -> dict[str, pd.Series]:
    """
    Build equal-weight target portfolio for each decile.
    """
    targets = {}

    for group in DECILE_GROUPS:
        codes = (
            group_df.loc[group_df["group"] == group, "code"]
            .dropna()
            .astype(str)
            .str.upper()
            .drop_duplicates()
            .tolist()
        )

        if len(codes) == 0:
            targets[group] = pd.Series(dtype=float)
        else:
            targets[group] = pd.Series(1.0 / len(codes), index=codes, dtype=float)

    return targets


# =============================================================================
# 4. Decile portfolio backtest, without applying limit constraint
# =============================================================================

def build_decile_group_returns(
    signal_panel: pd.DataFrame,
    signal_col: str,
    rebalance_periods: list[int] = DECILE_REBALANCE_PERIODS,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Q1-Q10 equal-weight backtest.

    This is the clean decile version of your previous Q1-Q5 test.
    It does not apply limit-up / limit-down constraints to returns.
    Limit failure is separately diagnosed below.
    """
    panel = signal_panel.copy()
    panel["trade_day"] = pd.to_datetime(panel["trade_day"]).dt.normalize()
    panel["code"] = panel["code"].astype(str).str.upper()

    panel = panel.sort_values(["code", "trade_day"]).reset_index(drop=True)

    panel["ret_vwap_1d"] = panel.groupby("code", sort=False)["vwap_price"].pct_change()
    panel["ret_vwap_1d"] = panel["ret_vwap_1d"].replace([np.inf, -np.inf], np.nan)

    trading_days = sorted(panel["trade_day"].dropna().unique())

    ret_by_day = {
        day: sub.set_index("code")["ret_vwap_1d"]
        for day, sub in panel[["code", "trade_day", "ret_vwap_1d"]].groupby("trade_day")
    }

    signal_by_day = {
        day: sub[["code", "trade_day", "volume", signal_col]].copy()
        for day, sub in panel[["code", "trade_day", "volume", signal_col]].groupby("trade_day")
    }

    daily_return_records = []
    rebalance_records = []

    for h in rebalance_periods:
        print(f"\nRunning Q1-Q10 decile backtest, rebalance_period={h}d ...")

        valid_start_end = len(trading_days) - h

        for start_i in range(0, valid_start_end, h):
            signal_day = trading_days[start_i]

            if start_i + h + 1 >= len(trading_days):
                break

            entry_day = trading_days[start_i + 1]
            exit_day = trading_days[start_i + h + 1]

            signal_day_df = signal_by_day.get(signal_day)

            if signal_day_df is None or signal_day_df.empty:
                continue

            group_df = assign_decile_groups_for_day(
                signal_day_df=signal_day_df,
                signal_col=signal_col,
            )

            if group_df.empty:
                continue

            if not set(DECILE_GROUPS).issubset(set(group_df["group"].unique())):
                continue

            group_members = {
                group: group_df.loc[group_df["group"] == group, "code"].astype(str).str.upper().tolist()
                for group in DECILE_GROUPS
            }

            for group, codes in group_members.items():
                temp = group_df[group_df["group"] == group]

                rebalance_records.append(
                    {
                        "horizon": h,
                        "signal_day": signal_day,
                        "entry_day": entry_day,
                        "exit_day": exit_day,
                        "group": group,
                        "n_stocks": len(codes),
                        "avg_signal": temp[signal_col].mean(),
                        "median_signal": temp[signal_col].median(),
                        "min_signal": temp[signal_col].min(),
                        "max_signal": temp[signal_col].max(),
                    }
                )

            for interval_i in range(start_i + 1, start_i + h + 1):
                return_day = trading_days[interval_i + 1]
                ret_series = ret_by_day.get(return_day)

                if ret_series is None:
                    continue

                for group, codes in group_members.items():
                    stock_ret = ret_series.reindex(codes).dropna()

                    if stock_ret.empty:
                        group_ret = np.nan
                        n_ret_obs = 0
                    else:
                        group_ret = float(stock_ret.mean())
                        n_ret_obs = int(stock_ret.shape[0])

                    daily_return_records.append(
                        {
                            "horizon": h,
                            "trade_day": return_day,
                            "signal_day": signal_day,
                            "entry_day": entry_day,
                            "exit_day": exit_day,
                            "group": group,
                            "ret": group_ret,
                            "n_ret_obs": n_ret_obs,
                            "n_stocks_at_rebalance": len(codes),
                        }
                    )

    daily_returns = pd.DataFrame(daily_return_records)
    rebalance_info = pd.DataFrame(rebalance_records)

    if daily_returns.empty:
        raise ValueError("No decile daily returns generated. Please check signal coverage.")

    return daily_returns, rebalance_info, panel


def decile_max_drawdown(nav: pd.Series) -> float:
    nav = pd.Series(nav).dropna()

    if nav.empty:
        return np.nan

    running_max = nav.cummax()
    drawdown = nav / running_max - 1.0

    return float(drawdown.min())


def build_decile_nav_and_summary(
    daily_returns: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Build cumulative NAV and performance summary for Q1-Q10.
    """
    nav_frames = []
    summary_records = []

    for h, sub in daily_returns.groupby("horizon"):
        wide = (
            sub.pivot_table(
                index="trade_day",
                columns="group",
                values="ret",
                aggfunc="mean",
            )
            .sort_index()
        )

        for group in DECILE_GROUPS:
            if group not in wide.columns:
                wide[group] = np.nan

        wide = wide[DECILE_GROUPS].copy()

        wide["Q10_minus_Q1"] = wide["Q10"] - wide["Q1"]

        ret_cols = DECILE_GROUPS + ["Q10_minus_Q1"]

        nav = (1.0 + wide[ret_cols].fillna(0.0)).cumprod()
        nav.insert(0, "trade_day", nav.index)
        nav.insert(0, "horizon", h)

        nav_frames.append(nav.reset_index(drop=True))

        for col in ret_cols:
            r = wide[col].dropna()

            if r.empty:
                summary_records.append(
                    {
                        "horizon": h,
                        "portfolio": col,
                        "n_days": 0,
                        "total_return": np.nan,
                        "annual_return": np.nan,
                        "annual_vol": np.nan,
                        "sharpe_no_rf": np.nan,
                        "max_drawdown": np.nan,
                        "avg_daily_return": np.nan,
                        "daily_win_rate": np.nan,
                    }
                )
                continue

            nav_col = (1.0 + r).cumprod()
            n_days = len(r)

            total_return = float(nav_col.iloc[-1] - 1.0)
            annual_return = float(nav_col.iloc[-1] ** (252.0 / n_days) - 1.0)
            annual_vol = float(r.std(ddof=1) * np.sqrt(252.0))
            sharpe = annual_return / annual_vol if annual_vol and annual_vol > 0 else np.nan

            summary_records.append(
                {
                    "horizon": h,
                    "portfolio": col,
                    "n_days": n_days,
                    "total_return": total_return,
                    "annual_return": annual_return,
                    "annual_vol": annual_vol,
                    "sharpe_no_rf": sharpe,
                    "max_drawdown": decile_max_drawdown(nav_col),
                    "avg_daily_return": float(r.mean()),
                    "daily_win_rate": float((r > 0).mean()),
                }
            )

    nav_df = pd.concat(nav_frames, ignore_index=True)
    summary_df = pd.DataFrame(summary_records)

    return nav_df, summary_df


# =============================================================================
# 5. Limit failure calculation
# =============================================================================

def _decile_clean_weight_series(w: pd.Series) -> pd.Series:
    if w is None or len(w) == 0:
        return pd.Series(dtype=float)

    out = pd.Series(w, dtype=float).copy()
    out.index = out.index.astype(str).str.upper()

    out = out.replace([np.inf, -np.inf], np.nan).dropna()
    out = out[out.abs() > 1e-15]

    return out


def _decile_execute_rebalance_with_limit_failure_stats(
    old_weights: pd.Series,
    desired_weights: pd.Series,
    entry_limit_status: pd.DataFrame,
) -> tuple[pd.Series, dict]:
    """
    Compare old weights with desired weights and identify failed trades.

    Rule:
    - If delta > 0, it is a buy.
      Buy fails if stock is limit-up on entry day.
    - If delta < 0, it is a sell.
      Sell fails if stock is limit-down on entry day.
    """
    old_w = _decile_clean_weight_series(old_weights)
    desired_w = _decile_clean_weight_series(desired_weights)

    all_codes = old_w.index.union(desired_w.index)

    old_aligned = old_w.reindex(all_codes).fillna(0.0)
    desired_aligned = desired_w.reindex(all_codes).fillna(0.0)

    desired_delta = desired_aligned - old_aligned

    if entry_limit_status is None or len(entry_limit_status) == 0:
        is_limit_up = pd.Series(False, index=all_codes)
        is_limit_down = pd.Series(False, index=all_codes)
    else:
        lim = entry_limit_status.copy()

        if "code" in lim.columns:
            lim["code"] = lim["code"].astype(str).str.upper()
            lim = lim.set_index("code")

        lim.index = lim.index.astype(str).str.upper()

        is_limit_up = lim.get("is_limit_up", pd.Series(False, index=lim.index))
        is_limit_down = lim.get("is_limit_down", pd.Series(False, index=lim.index))

        is_limit_up = is_limit_up.reindex(all_codes).fillna(False).astype(bool)
        is_limit_down = is_limit_down.reindex(all_codes).fillna(False).astype(bool)

    eps = 1e-15

    buy_mask = desired_delta > eps
    sell_mask = desired_delta < -eps

    buy_blocked = buy_mask & is_limit_up
    sell_blocked = sell_mask & is_limit_down

    desired_buy_turnover = float(desired_delta.clip(lower=0.0).sum())
    desired_sell_turnover = float((-desired_delta.clip(upper=0.0)).sum())
    desired_turnover = desired_buy_turnover + desired_sell_turnover

    failed_buy_turnover = float(desired_delta.where(buy_blocked, 0.0).clip(lower=0.0).sum())
    failed_sell_turnover = float((-desired_delta.where(sell_blocked, 0.0).clip(upper=0.0).sum()))
    failed_turnover = failed_buy_turnover + failed_sell_turnover

    allowed_delta = desired_delta.copy()
    allowed_delta.loc[buy_blocked | sell_blocked] = 0.0

    capital_scaled_buy_turnover = 0.0
    capital_scale = 1.0

    if DECILE_ENFORCE_LONG_ONLY_CAPITAL_LIMIT:
        sell_delta = allowed_delta.clip(upper=0.0)
        buy_delta = allowed_delta.clip(lower=0.0)

        after_allowed_sells = old_aligned + sell_delta
        after_allowed_sells = after_allowed_sells.clip(lower=0.0)

        target_long_exposure = desired_aligned[desired_aligned > 0].sum()

        if not np.isfinite(target_long_exposure) or target_long_exposure <= 0:
            target_long_exposure = 1.0

        current_exposure_after_sells = after_allowed_sells.sum()
        available_buy_capacity = max(target_long_exposure - current_exposure_after_sells, 0.0)

        buy_sum = buy_delta.sum()

        if buy_sum > available_buy_capacity + eps:
            capital_scale = available_buy_capacity / buy_sum if buy_sum > 0 else 0.0
            scaled_buy_delta = buy_delta * capital_scale
            capital_scaled_buy_turnover = float(buy_sum - scaled_buy_delta.sum())
            allowed_delta = sell_delta + scaled_buy_delta

    actual_aligned = old_aligned + allowed_delta
    actual_aligned = actual_aligned.where(actual_aligned.abs() > eps, 0.0)

    actual_w = _decile_clean_weight_series(actual_aligned)

    buy_order_count = int(buy_mask.sum())
    sell_order_count = int(sell_mask.sum())
    trade_order_count = buy_order_count + sell_order_count

    failed_buy_count = int(buy_blocked.sum())
    failed_sell_count = int(sell_blocked.sum())
    failed_trade_count = failed_buy_count + failed_sell_count

    info = {
        "desired_buy_turnover": desired_buy_turnover,
        "desired_sell_turnover": desired_sell_turnover,
        "desired_turnover": desired_turnover,

        "failed_buy_turnover": failed_buy_turnover,
        "failed_sell_turnover": failed_sell_turnover,
        "failed_turnover": failed_turnover,

        "failed_buy_turnover_ratio": (
            failed_buy_turnover / desired_buy_turnover
            if desired_buy_turnover > 1e-12 else np.nan
        ),
        "failed_sell_turnover_ratio": (
            failed_sell_turnover / desired_sell_turnover
            if desired_sell_turnover > 1e-12 else np.nan
        ),
        "failed_turnover_ratio": (
            failed_turnover / desired_turnover
            if desired_turnover > 1e-12 else np.nan
        ),

        "buy_order_count": buy_order_count,
        "sell_order_count": sell_order_count,
        "trade_order_count": trade_order_count,

        "failed_buy_count": failed_buy_count,
        "failed_sell_count": failed_sell_count,
        "failed_trade_count": failed_trade_count,

        "failed_buy_count_ratio": (
            failed_buy_count / buy_order_count
            if buy_order_count > 0 else np.nan
        ),
        "failed_sell_count_ratio": (
            failed_sell_count / sell_order_count
            if sell_order_count > 0 else np.nan
        ),
        "failed_trade_count_ratio": (
            failed_trade_count / trade_order_count
            if trade_order_count > 0 else np.nan
        ),

        "capital_scaled_buy_turnover": capital_scaled_buy_turnover,
        "capital_scale": capital_scale,

        "desired_long_exposure": float(desired_aligned[desired_aligned > 0].sum()),
        "actual_long_exposure": float(actual_aligned[actual_aligned > 0].sum()),
        "actual_gross_exposure": float(actual_aligned.abs().sum()),
        "actual_net_exposure": float(actual_aligned.sum()),
    }

    return actual_w, info


def _decile_drift_weights_one_day(
    weights: pd.Series,
    ret_series: pd.Series,
) -> pd.Series:
    """
    Drift risky weights by one-day stock returns.
    Cash weight is implicit and earns zero.
    """
    w = _decile_clean_weight_series(weights)

    if w.empty:
        return w

    r = pd.to_numeric(ret_series.reindex(w.index), errors="coerce")
    r = r.replace([np.inf, -np.inf], np.nan).fillna(0.0)

    port_ret = float((w * r).sum())

    if not np.isfinite(port_ret) or not np.isfinite(1.0 + port_ret) or (1.0 + port_ret) <= 0:
        return w

    drifted = w * (1.0 + r) / (1.0 + port_ret)

    return _decile_clean_weight_series(drifted)


def _decile_target_limit_stats(
    target_weights: pd.Series,
    entry_limit_status: pd.DataFrame,
) -> dict:
    """
    Calculate limit-up / limit-down fraction among target group constituents.
    """
    target_w = _decile_clean_weight_series(target_weights)
    target_codes = target_w.index.astype(str).str.upper()

    if len(target_codes) == 0:
        return {
            "target_n_names": 0,
            "target_limit_up_count": 0,
            "target_limit_down_count": 0,
            "target_limit_up_ratio": np.nan,
            "target_limit_down_ratio": np.nan,
        }

    if entry_limit_status is None or len(entry_limit_status) == 0:
        return {
            "target_n_names": int(len(target_codes)),
            "target_limit_up_count": 0,
            "target_limit_down_count": 0,
            "target_limit_up_ratio": 0.0,
            "target_limit_down_ratio": 0.0,
        }

    lim = entry_limit_status.copy()

    if "code" in lim.columns:
        lim["code"] = lim["code"].astype(str).str.upper()
        lim = lim.set_index("code")

    lim.index = lim.index.astype(str).str.upper()

    up = lim.get("is_limit_up", pd.Series(False, index=lim.index)).reindex(target_codes).fillna(False).astype(bool)
    down = lim.get("is_limit_down", pd.Series(False, index=lim.index)).reindex(target_codes).fillna(False).astype(bool)

    up_count = int(up.sum())
    down_count = int(down.sum())
    n = int(len(target_codes))

    return {
        "target_n_names": n,
        "target_limit_up_count": up_count,
        "target_limit_down_count": down_count,
        "target_limit_up_ratio": up_count / n if n > 0 else np.nan,
        "target_limit_down_ratio": down_count / n if n > 0 else np.nan,
    }


def calculate_decile_limit_failure_stats(
    signal_panel_with_limit: pd.DataFrame,
    signal_col: str,
    rebalance_periods: list[int] = DECILE_REBALANCE_PERIODS,
) -> pd.DataFrame:
    """
    For Q1-Q10, calculate trade failure ratios caused by limit-up / limit-down.

    This diagnostic simulates actual blocked trades only to measure failure ratios
    and actual exposure. It is not meant to replace the clean decile NAV test above.
    """
    panel = signal_panel_with_limit.copy()
    panel["trade_day"] = pd.to_datetime(panel["trade_day"]).dt.normalize()
    panel["code"] = panel["code"].astype(str).str.upper()

    panel = panel.sort_values(["code", "trade_day"]).reset_index(drop=True)

    panel["ret_vwap_1d"] = panel.groupby("code", sort=False)["vwap_price"].pct_change()
    panel["ret_vwap_1d"] = panel["ret_vwap_1d"].replace([np.inf, -np.inf], np.nan)

    trading_days = sorted(panel["trade_day"].dropna().unique())

    ret_by_day = {
        day: sub.set_index("code")["ret_vwap_1d"]
        for day, sub in panel[["code", "trade_day", "ret_vwap_1d"]].groupby("trade_day")
    }

    signal_by_day = {
        day: sub[["code", "trade_day", "volume", signal_col]].copy()
        for day, sub in panel[["code", "trade_day", "volume", signal_col]].groupby("trade_day")
    }

    limit_by_day = {
        day: sub[[
            "code",
            "trade_day",
            "is_limit_up",
            "is_limit_down",
            "limit_up_price",
            "limit_down_price",
        ]].copy()
        for day, sub in panel[[
            "code",
            "trade_day",
            "is_limit_up",
            "is_limit_down",
            "limit_up_price",
            "limit_down_price",
        ]].groupby("trade_day")
    }

    records = []

    for h in rebalance_periods:
        print(f"\nCalculating Q1-Q10 limit failure stats, rebalance_period={h}d ...")

        current_weights = {g: pd.Series(dtype=float) for g in DECILE_GROUPS}

        rebalance_id = 0
        valid_start_end = len(trading_days) - h

        for start_i in range(0, valid_start_end, h):
            signal_day = trading_days[start_i]

            if start_i + h + 1 >= len(trading_days):
                break

            entry_day = trading_days[start_i + 1]
            exit_day = trading_days[start_i + h + 1]

            signal_day_df = signal_by_day.get(signal_day)

            if signal_day_df is None or signal_day_df.empty:
                continue

            group_df = assign_decile_groups_for_day(
                signal_day_df=signal_day_df,
                signal_col=signal_col,
            )

            if group_df.empty:
                continue

            if not set(DECILE_GROUPS).issubset(set(group_df["group"].unique())):
                continue

            target_weights = build_decile_target_weights(group_df)
            entry_limit_status = limit_by_day.get(entry_day, pd.DataFrame())

            rebalance_id += 1

            for group in DECILE_GROUPS:
                old_w = current_weights[group]
                desired_w = target_weights[group]

                is_initial_rebalance = old_w.empty

                actual_w, failure_info = _decile_execute_rebalance_with_limit_failure_stats(
                    old_weights=old_w,
                    desired_weights=desired_w,
                    entry_limit_status=entry_limit_status,
                )

                target_limit_info = _decile_target_limit_stats(
                    target_weights=desired_w,
                    entry_limit_status=entry_limit_status,
                )

                group_signal = group_df[group_df["group"] == group][signal_col]

                records.append(
                    {
                        "horizon": int(h),
                        "rebalance_id": rebalance_id,
                        "signal_day": signal_day,
                        "entry_day": entry_day,
                        "exit_day": exit_day,
                        "group": group,
                        "is_initial_rebalance": is_initial_rebalance,

                        "avg_signal": float(group_signal.mean()),
                        "median_signal": float(group_signal.median()),
                        "min_signal": float(group_signal.min()),
                        "max_signal": float(group_signal.max()),

                        **target_limit_info,
                        **failure_info,
                    }
                )

                current_weights[group] = actual_w.copy()

            # Drift actual weights during the holding period.
            for interval_i in range(start_i + 1, start_i + h + 1):
                return_day = trading_days[interval_i + 1]
                ret_series = ret_by_day.get(return_day)

                if ret_series is None:
                    continue

                for group in DECILE_GROUPS:
                    current_weights[group] = _decile_drift_weights_one_day(
                        weights=current_weights[group],
                        ret_series=ret_series,
                    )

    failure_by_rebalance = pd.DataFrame(records)

    if failure_by_rebalance.empty:
        raise ValueError("No decile limit failure records generated.")

    failure_by_rebalance = failure_by_rebalance.sort_values(
        ["horizon", "rebalance_id", "group"]
    ).reset_index(drop=True)

    return failure_by_rebalance


# =============================================================================
# 6. Failure summary and comparison tables
# =============================================================================

def _safe_divide(a, b):
    return np.where(np.abs(b) > 1e-12, a / b, np.nan)


def build_decile_limit_failure_summary(
    failure_by_rebalance: pd.DataFrame,
    exclude_initial: bool = DECILE_EXCLUDE_INITIAL_REBALANCE_IN_FAILURE_SUMMARY,
) -> pd.DataFrame:
    """
    Build summary by horizon and group.
    """
    work = failure_by_rebalance.copy()

    if exclude_initial:
        work = work[~work["is_initial_rebalance"].astype(bool)].copy()

    if work.empty:
        print("Warning: no non-initial rebalances. Using all rebalances instead.")
        work = failure_by_rebalance.copy()

    summary = (
        work
        .groupby(["horizon", "group"], as_index=False)
        .agg(
            rebalance_count=("rebalance_id", "count"),

            avg_target_n_names=("target_n_names", "mean"),

            avg_target_limit_up_ratio=("target_limit_up_ratio", "mean"),
            avg_target_limit_down_ratio=("target_limit_down_ratio", "mean"),
            total_target_limit_up_count=("target_limit_up_count", "sum"),
            total_target_limit_down_count=("target_limit_down_count", "sum"),

            total_desired_buy_turnover=("desired_buy_turnover", "sum"),
            total_desired_sell_turnover=("desired_sell_turnover", "sum"),
            total_desired_turnover=("desired_turnover", "sum"),

            total_failed_buy_turnover=("failed_buy_turnover", "sum"),
            total_failed_sell_turnover=("failed_sell_turnover", "sum"),
            total_failed_turnover=("failed_turnover", "sum"),

            avg_failed_buy_turnover_ratio=("failed_buy_turnover_ratio", "mean"),
            avg_failed_sell_turnover_ratio=("failed_sell_turnover_ratio", "mean"),
            avg_failed_turnover_ratio=("failed_turnover_ratio", "mean"),

            total_buy_order_count=("buy_order_count", "sum"),
            total_sell_order_count=("sell_order_count", "sum"),
            total_trade_order_count=("trade_order_count", "sum"),

            total_failed_buy_count=("failed_buy_count", "sum"),
            total_failed_sell_count=("failed_sell_count", "sum"),
            total_failed_trade_count=("failed_trade_count", "sum"),

            avg_failed_buy_count_ratio=("failed_buy_count_ratio", "mean"),
            avg_failed_sell_count_ratio=("failed_sell_count_ratio", "mean"),
            avg_failed_trade_count_ratio=("failed_trade_count_ratio", "mean"),

            avg_actual_long_exposure=("actual_long_exposure", "mean"),
            avg_actual_gross_exposure=("actual_gross_exposure", "mean"),
            avg_actual_net_exposure=("actual_net_exposure", "mean"),

            avg_capital_scaled_buy_turnover=("capital_scaled_buy_turnover", "mean"),
        )
    )

    summary["total_failed_buy_turnover_ratio"] = _safe_divide(
        summary["total_failed_buy_turnover"],
        summary["total_desired_buy_turnover"],
    )

    summary["total_failed_sell_turnover_ratio"] = _safe_divide(
        summary["total_failed_sell_turnover"],
        summary["total_desired_sell_turnover"],
    )

    summary["total_failed_turnover_ratio"] = _safe_divide(
        summary["total_failed_turnover"],
        summary["total_desired_turnover"],
    )

    summary["total_failed_buy_count_ratio"] = _safe_divide(
        summary["total_failed_buy_count"],
        summary["total_buy_order_count"],
    )

    summary["total_failed_sell_count_ratio"] = _safe_divide(
        summary["total_failed_sell_count"],
        summary["total_sell_order_count"],
    )

    summary["total_failed_trade_count_ratio"] = _safe_divide(
        summary["total_failed_trade_count"],
        summary["total_trade_order_count"],
    )

    summary["group_order"] = summary["group"].str.extract(r"Q(\d+)").astype(int)
    summary = summary.sort_values(["horizon", "group_order"]).drop(columns=["group_order"]).reset_index(drop=True)

    return summary


def build_decile_extreme_layer_check(
    failure_summary: pd.DataFrame,
    performance_summary: pd.DataFrame,
) -> pd.DataFrame:
    """
    Compare Q1 with Q2-Q10 average, and Q10 with Q1.

    This sheet is the most useful one for teacher discussion.
    """
    metrics = [
        "avg_target_limit_down_ratio",
        "avg_target_limit_up_ratio",
        "total_failed_sell_turnover_ratio",
        "total_failed_buy_turnover_ratio",
        "total_failed_turnover_ratio",
        "total_failed_sell_count_ratio",
        "total_failed_buy_count_ratio",
        "avg_actual_gross_exposure",
    ]

    records = []

    perf = performance_summary.copy()

    for h, sub in failure_summary.groupby("horizon"):
        q1 = sub[sub["group"] == "Q1"]
        q10 = sub[sub["group"] == "Q10"]
        q2_q10 = sub[sub["group"].isin([f"Q{i}" for i in range(2, 11)])]
        q2_q9 = sub[sub["group"].isin([f"Q{i}" for i in range(2, 10)])]

        if q1.empty:
            continue

        row = {"horizon": int(h)}

        for m in metrics:
            q1_val = float(q1.iloc[0][m]) if m in q1.columns else np.nan
            q10_val = float(q10.iloc[0][m]) if (not q10.empty and m in q10.columns) else np.nan
            q2_q10_avg = float(q2_q10[m].mean()) if (not q2_q10.empty and m in q2_q10.columns) else np.nan
            q2_q9_avg = float(q2_q9[m].mean()) if (not q2_q9.empty and m in q2_q9.columns) else np.nan

            row[f"q1_{m}"] = q1_val
            row[f"q10_{m}"] = q10_val
            row[f"q2_q10_avg_{m}"] = q2_q10_avg
            row[f"q2_q9_avg_{m}"] = q2_q9_avg

            row[f"q1_minus_q2_q10_avg_{m}"] = q1_val - q2_q10_avg if pd.notna(q2_q10_avg) else np.nan
            row[f"q1_div_q2_q10_avg_{m}"] = (
                q1_val / q2_q10_avg if pd.notna(q2_q10_avg) and abs(q2_q10_avg) > 1e-12 else np.nan
            )

        # Add performance of Q1 / Q2-Q10 / Q10 if available.
        perf_h = perf[perf["horizon"] == h].copy()

        if not perf_h.empty:
            for p in ["Q1", "Q10", "Q10_minus_Q1"]:
                p_row = perf_h[perf_h["portfolio"] == p]

                if not p_row.empty:
                    row[f"{p}_annual_return"] = p_row.iloc[0].get("annual_return", np.nan)
                    row[f"{p}_sharpe_no_rf"] = p_row.iloc[0].get("sharpe_no_rf", np.nan)
                    row[f"{p}_max_drawdown"] = p_row.iloc[0].get("max_drawdown", np.nan)
                    row[f"{p}_avg_daily_return"] = p_row.iloc[0].get("avg_daily_return", np.nan)

            middle_perf = perf_h[perf_h["portfolio"].isin([f"Q{i}" for i in range(2, 10)])]
            if not middle_perf.empty:
                row["Q2_Q9_avg_annual_return"] = middle_perf["annual_return"].mean()
                row["Q2_Q9_avg_avg_daily_return"] = middle_perf["avg_daily_return"].mean()

        records.append(row)

    out = pd.DataFrame(records)

    if not out.empty:
        out = out.sort_values("horizon").reset_index(drop=True)

    return out


# =============================================================================
# 7. Plot
# =============================================================================

def plot_decile_nav_curves(nav_df: pd.DataFrame, output_dir: Path):
    output_dir.mkdir(parents=True, exist_ok=True)

    plot_cols = DECILE_GROUPS + ["Q10_minus_Q1"]

    for h, sub in nav_df.groupby("horizon"):
        sub = sub.sort_values("trade_day").copy()
        sub["trade_day"] = pd.to_datetime(sub["trade_day"])

        fig, ax = plt.subplots(figsize=(13, 7))

        for col in plot_cols:
            if col in sub.columns:
                ax.plot(sub["trade_day"], sub[col], label=col)

        ax.set_title(f"Q1-Q10 decile portfolio cumulative return, rebalance every {h} trading day(s)")
        ax.set_xlabel("Date")
        ax.set_ylabel("Cumulative NAV")
        ax.legend(ncol=2)
        ax.grid(True, alpha=0.3)

        fig.tight_layout()

        fig_path = output_dir / f"decile_q1_q10_nav_h{h}.png"
        fig.savefig(fig_path, dpi=300, bbox_inches="tight")
        plt.close(fig)

        print(f"Saved plot: {fig_path}")


def plot_decile_limit_failure_bars(failure_summary: pd.DataFrame, output_dir: Path):
    output_dir.mkdir(parents=True, exist_ok=True)

    metrics = [
        (
            "avg_target_limit_down_ratio",
            "Average Target Constituents Limit-down Ratio",
            "avg_target_limit_down_ratio",
        ),
        (
            "total_failed_sell_turnover_ratio",
            "Total Failed Sell Turnover Ratio",
            "failed_sell_turnover_ratio",
        ),
        (
            "avg_target_limit_up_ratio",
            "Average Target Constituents Limit-up Ratio",
            "avg_target_limit_up_ratio",
        ),
        (
            "total_failed_buy_turnover_ratio",
            "Total Failed Buy Turnover Ratio",
            "failed_buy_turnover_ratio",
        ),
        (
            "total_failed_turnover_ratio",
            "Total Failed Turnover Ratio",
            "failed_total_turnover_ratio",
        ),
    ]

    for metric, title, file_stub in metrics:
        if metric not in failure_summary.columns:
            continue

        for h, sub in failure_summary.groupby("horizon"):
            sub = sub.copy()
            sub["group_order"] = sub["group"].str.extract(r"Q(\d+)").astype(int)
            sub = sub.sort_values("group_order")

            fig, ax = plt.subplots(figsize=(10, 5))

            ax.bar(sub["group"], sub[metric])

            ax.set_title(f"{title}, rebalance every {h} trading day(s)")
            ax.set_xlabel("Decile group")
            ax.set_ylabel(metric)
            ax.grid(True, axis="y", alpha=0.3)

            fig.tight_layout()

            fig_path = output_dir / f"{file_stub}_h{h}.png"
            fig.savefig(fig_path, dpi=300, bbox_inches="tight")
            plt.close(fig)

            print(f"Saved plot: {fig_path}")


# =============================================================================
# 8. Save Excel and format
# =============================================================================

def save_decile_outputs(
    decile_performance_summary: pd.DataFrame,
    decile_nav: pd.DataFrame,
    decile_daily_returns: pd.DataFrame,
    decile_rebalance_info: pd.DataFrame,
    failure_by_rebalance: pd.DataFrame,
    failure_summary: pd.DataFrame,
    failure_summary_all: pd.DataFrame,
    extreme_layer_check: pd.DataFrame,
    signal_col: str,
    signal_mode_used: str,
):
    config = pd.DataFrame(
        [
            {"item": "signal_mode_used", "value": signal_mode_used},
            {"item": "signal_column", "value": signal_col},
            {"item": "n_groups", "value": DECILE_N_GROUPS},
            {"item": "groups", "value": str(DECILE_GROUPS)},
            {"item": "rebalance_periods", "value": str(DECILE_REBALANCE_PERIODS)},
            {"item": "min_stocks_to_group", "value": DECILE_MIN_STOCKS_TO_GROUP},
            {"item": "Q1", "value": "lowest signal"},
            {"item": "Q10", "value": "highest signal"},
            {"item": "execution_assumption", "value": "signal at t, enter at VWAP[t+1]"},
            {"item": "buy_failure_rule", "value": "buy delta fails if entry day is limit-up"},
            {"item": "sell_failure_rule", "value": "sell delta fails if entry day is limit-down"},
            {"item": "limit_price_rtol", "value": DECILE_LIMIT_PRICE_RTOL},
            {"item": "limit_price_atol", "value": DECILE_LIMIT_PRICE_ATOL},
            {"item": "failure_summary_excludes_initial_rebalance", "value": DECILE_EXCLUDE_INITIAL_REBALANCE_IN_FAILURE_SUMMARY},
        ]
    )

    methodology = pd.DataFrame(
        [
            {
                "Item": "Decile grouping",
                "Description": "Stocks are sorted by composite signal each rebalance date and split into Q1-Q10. Q1 is the lowest-signal group and Q10 is the highest-signal group.",
            },
            {
                "Item": "Clean decile backtest",
                "Description": "Q1-Q10 NAV uses equal-weight groups without applying limit-up / limit-down execution constraints. This is used to check whether only Q1 is separated.",
            },
            {
                "Item": "Target limit-down ratio",
                "Description": "Share of target group constituents that are limit-down on the entry day.",
            },
            {
                "Item": "Target limit-up ratio",
                "Description": "Share of target group constituents that are limit-up on the entry day.",
            },
            {
                "Item": "Failed sell turnover ratio",
                "Description": "Among desired sell turnover, the fraction that cannot be executed because stocks are limit-down on the entry day.",
            },
            {
                "Item": "Failed buy turnover ratio",
                "Description": "Among desired buy turnover, the fraction that cannot be executed because stocks are limit-up on the entry day.",
            },
            {
                "Item": "Initial rebalance",
                "Description": "Main failure summary excludes initial opening rebalance because opening from cash is mechanically all buy orders.",
            },
            {
                "Item": "Main diagnostic",
                "Description": "If Q1 has higher target limit-down ratio and higher failed sell turnover ratio than Q2-Q10, the Q1 separation may be related to downside illiquidity / limit-down sell failure.",
            },
        ]
    )

    with pd.ExcelWriter(DECILE_EXCEL, engine="openpyxl") as writer:
        extreme_layer_check.to_excel(writer, sheet_name="Extreme_Layer_Check", index=False)
        failure_summary.to_excel(writer, sheet_name="Limit_Failure_Summary", index=False)
        failure_summary_all.to_excel(writer, sheet_name="Failure_Summary_All", index=False)
        decile_performance_summary.to_excel(writer, sheet_name="Performance_Summary", index=False)
        decile_nav.to_excel(writer, sheet_name="Cumulative_NAV", index=False)
        decile_rebalance_info.to_excel(writer, sheet_name="Rebalance_Info", index=False)
        failure_by_rebalance.to_excel(writer, sheet_name="Failure_By_Rebalance", index=False)
        decile_daily_returns.to_excel(writer, sheet_name="Daily_Group_Returns", index=False)
        config.to_excel(writer, sheet_name="Config", index=False)
        methodology.to_excel(writer, sheet_name="Methodology", index=False)

    format_decile_excel(DECILE_EXCEL)

    print(f"Saved decile Excel: {DECILE_EXCEL}")


def format_decile_excel(excel_path: Path):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(excel_path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    percent_keywords = [
        "ratio",
        "return",
        "vol",
        "drawdown",
        "turnover",
        "exposure",
        "daily_win_rate",
    ]

    integer_cols = {
        "horizon",
        "rebalance_id",
        "rebalance_count",
        "target_n_names",
        "target_limit_up_count",
        "target_limit_down_count",
        "buy_order_count",
        "sell_order_count",
        "trade_order_count",
        "failed_buy_count",
        "failed_sell_count",
        "failed_trade_count",
        "n_days",
        "n_ret_obs",
        "n_stocks",
        "n_stocks_at_rebalance",
    }

    date_cols = {
        "trade_day",
        "signal_day",
        "entry_day",
        "exit_day",
    }

    numeric_4_keywords = [
        "signal",
        "sharpe",
        "nav",
    ]

    for ws in wb.worksheets:
        if ws.max_row < 1 or ws.max_column < 1:
            continue

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        header_to_col = {
            str(ws.cell(row=1, column=col_idx).value): col_idx
            for col_idx in range(1, ws.max_column + 1)
        }

        for header, col_idx in header_to_col.items():
            h_lower = header.lower()

            if header in integer_cols:
                fmt = "0"
            elif header in date_cols:
                fmt = "yyyy-mm-dd"
            elif any(k in h_lower for k in percent_keywords):
                fmt = "0.00%"
            elif any(k in h_lower for k in numeric_4_keywords):
                fmt = "0.0000"
            else:
                fmt = None

            if fmt is not None:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = fmt

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            scan_rows = min(ws.max_row, 3000)

            for row_idx in range(1, scan_rows + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))

            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 38)

    wb.save(excel_path)


# =============================================================================
# 9. Run
# =============================================================================

def run_decile_q1_q10_limit_failure_test():
    signal_panel = get_signal_panel_for_decile_test()
    signal_col, signal_mode_used = resolve_decile_signal_col(signal_panel)

    print(f"\nDecile signal mode: {signal_mode_used}")
    print(f"Decile signal column: {signal_col}")

    # -------------------------------------------------------------------------
    # Q1-Q10 clean backtest
    # -------------------------------------------------------------------------
    decile_daily_returns, decile_rebalance_info, panel_for_backtest = build_decile_group_returns(
        signal_panel=signal_panel,
        signal_col=signal_col,
        rebalance_periods=DECILE_REBALANCE_PERIODS,
    )

    decile_nav, decile_performance_summary = build_decile_nav_and_summary(
        daily_returns=decile_daily_returns,
    )

    # -------------------------------------------------------------------------
    # Limit-up / limit-down failure diagnostics
    # -------------------------------------------------------------------------
    signal_panel_with_limit = add_limit_status_to_decile_signal_panel(signal_panel)

    failure_by_rebalance = calculate_decile_limit_failure_stats(
        signal_panel_with_limit=signal_panel_with_limit,
        signal_col=signal_col,
        rebalance_periods=DECILE_REBALANCE_PERIODS,
    )

    failure_summary = build_decile_limit_failure_summary(
        failure_by_rebalance=failure_by_rebalance,
        exclude_initial=True,
    )

    failure_summary_all = build_decile_limit_failure_summary(
        failure_by_rebalance=failure_by_rebalance,
        exclude_initial=False,
    )

    extreme_layer_check = build_decile_extreme_layer_check(
        failure_summary=failure_summary,
        performance_summary=decile_performance_summary,
    )

    # -------------------------------------------------------------------------
    # Save CSV
    # -------------------------------------------------------------------------
    decile_performance_summary.to_csv(
        DECILE_OUTDIR / "decile_performance_summary.csv",
        index=False,
        encoding="utf-8-sig",
    )

    decile_nav.to_csv(
        DECILE_OUTDIR / "decile_cumulative_nav.csv",
        index=False,
        encoding="utf-8-sig",
    )

    decile_daily_returns.to_csv(
        DECILE_OUTDIR / "decile_daily_group_returns.csv",
        index=False,
        encoding="utf-8-sig",
    )

    decile_rebalance_info.to_csv(
        DECILE_OUTDIR / "decile_rebalance_info.csv",
        index=False,
        encoding="utf-8-sig",
    )

    failure_by_rebalance.to_csv(
        DECILE_OUTDIR / "decile_limit_failure_by_rebalance.csv",
        index=False,
        encoding="utf-8-sig",
    )

    failure_summary.to_csv(
        DECILE_OUTDIR / "decile_limit_failure_summary.csv",
        index=False,
        encoding="utf-8-sig",
    )

    failure_summary_all.to_csv(
        DECILE_OUTDIR / "decile_limit_failure_summary_all.csv",
        index=False,
        encoding="utf-8-sig",
    )

    extreme_layer_check.to_csv(
        DECILE_OUTDIR / "decile_extreme_layer_check.csv",
        index=False,
        encoding="utf-8-sig",
    )

    # -------------------------------------------------------------------------
    # Save Excel and plots
    # -------------------------------------------------------------------------
    save_decile_outputs(
        decile_performance_summary=decile_performance_summary,
        decile_nav=decile_nav,
        decile_daily_returns=decile_daily_returns,
        decile_rebalance_info=decile_rebalance_info,
        failure_by_rebalance=failure_by_rebalance,
        failure_summary=failure_summary,
        failure_summary_all=failure_summary_all,
        extreme_layer_check=extreme_layer_check,
        signal_col=signal_col,
        signal_mode_used=signal_mode_used,
    )

    plot_decile_nav_curves(
        nav_df=decile_nav,
        output_dir=DECILE_OUTDIR,
    )

    plot_decile_limit_failure_bars(
        failure_summary=failure_summary,
        output_dir=DECILE_OUTDIR,
    )

    # -------------------------------------------------------------------------
    # Console output
    # -------------------------------------------------------------------------
    print("\nDecile performance summary:")
    show_perf_cols = [
        "horizon",
        "portfolio",
        "annual_return",
        "sharpe_no_rf",
        "max_drawdown",
        "avg_daily_return",
        "daily_win_rate",
    ]
    print(decile_performance_summary[[c for c in show_perf_cols if c in decile_performance_summary.columns]].to_string(index=False))

    print("\nLimit failure summary:")
    show_failure_cols = [
        "horizon",
        "group",
        "rebalance_count",
        "avg_target_limit_down_ratio",
        "total_failed_sell_turnover_ratio",
        "avg_target_limit_up_ratio",
        "total_failed_buy_turnover_ratio",
        "total_failed_turnover_ratio",
        "avg_actual_gross_exposure",
    ]
    print(failure_summary[[c for c in show_failure_cols if c in failure_summary.columns]].to_string(index=False))

    print("\nExtreme layer check:")
    print(extreme_layer_check.to_string(index=False))

    return {
        "decile_daily_returns": decile_daily_returns,
        "decile_rebalance_info": decile_rebalance_info,
        "decile_nav": decile_nav,
        "decile_performance_summary": decile_performance_summary,
        "failure_by_rebalance": failure_by_rebalance,
        "failure_summary": failure_summary,
        "failure_summary_all": failure_summary_all,
        "extreme_layer_check": extreme_layer_check,
        "signal_panel_with_limit": signal_panel_with_limit,
        "signal_col": signal_col,
        "signal_mode_used": signal_mode_used,
    }


if __name__ == "__main__":
    decile_limit_failure_result = run_decile_q1_q10_limit_failure_test()